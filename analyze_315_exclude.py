import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# 读取预测数据
df = pd.read_excel('d:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx')

def to_float(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

df['主胜概率'] = df['主胜概率'].apply(to_float)
df['平局概率'] = df['平局概率'].apply(to_float)
df['客胜概率'] = df['客胜概率'].apply(to_float)

actual_results = {
    '周日001': ('日本女', 1), '周日002': ('福冈黄蜂', 1), '周日003': ('首尔FC', 0),
    '周日004': ('仁川联', 2), '周日005': ('墨胜利', 1), '周日006': ('乌德勒支', 0),
    '周日007': ('热那亚', 0), '周日008': ('汉诺威96', 2), '周日009': ('马洛卡', 1),
    '周日010': ('费耶诺德', 1), '周日011': ('布兰', 0), '周日012': ('水晶宫', 2),
    '周日013': ('诺丁汉', 2), '周日014': ('曼联', 1), '周日015': ('比萨', 1),
    '周日016': ('博洛尼亚', 0), '周日017': ('美因茨', 0), '周日018': ('巴萨', 1),
    '周日019': ('瓦勒伦加', 1), '周日020': ('里昂', 2), '周日021': ('利物浦', 2),
    '周日022': ('柏林联合', 0), '周日023': ('科莫', 1), '周日024': ('贝蒂斯', 2),
    '周日025': ('斯图加特', 1), '周日026': ('拉齐奥', 1), '周日027': ('皇家社会', 1),
    '周日028': ('波尔图', 1), '周日029': ('温哥华', 1),
}

results = []
for idx, row in df.iterrows():
    match_id = row['编号'].split('_')[0]
    
    first_choice = str(row['首选']).strip()
    if '主胜' in first_choice:
        pred_code, pred_str = 0, '主胜'
    elif '客胜' in first_choice:
        pred_code, pred_str = 1, '客胜'
    else:
        pred_code, pred_str = 2, '平局'
    
    probs = {'主胜': row['主胜概率'], '平局': row['平局概率'], '客胜': row['客胜概率']}
    max_prob_str = max(probs, key=probs.get)
    max_prob = probs[max_prob_str]
    
    panxing = row['盘型']
    
    actual = actual_results.get(match_id)
    if not actual:
        continue
    actual_code = actual[1]
    actual_str = ['主胜', '客胜', '平局'][actual_code]
    
    is_correct_orig = pred_code == actual_code
    confidence = max_prob
    
    # 排除法：找到除原预测外概率最高的，然后选最低的
    other_options = {k: v for k, v in probs.items() if k != pred_str}
    second_best = max(other_options, key=other_options.get)
    exclude_options = [pred_str, second_best]
    final_pick = [k for k in probs.keys() if k not in exclude_options][0]
    
    is_correct_exclude = final_pick == actual_str
    
    results.append({
        '编号': match_id,
        '盘型': panxing,
        '置信度': f"{max_prob:.1f}%",
        '原预测': pred_str,
        '排除': f"{pred_str},{second_best}",
        '最终推荐': final_pick,
        '实际': actual_str,
        '原对错': 'O' if is_correct_orig else 'X',
        '排除对错': 'O' if is_correct_exclude else 'X'
    })

result_df = pd.DataFrame(results)

print("=" * 100)
print("排除法验证 (实盘+低置信度)")
print("=" * 100)

# 只看实盘+低置信度
low_conf_real = result_df[(result_df['盘型'] == '实盘') & (result_df['置信度'].apply(lambda x: float(x.replace('%','')) <= 50))]
print(f"\n实盘+低置信度 共{len(low_conf_real)}场:")
print(low_conf_real[['编号', '置信度', '原预测', '排除', '最终推荐', '实际', '原对错', '排除对错']].to_string(index=False))

orig_c = (low_conf_real['原对错'] == 'O').sum()
exclude_c = (low_conf_real['排除对错'] == 'O').sum()
print(f"\n原预测正确: {orig_c}/{len(low_conf_real)} = {orig_c/len(low_conf_real)*100:.1f}%")
print(f"排除法正确: {exclude_c}/{len(low_conf_real)} = {exclude_c/len(low_conf_real)*100:.1f}%")

# 完整表格
print("\n" + "=" * 100)
print("全部29场验证")
print("=" * 100)
orig_correct = (result_df['原对错'] == 'O').sum()
exclude_correct = (result_df['排除对错'] == 'O').sum()
print(f"\n原单选准确率: {orig_correct}/29 = {orig_correct/29*100:.1f}%")
print(f"排除法准确率: {exclude_correct}/29 = {exclude_correct/29*100:.1f}%")

print(result_df[['编号', '盘型', '置信度', '原预测', '排除', '最终推荐', '实际', '原对错', '排除对错']].to_string(index=False))

# 保存
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "排除法"

headers = ['编号', '盘型', '置信度', '原预测', '排除', '最终推荐', '实际', '原对错', '排除对错']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = PatternFill('solid', start_color='366092')
    cell.font = Font(bold=True, color='FFFFFF')

for row_idx, r in enumerate(results, 2):
    for col_idx, key in enumerate(['编号', '盘型', '置信度', '原预测', '排除', '最终推荐', '实际', '原对错', '排除对错'], 1):
        ws.cell(row_idx, col_idx, r[key])
    
    if r['排除对错'] == 'O':
        ws.cell(row_idx, 7).fill = PatternFill('solid', start_color='00FF00')
    else:
        ws.cell(row_idx, 7).fill = PatternFill('solid', start_color='FF6666')

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 10

wb.save('3.15_排除法验证.xlsx')
print("\n已保存: 3.15_排除法验证.xlsx")
