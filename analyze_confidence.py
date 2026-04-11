# 比赛把握度分析 + 重点推荐
import os
import re
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def analyze_with_confidence(info):
    """带把握度的分析"""
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "把握度": "低", "详情": []}
    
    initial_home = [x[0] for x in initial]
    initial_draw = [x[1] for x in initial]
    initial_away = [x[2] for x in initial]
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_realtime_home = sum(realtime_home) / len(realtime_home)
    avg_realtime_draw = sum(realtime_draw) / len(realtime_draw)
    avg_realtime_away = sum(realtime_away) / len(realtime_away)
    
    prob_home = 1 / avg_realtime_home / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_draw = 1 / avg_realtime_draw / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_away = 1 / avg_realtime_away / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    
    home_pct_change = (avg_realtime_home - sum(initial_home)/len(initial_home)) / (sum(initial_home)/len(initial_home)) * 100
    away_pct_change = (avg_realtime_away - sum(initial_away)/len(initial_away)) / (sum(initial_away)/len(initial_away)) * 100
    
    draw_up = sum(1 for i in range(len(initial)) if realtime[i][1] > initial[i][1])
    draw_down_pct = (len(initial) - draw_up) / len(initial) * 100
    
    details = []
    prediction = ""
    confidence = "低"  # 默认低把握
    
    # ===== 高把握比赛 =====
    
    # 把握度A: 强胆客 (<1.5) - 历史100%准确
    if avg_realtime_away < 1.5:
        prediction = f"{away}客胜"
        details.append("强胆客")
        confidence = "A"
    
    # 把握度A: 强胆主 (<1.5) - 历史57%但赔率极端
    elif avg_realtime_home < 1.5:
        prediction = f"{home}主胜"
        details.append("强胆主")
        confidence = "A"
    
    # 把握度B: 强队主场 + 近况好 + 概率高
    elif 1.5 < avg_realtime_home < 2.0 and home_w >= home_l and prob_home > 0.45:
        prediction = f"{home}主胜"
        details.append(f"强主场,概率{prob_home*100:.0f}%")
        confidence = "B"
    
    # 把握度B: 强队客场 + 近况好 + 概率高
    elif 1.5 < avg_realtime_away < 2.0 and away_w >= away_l and prob_away > 0.45:
        prediction = f"{away}客胜"
        details.append(f"强客场,概率{prob_away*100:.0f}%")
        confidence = "B"
    
    # 把握度B: 主胜概率极高 (>55%)
    elif prob_home > 0.55:
        prediction = f"{home}主胜"
        details.append(f"主胜概率{prob_home*100:.0f}%")
        confidence = "B"
    
    # 把握度B: 客胜概率极高 (>55%)
    elif prob_away > 0.55:
        prediction = f"{away}客胜"
        details.append(f"客胜概率{prob_away*100:.0f}%")
        confidence = "B"
    
    # ===== 中把握比赛 =====
    
    # 把握度C: 概率偏向 + 变化支持
    elif prob_home > prob_away and prob_home > 0.35 and home_pct_change < 5:
        prediction = f"{home}主胜"
        details.append(f"主胜概率{prob_home*100:.0f}%")
        confidence = "C"
    
    elif prob_away > prob_home and prob_away > 0.35 and away_pct_change < 5:
        prediction = f"{away}客胜"
        details.append(f"客胜概率{prob_away*100:.0f}%")
        confidence = "C"
    
    # ===== 低把握比赛 =====
    
    # 强客场 - 历史0%准确
    elif 1.5 < avg_realtime_away < 2.0:
        if prob_draw > 0.30:
            prediction = "平局"
            details.append("强客场防冷")
        else:
            prediction = f"{home}主胜"
            details.append("强客场防冷")
        confidence = "D"
    
    # 其他情况默认概率最高
    else:
        if prob_home > prob_away and prob_home > prob_draw:
            prediction = f"{home}主胜"
            details.append("主胜概率最高")
        elif prob_away > prob_home and prob_away > prob_draw:
            prediction = f"{away}客胜"
            details.append("客胜概率最高")
        else:
            prediction = "平局"
            details.append("平局概率最高")
        confidence = "D"
    
    panxing = "实盘" if draw_down_pct > 50 else "诱盘"
    
    return {
        "预测": prediction,
        "把握度": confidence,
        "盘型": panxing,
        "详情": details,
        "主胜概率": f"{prob_home*100:.0f}%",
        "平局概率": f"{prob_draw*100:.0f}%",
        "客胜概率": f"{prob_away*100:.0f}%"
    }

def parse_source_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    match_info = {}
    
    filename = Path(filepath).stem
    match = re.match(r'周六(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        match_info['match_id'] = f"周六{match.group(1)}"
        match_info['home_team'] = match.group(2)
        match_info['away_team'] = match.group(3)
    
    match = re.search(r'主队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['home_w'] = int(match.group(1))
        match_info['home_l'] = int(match.group(3))
    
    match = re.search(r'客队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['away_w'] = int(match.group(1))
        match_info['away_l'] = int(match.group(3))
    
    initial_odds = []
    realtime_odds = []
    
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            initial_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            realtime_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match_info['initial_odds'] = initial_odds
    match_info['realtime_odds'] = realtime_odds
    
    return match_info

# 实际结果
actual_14 = {
    "周六001": "平局", "周六002": "客胜", "周六003": "平局", "周六004": "客胜",
    "周六005": "主胜", "周六006": "主胜", "周六007": "平局", "周六008": "主胜",
    "周六009": "主胜", "周六010": "客胜", "周六011": "主胜", "周六012": "平局",
    "周六013": "平局", "周六014": "主胜", "周六015": "主胜", "周六016": "平局",
    "周六017": "平局", "周六018": "主胜", "周六019": "主胜", "周六020": "主胜",
    "周六021": "主胜", "周六022": "主胜", "周六023": "客胜", "周六024": "平局",
    "周六025": "主胜", "周六026": "客胜", "周六027": "客胜", "周六028": "客胜",
    "周六029": "平局", "周六030": "主胜", "周六031": "客胜", "周六032": "平局",
}

def check_match(pred, actual):
    pred = str(pred).lower()
    if "主胜" in pred:
        return "主胜" == actual
    elif "客胜" in pred:
        return "客胜" == actual
    elif "平局" in pred:
        return "平局" == actual
    return False

# 分析3.14
folder = "分析模板/3.14"
files = sorted(Path(folder).glob("*.md"))
results = []

for f in files:
    info = parse_source_file(f)
    if not info.get('home_team') or not info.get('initial_odds'):
        continue
    result = analyze_with_confidence(info)
    results.append({
        '编号': info.get('match_id'),
        '对阵': f"{info.get('home_team')} vs {info.get('away_team')}",
        '预测': result['预测'],
        '把握度': result['把握度'],
        '盘型': result['盘型'],
        '主胜概率': result['主胜概率'],
        '平局概率': result['平局概率'],
        '客胜概率': result['客胜概率'],
        '详情': ', '.join(result['详情'])
    })

# 按把握度分组统计
conf_a = [r for r in results if r['把握度'] == 'A']
conf_b = [r for r in results if r['把握度'] == 'B']
conf_c = [r for r in results if r['把握度'] == 'C']
conf_d = [r for r in results if r['把握度'] == 'D']

print("=" * 60)
print("3.14 比赛把握度分析")
print("=" * 60)

# 统计各把握度准确率
def count_confidence(results_list, actual_dict):
    correct = 0
    total = len(results_list)
    for r in results_list:
        actual = actual_dict.get(r['编号'], "")
        if check_match(r['预测'], actual):
            correct += 1
    return correct, total

print("\n【把握度A - 高把握】(强胆赔率)")
if conf_a:
    c, t = count_confidence(conf_a, actual_14)
    print(f"  共{t}场, 准确{c}场, 准确率{c/t*100:.0f}%")
    for r in conf_a:
        actual = actual_14.get(r['编号'], "")
        ok = "OK" if check_match(r['预测'], actual) else "X"
        print(f"    {ok} {r['编号']}: {r['预测']} vs {actual}")

print("\n【把握度B - 中高把握】(强队+高概率)")
if conf_b:
    c, t = count_confidence(conf_b, actual_14)
    print(f"  共{t}场, 准确{c}场, 准确率{c/t*100:.0f}%")
    for r in conf_b:
        actual = actual_14.get(r['编号'], "")
        ok = "OK" if check_match(r['预测'], actual) else "X"
        print(f"    {ok} {r['编号']}: {r['预测']} vs {actual}")

print("\n【把握度C - 中把握】(概率偏向)")
if conf_c:
    c, t = count_confidence(conf_c, actual_14)
    print(f"  共{t}场, 准确{c}场, 准确率{c/t*100:.0f}%")
    for r in conf_c:
        actual = actual_14.get(r['编号'], "")
        ok = "OK" if check_match(r['预测'], actual) else "X"
        print(f"    {ok} {r['编号']}: {r['预测']} vs {actual}")

print("\n【把握度D - 低把握】(其他)")
if conf_d:
    c, t = count_confidence(conf_d, actual_14)
    print(f"  共{t}场, 准确{c}场, 准确率{c/t*100:.0f}%")

# 总准确率
all_correct = sum(1 for r in results if check_match(r['预测'], actual_14.get(r['编号'], "")))
print(f"\n总体准确率: {all_correct}/32 = {all_correct/32*100:.1f}%")

# 重点推荐 (A+B)
print("\n" + "=" * 60)
print("【重点推荐】把握度A+B")
print("=" * 60)
high_conf = conf_a + conf_b
high_correct = sum(1 for r in high_conf if check_match(r['预测'], actual_14.get(r['编号'], "")))
print(f"共{len(high_conf)}场, 准确{high_correct}场, 准确率{high_correct/len(high_conf)*100:.0f}%")

# 保存Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.14_把握度分析"

headers = ['编号', '对阵', '预测', '把握度', '盘型', '主胜概率', '平局概率', '客胜概率', '详情']
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal='center')

for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['预测'])
    # 把握度颜色
    ws.cell(row, 4, r['把握度'])
    ws.cell(row, 5, r['盘型'])
    ws.cell(row, 6, r['主胜概率'])
    ws.cell(row, 7, r['平局概率'])
    ws.cell(row, 8, r['客胜概率'])
    ws.cell(row, 9, r['详情'])
    
    # 把握度颜色
    conf_cell = ws.cell(row, 4)
    if r['把握度'] == 'A':
        conf_cell.fill = PatternFill("solid", fgColor="00FF00")  # 绿色
    elif r['把握度'] == 'B':
        conf_cell.fill = PatternFill("solid", fgColor="92D050")  # 浅绿
    elif r['把握度'] == 'C':
        conf_cell.fill = PatternFill("solid", fgColor="FFFF00")  # 黄色
    else:
        conf_cell.fill = PatternFill("solid", fgColor="FF0000")  # 红色

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 30

wb.save("3.14_把握度分析.xlsx")
print("\n已保存: 3.14_把握度分析.xlsx")
