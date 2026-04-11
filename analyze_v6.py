# V6算法 - 尝试不同策略
# 策略：识别冷门模式 + 增强平局

import os
import re
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def analyze_v6(info):
    """V6算法 - 冷门识别 + 平局增强"""
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    macao_tip = info.get('macao_tip', '')
    
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "盘型": "未知", "详情": []}
    
    initial_home = [x[0] for x in initial]
    initial_draw = [x[1] for x in initial]
    initial_away = [x[2] for x in initial]
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_initial_home = sum(initial_home) / len(initial_home)
    avg_initial_draw = sum(initial_draw) / len(initial_draw)
    avg_initial_away = sum(initial_away) / len(initial_away)
    
    avg_realtime_home = sum(realtime_home) / len(realtime_home)
    avg_realtime_draw = sum(realtime_draw) / len(realtime_draw)
    avg_realtime_away = sum(realtime_away) / len(realtime_away)
    
    prob_home = 1 / avg_realtime_home / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_draw = 1 / avg_realtime_draw / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_away = 1 / avg_realtime_away / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    
    home_pct_change = (avg_realtime_home - avg_initial_home) / avg_initial_home * 100
    draw_pct_change = (avg_realtime_draw - avg_initial_draw) / avg_initial_draw * 100
    away_pct_change = (avg_realtime_away - avg_initial_away) / avg_initial_away * 100
    
    home_up = sum(1 for i in range(len(initial)) if realtime[i][0] > initial[i][0])
    draw_up = sum(1 for i in range(len(initial)) if realtime[i][1] > initial[i][1])
    away_up = sum(1 for i in range(len(initial)) if realtime[i][2] > initial[i][2])
    
    home_up_pct = home_up / len(initial) * 100
    away_up_pct = away_up / len(initial) * 100
    draw_down = len(initial) - draw_up
    draw_down_pct = draw_down / len(initial) * 100
    
    macao_home = realtime[0][0] if realtime else avg_realtime_home
    macao_away = realtime[0][2] if realtime else avg_realtime_away
    
    details = []
    prediction = ""
    
    # ===== V6新策略：冷门识别 =====
    # 冷门模式1: 强队近况好但赔率升 -> 反向
    if home_w >= 3 and home_pct_change > 5:
        prediction = "平局" if prob_draw > 0.28 else f"{away}客胜"
        details.append("冷门: 主队强但升赔")
    elif away_w >= 3 and away_pct_change > 5:
        prediction = "平局" if prob_draw > 0.28 else f"{home}主胜"
        details.append("冷门: 客队强但升赔")
    
    # 冷门模式2: 主客胜同时升赔 -> 平局
    elif home_pct_change > 0 and away_pct_change > 0 and prob_draw > 0.28:
        prediction = "平局"
        details.append("模式: 双升")
    
    # 冷门模式3: 强队低赔但升赔 -> 可能不胜
    elif avg_realtime_home < 1.5 and home_pct_change > 3:
        prediction = "平局" if prob_draw > 0.30 else f"{away}客胜"
        details.append("冷门: 强队低赔升")
    elif avg_realtime_away < 1.5 and away_pct_change > 3:
        prediction = "平局" if prob_draw > 0.30 else f"{home}主胜"
        details.append("冷门: 强队低赔升")
    
    # 规则: 强胆
    elif avg_realtime_home < 1.5:
        prediction = f"{home}主胜"
        details.append("强胆")
    elif avg_realtime_away < 1.5:
        prediction = f"{away}客胜"
        details.append("强胆")
    
    # 规则: 强队 + 近况好
    elif 1.5 < avg_realtime_home < 2.0 and home_w > home_l:
        prediction = f"{home}主胜"
        details.append("强队主场")
    elif 1.5 < avg_realtime_away < 2.0 and away_w > away_l:
        prediction = f"{away}客胜"
        details.append("强队客场")
    
    # 规则: 大幅变化
    elif home_pct_change > 15 and home_up_pct > 80:
        prediction = f"{away}客胜"
        details.append(f"主胜升{home_pct_change:.0f}%")
    elif away_pct_change > 15 and away_up_pct > 70:
        prediction = f"{home}主胜"
        details.append(f"客胜升{away_pct_change:.0f}%")
    
    # 规则: 近况好
    elif home_w >= 3 and avg_realtime_home < 2.3:
        prediction = f"{home}主胜"
        details.append(f"主队W{home_w}")
    elif away_w >= 3 and avg_realtime_away < 2.3:
        prediction = f"{away}客胜"
        details.append(f"客队W{away_w}")
    
    # 规则: 平局防范增强
    elif draw_down_pct > 55 and draw_pct_change < -2 and prob_draw > 0.28:
        prediction = "平局"
        details.append(f"平局降{abs(draw_pct_change):.1f}%")
    elif prob_draw > 0.32 and home_pct_change > 3:
        prediction = "平局"
        details.append("高平局概率+主胜不稳")
    
    # 规则: 澳门
    elif macao_away < avg_realtime_away * 0.85:
        prediction = f"{away}客胜"
        details.append("澳门客胜低")
    elif macao_home < avg_realtime_home * 0.90:
        prediction = f"{home}主胜"
        details.append("澳门主胜低")
    elif macao_tip and home in macao_tip:
        prediction = f"{home}主胜"
        details.append("澳门推荐")
    elif macao_tip and away in macao_tip:
        prediction = f"{away}客胜"
        details.append("澳门推荐")
    
    # 规则: 主胜不稳
    elif home_pct_change > 5 and away_pct_change < 0:
        prediction = f"{away}客胜"
        details.append("主胜不稳")
    
    # 默认
    else:
        if prob_home > prob_away and prob_home > prob_draw:
            prediction = f"{home}主胜"
            details.append("主胜概率最高")
        elif prob_away > prob_home and prob_away > prob_draw:
            prediction = f"{away}客胜"
            details.append("客胜概率最高")
        else:
            prediction = "平局"
            details.append("平局概率最高")
    
    panxing = "实盘" if home_up_pct > 80 and draw_down_pct > 50 else "诱盘"
    
    return {
        "预测": prediction,
        "盘型": panxing,
        "详情": details,
    }

def parse_source_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    match_info = {}
    
    filename = Path(filepath).stem
    match = re.match(r'周六(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        match_info['match_id'] = f"周六{match.group(1)}"
        match_info['home_team'] = match.group(2)
        match_info['away_team'] = match.group(3)
    
    match = re.search(r'主队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['home_w'] = int(match.group(1))
        match_info['home_l'] = int(match.group(3))
    
    match = re.search(r'客队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['away_w'] = int(match.group(1))
        match_info['away_l'] = int(match.group(3))
    
    match = re.search(r'澳门推荐.*?\|.*?(\S+)', content)
    if match:
        match_info['macao_tip'] = match.group(1)
    
    initial_odds = []
    realtime_odds = []
    
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            initial_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            realtime_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match_info['initial_odds'] = initial_odds
    match_info['realtime_odds'] = realtime_odds
    
    return match_info

# 测试V6
folder = "分析模板/3.14"
files = sorted(Path(folder).glob("*.md"))
results = []

for f in files:
    info = parse_source_file(f)
    if not info.get('home_team') or not info.get('initial_odds'):
        continue
    result = analyze_v6(info)
    results.append({
        '编号': info.get('match_id'),
        '对阵': f"{info.get('home_team')} vs {info.get('away_team')}",
        '预测': result['预测'],
        '详情': ', '.join(result['详情'])
    })

actual_14 = {
    "周六001": "平局", "周六002": "客胜", "周六003": "平局", "周六004": "客胜",
    "周六005": "主胜", "周六006": "主胜", "周六007": "平局", "周六008": "主胜",
    "周六009": "主胜", "周六010": "客胜", "周六011": "主胜", "周六012": "平局",
    "周六013": "平局", "周六014": "主胜", "周六015": "主胜", "周六016": "平局",
    "周六017": "平局", "周六018": "主胜", "周六019": "主胜", "周六020": "主胜",
    "周六021": "主胜", "周六022": "主胜", "周六023": "客胜", "周六024": "平局",
    "周六025": "主胜", "周六026": "客胜", "周六027": "客胜", "周六028": "客胜",
    "周六029": "平局", "周六030": "主胜", "周六031": "客胜", "周六032": "平局",
}

def check_match(pred, actual):
    pred = str(pred).lower()
    if "主胜" in pred:
        return "主胜" == actual
    elif "客胜" in pred:
        return "客胜" == actual
    elif "平局" in pred:
        return "平局" == actual
    return False

correct14 = 0
draw_preds = 0
for r in results:
    actual = actual_14.get(r['编号'], "")
    pred = r['预测']
    if "平局" in pred:
        draw_preds += 1
    if check_match(pred, actual):
        correct14 += 1

print(f"V6算法 - 3.14准确率: {correct14/32*100:.1f}% ({correct14}/32)")
print(f"预测平局数: {draw_preds}/32")
print(f"实际平局数: 9/32 (28%)")

# 保存结果
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.14_V6预测"

for col, header in enumerate(['编号', '对阵', '预测', '详情'], 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")

for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['预测'])
    ws.cell(row, 4, r['详情'])

wb.save("3.14_V6预测.xlsx")
print("\n已保存: 3.14_V6预测.xlsx")
