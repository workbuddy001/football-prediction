# 3.14比赛复盘分析
import openpyxl

# 实际比赛结果
actual_results = {
    "周六001": ("中国女", "中国台女", "平局"),
    "周六002": ("名古屋鲸", "神户胜利", "客胜"),
    "周六003": ("光州FC", "全北现代", "平局"),
    "周六004": ("纽喷气机", "奥克兰FC", "客胜"),
    "周六005": ("鹿岛鹿角", "川崎前锋", "主胜"),
    "周六006": ("东京绿茵", "浦和红钻", "主胜"),
    "周六007": ("大田市民", "金泉尚武", "平局"),
    "周六008": ("韩国女", "乌兹别克斯坦女主胜", "主胜"),
    "周六009": ("不伦瑞克", "杜塞多夫", "主胜"),
    "周六010": ("考文垂", "南安普敦", "客胜"),
    "周六011": ("赫罗纳", "毕尔巴鄂", "主胜"),
    "周六012": ("国际米兰", "亚特兰大", "平局"),
    "周六013": ("霍芬海姆", "沃夫斯堡", "平局"),
    "周六014": ("多特蒙德", "奥格斯堡", "主胜"),
    "周六015": ("法兰克福", "海登海姆", "主胜"),
    "周六016": ("勒沃库森", "拜仁", "平局"),
    "周六017": ("伯恩利", "伯恩茅斯", "平局"),
    "周六018": ("马竞", "赫塔费", "主胜"),
    "周六019": ("洛里昂", "朗斯", "主胜"),
    "周六020": ("那不勒斯", "莱切", "主胜"),
    "周六021": ("莫尔德", "罗森博格", "主胜"),
    "周六022": ("阿森纳", "埃弗顿", "主胜"),
    "周六023": ("切尔西", "纽卡斯尔", "客胜"),
    "周六024": ("汉堡", "科隆", "平局"),
    "周六025": ("奥维耶多", "巴伦西亚", "主胜"),
    "周六026": ("埃因霍温", "奈梅亨", "客胜"),
    "周六027": ("赛哈特海湾", "利雅得胜利", "客胜"),
    "周六028": ("乌迪内斯", "尤文图斯", "客胜"),
    "周六029": ("西汉姆联", "曼城", "平局"),
    "周六030": ("皇马", "埃尔切", "主胜"),
    "周六031": ("阿罗卡", "本菲卡", "客胜"),
    "周六032": ("达拉斯", "圣迭戈FC", "平局"),
}

# 读取预测结果
wb = openpyxl.load_workbook('3.14_比赛预测汇总.xlsx')
ws = wb.active

predictions = {}
for row in range(2, 34):
    id_ = ws.cell(row, 1).value
    pred = ws.cell(row, 4).value
    predictions[id_] = pred

# 检查预测是否正确
def check_match(pred, actual):
    pred = str(pred).lower()
    actual = str(actual).lower()
    
    if "主胜" in pred:
        return "主胜" in actual
    elif "客胜" in pred:
        return "客胜" in actual
    elif "平局" in pred:
        return "平局" in actual
    return False

# 计算准确率
correct = 0
wrong = 0
details = []

for id_ in sorted(actual_results.keys()):
    actual = actual_results[id_][2]
    pred = predictions.get(id_, "")
    is_correct = check_match(pred, actual)
    
    if is_correct:
        correct += 1
        result = "OK"
    else:
        wrong += 1
        result = "X"
    
    details.append(f"{result} {id_}: pred={pred}, actual={actual}")

accuracy = correct / (correct + wrong) * 100
print(f"准确率: {accuracy:.1f}% ({correct}/{correct+wrong})")
print()
for d in details:
    print(d)
