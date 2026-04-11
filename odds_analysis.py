import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 初盘赔率数据
initial_odds = [
    (2.88, 2.80, 2.35), (3.00, 2.75, 2.50), (3.13, 2.91, 2.23), (3.10, 2.87, 2.50),
    (3.20, 3.00, 2.35), (3.15, 3.05, 2.45), (3.15, 3.05, 2.35), (3.15, 3.25, 2.41),
    (3.40, 3.10, 2.38), (3.10, 2.90, 2.45), (3.10, 2.87, 2.55), (3.02, 2.95, 2.57),
    (3.20, 3.00, 2.35), (3.10, 2.90, 2.50), (2.99, 2.92, 2.35), (3.15, 2.90, 2.38),
    (3.10, 2.95, 2.50), (3.15, 3.05, 2.32), (3.15, 3.05, 2.32), (3.15, 3.25, 2.41),
    (3.15, 2.80, 2.18), (3.15, 3.05, 2.32), (3.25, 2.98, 2.44), (2.80, 3.05, 2.60),
    (3.30, 3.09, 2.58), (3.15, 3.05, 2.35), (3.00, 3.10, 2.35), (3.10, 2.80, 2.30),
    (3.15, 3.05, 2.35), (3.15, 2.98, 2.52), (2.88, 3.00, 2.50), (3.10, 3.00, 2.30),
    (3.37, 3.00, 2.50), (3.20, 3.05, 2.41), (3.25, 3.15, 2.40), (3.15, 3.10, 2.45),
    (3.20, 2.90, 2.30), (3.18, 3.05, 2.52), (3.35, 3.10, 2.33), (3.15, 3.00, 2.40),
    (3.00, 2.88, 2.50), (2.85, 2.70, 2.40), (3.20, 3.10, 2.35), (2.80, 2.80, 2.70),
    (3.15, 3.00, 2.38), (3.20, 3.20, 2.30), (3.30, 3.04, 2.42), (3.20, 3.00, 2.35),
    (3.26, 2.98, 2.35), (3.15, 2.95, 2.45), (3.10, 2.70, 2.50)
]

realtime_odds = [
    (3.30, 2.69, 2.20), (3.30, 2.80, 2.30), (3.13, 2.91, 2.23), (3.30, 2.87, 2.40),
    (3.30, 3.00, 2.37), (3.35, 2.95, 2.45), (3.40, 2.90, 2.45), (3.60, 2.95, 2.38),
    (3.50, 3.00, 2.40), (3.50, 2.80, 2.38), (3.30, 2.87, 2.40), (3.56, 3.00, 2.41),
    (3.45, 2.95, 2.38), (3.40, 2.87, 2.45), (3.36, 2.90, 2.44), (3.35, 2.90, 2.40),
    (3.30, 2.87, 2.40), (3.20, 2.94, 2.51), (3.35, 2.90, 2.47), (3.50, 2.95, 2.40),
    (3.20, 2.77, 2.18), (3.35, 2.90, 2.47), (3.60, 3.05, 2.48), (3.40, 2.95, 2.40),
    (3.38, 3.08, 2.44), (3.40, 2.95, 2.40), (3.05, 3.10, 2.35), (3.40, 2.90, 2.40),
    (3.35, 2.95, 2.40), (3.25, 3.05, 2.42), (3.40, 2.90, 2.38), (3.30, 2.88, 2.30),
    (3.47, 2.95, 2.48), (3.40, 2.95, 2.46), (3.40, 2.95, 2.40), (3.15, 3.05, 2.45),
    (3.40, 2.90, 2.40), (3.44, 2.99, 2.44), (3.30, 2.90, 2.48), (3.30, 2.90, 2.40),
    (3.30, 2.80, 2.38), (3.35, 2.85, 2.35), (3.40, 2.90, 2.40), (3.40, 2.90, 2.30),
    (3.30, 2.95, 2.43), (3.50, 2.95, 2.30), (3.42, 3.01, 2.48), (3.45, 3.00, 2.46),
    (3.46, 2.91, 2.36), (3.45, 2.95, 2.35), (3.30, 2.80, 2.30)
]

init_home = [o[0] for o in initial_odds]
init_draw = [o[1] for o in initial_odds]
init_away = [o[2] for o in initial_odds]

real_home = [o[0] for o in realtime_odds]
real_draw = [o[1] for o in realtime_odds]
real_away = [o[2] for o in realtime_odds]

home_change = [real_home[i] - init_home[i] for i in range(len(initial_odds))]
draw_change = [real_draw[i] - init_draw[i] for i in range(len(initial_odds))]
away_change = [real_away[i] - init_away[i] for i in range(len(initial_odds))]

wb = Workbook()
ws = wb.active
ws.title = "赔率分析"

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
title_font = Font(bold=True, size=14)
center = Alignment(horizontal='center', vertical='center')

ws['A1'] = '维罗纳 vs 热那亚 赔率统计分析'
ws['A1'].font = title_font
ws.merge_cells('A1:I1')
ws['A1'].alignment = center

ws['A3'] = '一、赔率统计汇总'
ws['A3'].font = Font(bold=True, size=12)

headers = ['指标', '初盘主胜', '初盘平局', '初盘客胜', '即时主胜', '即时平局', '即时客胜']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

stats_data = [
    ['平均值', round(np.mean(init_home), 2), round(np.mean(init_draw), 2), round(np.mean(init_away), 2),
             round(np.mean(real_home), 2), round(np.mean(real_draw), 2), round(np.mean(real_away), 2)],
    ['中位数', round(np.median(init_home), 2), round(np.median(init_draw), 2), round(np.median(init_away), 2),
             round(np.median(real_home), 2), round(np.median(real_draw), 2), round(np.median(real_away), 2)],
    ['最大值', max(init_home), max(init_draw), max(init_away),
             max(real_home), max(real_draw), max(real_away)],
    ['最小值', min(init_home), min(init_draw), min(init_away),
             min(real_home), min(real_draw), min(real_away)],
    ['标准差', round(np.std(init_home), 2), round(np.std(init_draw), 2), round(np.std(init_away), 2),
             round(np.std(real_home), 2), round(np.std(real_draw), 2), round(np.std(real_away), 2)],
]

for r, row_data in enumerate(stats_data, 5):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

ws['A11'] = '二、赔率变化统计'
ws['A11'].font = Font(bold=True, size=12)

change_headers = ['变化指标', '平均变化', '中位数变化', '上升次数', '下降次数', '不变次数']
for c, h in enumerate(change_headers, 1):
    cell = ws.cell(row=12, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

change_stats = [
    ['主胜变化', round(np.mean(home_change), 2), round(np.median(home_change), 2),
             sum(1 for x in home_change if x > 0), sum(1 for x in home_change if x < 0), sum(1 for x in home_change if x == 0)],
    ['平局变化', round(np.mean(draw_change), 2), round(np.median(draw_change), 2),
             sum(1 for x in draw_change if x > 0), sum(1 for x in draw_change if x < 0), sum(1 for x in draw_change if x == 0)],
    ['客胜变化', round(np.mean(away_change), 2), round(np.median(away_change), 2),
             sum(1 for x in away_change if x > 0), sum(1 for x in away_change if x < 0), sum(1 for x in away_change if x == 0)],
]

for r, row_data in enumerate(change_stats, 13):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

ws['A17'] = '三、概率与返还率分析'
ws['A17'].font = Font(bold=True, size=12)

init_home_prob = [1/x*100 for x in init_home]
init_draw_prob = [1/x*100 for x in init_draw]
init_away_prob = [1/x*100 for x in init_away]

real_home_prob = [1/x*100 for x in real_home]
real_draw_prob = [1/x*100 for x in real_draw]
real_away_prob = [1/x*100 for x in real_away]

init_return = [(init_home_prob[i] + init_draw_prob[i] + init_away_prob[i])/100 for i in range(len(initial_odds))]
real_return = [(real_home_prob[i] + real_draw_prob[i] + real_away_prob[i])/100 for i in range(len(initial_odds))]

prob_headers = ['分析项', '主胜概率%', '平局概率%', '客胜概率%', '返还率%']
for c, h in enumerate(prob_headers, 1):
    cell = ws.cell(row=18, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

prob_data = [
    ['初盘平均值', round(np.mean(init_home_prob), 2), round(np.mean(init_draw_prob), 2), 
             round(np.mean(init_away_prob), 2), round(np.mean(init_return)*100, 2)],
    ['即时平均值', round(np.mean(real_home_prob), 2), round(np.mean(real_draw_prob), 2), 
             round(np.mean(real_away_prob), 2), round(np.mean(real_return)*100, 2)],
    ['变化', round(np.mean(real_home_prob) - np.mean(init_home_prob), 2),
             round(np.mean(real_draw_prob) - np.mean(init_draw_prob), 2),
             round(np.mean(real_away_prob) - np.mean(init_away_prob), 2),
             round((np.mean(real_return) - np.mean(init_return))*100, 2)],
]

for r, row_data in enumerate(prob_data, 19):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

ws['A23'] = '四、平局赔率专项分析'
ws['A23'].font = Font(bold=True, size=12)

draw_headers = ['分析项', '平均值', '中位数', '最低', '最高', '平均变化']
for c, h in enumerate(draw_headers, 1):
    cell = ws.cell(row=24, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

draw_data = [
    ['初盘平局', round(np.mean(init_draw), 2), round(np.median(init_draw), 2), min(init_draw), max(init_draw), '-'],
    ['即时平局', round(np.mean(real_draw), 2), round(np.median(real_draw), 2), min(real_draw), max(real_draw), round(np.mean(draw_change), 2)],
]

for r, row_data in enumerate(draw_data, 25):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

ws['A28'] = '五、详细赔率数据'
ws['A28'].font = Font(bold=True, size=12)

detail_headers = ['公司', '初盘主胜', '初盘平局', '初盘客胜', '即时主胜', '即时平局', '即时客胜', '主胜变化', '平局变化', '客胜变化']
for c, h in enumerate(detail_headers, 1):
    cell = ws.cell(row=29, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

for i in range(len(initial_odds)):
    row = 30 + i
    data = [i+1, init_home[i], init_draw[i], init_away[i], real_home[i], real_draw[i], real_away[i],
            round(home_change[i], 2), round(draw_change[i], 2), round(away_change[i], 2)]
    for c, val in enumerate(data, 1):
        ws.cell(row=row, column=c, value=val).alignment = center

for col in 'ABCDEFGHIJ':
    ws.column_dimensions[col].width = 12
ws.column_dimensions['A'].width = 8

wb.save('维罗纳vs热那亚_赔率分析.xlsx')
print("Excel文件已生成: 维罗纳vs热那亚_赔率分析.xlsx")
