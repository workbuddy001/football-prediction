# -*- coding: utf-8 -*-
"""
3.16比赛预测分析脚本
基于V7/V8分析 + 排除法策略
方法说明:
- 置信度: 从赔率计算的概率
- 胜率差: 主胜概率 - 客胜概率
- 8变化: 赔率下降的公司数量变化
- 排除法: 实盘+低置信度 → 选最低概率
"""

import re
import ast
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== 配置 ====================
DATA_DIR = Path("d:/work/workbuddy/足球预测/分析模板/3.16")
OUTPUT_FILE = "3.16_V7V8预测分析.xlsx"

# ==================== 解析源数据 ====================
def parse_source_file(filepath):
    """解析源数据文件"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    info = {}
    filename = filepath.stem
    
    # 解析编号和球队
    match = re.match(r'(周一|周二|周三|周四|周五|周六)(\d+)_([^vs]+)vs(.+?)_源数据', filename)
    if match:
        info['date'] = match.group(1)
        info['num'] = match.group(2)
        info['home_team'] = match.group(3).strip()
        info['away_team'] = match.group(4).strip()
        info['match_id'] = f"{match.group(1)}{match.group(2)}"
    
    # 解析近况走势
    match = re.search(r'主队近况走势.*?([WDL]+)', content)
    if match:
        form = match.group(1).upper()
        info['home_w'] = form.count('W')
        info['home_d'] = form.count('D')
        info['home_l'] = form.count('L')
    
    match = re.search(r'客队近况走势.*?([WDL]+)', content)
    if match:
        form = match.group(1).upper()
        info['away_w'] = form.count('W')
        info['away_d'] = form.count('D')
        info['away_l'] = form.count('L')
    
    # 解析澳门推荐
    match = re.search(r'澳门推荐.*?([^\n]+)', content)
    if match:
        info['macao_tip'] = match.group(1).strip()
    
    # 解析初盘赔率
    initial_odds = []
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            # 提取所有赔率tuple
            odds_str = '[' + match.group(1) + ']'
            # 清理注释
            odds_str = re.sub(r'#.*', '', odds_str)
            initial_odds = ast.literal_eval(odds_str)
        except Exception as e:
            print(f"解析初盘赔率失败: {e}")
    
    # 解析即时赔率
    realtime_odds = []
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            odds_str = '[' + match.group(1) + ']'
            odds_str = re.sub(r'#.*', '', odds_str)
            realtime_odds = ast.literal_eval(odds_str)
        except Exception as e:
            print(f"解析即时赔率失败: {e}")
    
    info['initial_odds'] = initial_odds
    info['realtime_odds'] = realtime_odds
    
    return info

def calculate_v7_v8(info):
    """计算V7/V8分析指标"""
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    
    if not initial or not realtime:
        return None
    
    # 计算平均赔率
    init_home = sum(x[0] for x in initial) / len(initial)
    init_draw = sum(x[1] for x in initial) / len(initial)
    init_away = sum(x[2] for x in initial) / len(initial)
    
    real_home = sum(x[0] for x in realtime) / len(realtime)
    real_draw = sum(x[1] for x in realtime) / len(realtime)
    real_away = sum(x[2] for x in realtime) / len(realtime)
    
    # 计算概率(倒数/返还率)
    prob_home = 1/init_home / (1/init_home + 1/init_draw + 1/init_away)
    prob_draw = 1/init_draw / (1/init_home + 1/init_draw + 1/init_away)
    prob_away = 1/init_away / (1/init_home + 1/init_draw + 1/init_away)
    
    # 即时概率
    real_prob_home = 1/real_home / (1/real_home + 1/real_draw + 1/real_away)
    real_prob_draw = 1/real_draw / (1/real_home + 1/real_draw + 1/real_away)
    real_prob_away = 1/real_away / (1/real_home + 1/real_draw + 1/real_away)
    
    # 置信度 = 最高概率
    confidence = max(real_prob_home, real_prob_draw, real_prob_away) * 100
    
    # 胜率差 = 主胜概率 - 客胜概率
    diff = (real_prob_home - real_prob_away) * 100
    
    # 计算8变化(赔率下降的公司数)
    home_8 = sum(1 for i in range(len(initial)) if realtime[i][0] < initial[i][0])
    draw_8 = sum(1 for i in range(len(initial)) if realtime[i][1] < initial[i][1])
    away_8 = sum(1 for i in range(len(initial)) if realtime[i][2] < initial[i][2])
    
    return {
        'confidence': confidence,
        'diff': diff,
        'home_8': home_8,
        'draw_8': draw_8,
        'away_8': away_8,
        'prob_home': real_prob_home * 100,
        'prob_draw': real_prob_draw * 100,
        'prob_away': real_prob_away * 100,
    }

def apply_prediction_strategy(v7v8, info):
    """应用V7/V8 + 排除法策略预测"""
    if not v7v8:
        return {'预测': '数据不足', '策略': '-', '理由': '数据不足'}
    
    conf = v7v8['confidence']
    diff = v7v8['diff']
    home_8 = v7v8['home_8']
    draw_8 = v7v8['draw_8']
    away_8 = v7v8['away_8']
    
    # 获取概率
    probs = {
        '主胜': v7v8['prob_home'],
        '平局': v7v8['prob_draw'],
        '客胜': v7v8['prob_away']
    }
    
    # 找到最高和最低概率
    max_option = max(probs, key=probs.get)
    min_option = min(probs, key=probs.get)
    
    # 判断盘型(基于8变化)
    # 诱盘: 某选项8增加但实际可能不胜
    # 实盘: 概率分布正常
    is_trap = False
    if home_8 >= 2 and conf > 50 and diff < 30:
        is_trap = True  # 主胜8增加但胜率差小 → 诱盘
    elif away_8 >= 2 and conf > 50 and diff > -30:
        is_trap = True  # 客胜8增加但胜率差大 → 诱盘
    
    # ==================== 策略应用 ====================
    
    # 策略1: 高置信度(>=55%) + 胜率差明显 → 正向
    if conf >= 55:
        if diff >= 30:
            pred = '主胜'
            reason = f"高置信({conf:.0f}%)+胜率差(+{diff:.0f}%)"
            strategy = "高置信正向"
        elif diff <= -30:
            pred = '客胜'
            reason = f"高置信({conf:.0f}%)+胜率差({diff:.0f}%)"
            strategy = "高置信正向"
        else:
            # 胜率差小，看8变化
            if home_8 >= 2:
                pred = '主胜'
                reason = f"主胜8变化({home_8})>=2"
                strategy = "8强化主胜"
            elif away_8 >= 2:
                pred = '客胜'
                reason = f"客胜8变化({away_8})>=2"
                strategy = "8强化客胜"
            else:
                pred = max_option
                reason = f"高置信({conf:.0f}%)默认"
                strategy = "高置信默认"
    
    # 策略2: 低置信度(<=45%) → 排除法(选最低)
    elif conf <= 45:
        pred = min_option
        reason = f"低置信度({conf:.0f}%)→排除法"
        strategy = "排除法"
    
    # 策略3: 中等置信度(45%-55%)
    else:
        if abs(diff) >= 20:
            if diff > 0:
                pred = '主胜'
                reason = f"胜率差+{diff:.0f}%"
                strategy = "胜率差偏向"
            else:
                pred = '客胜'
                reason = f"胜率差{diff:.0f}%"
                strategy = "胜率差偏向"
        else:
            # 胜率差小，看8变化
            if home_8 - away_8 >= 2:
                pred = '主胜'
                reason = f"主胜8优势({home_8} vs {away_8})"
                strategy = "8优势"
            elif away_8 - home_8 >= 2:
                pred = '客胜'
                reason = f"客胜8优势({away_8} vs {home_8})"
                strategy = "8优势"
            else:
                pred = max_option
                reason = f"中等置信默认"
                strategy = "中等置信"
    
    return {
        '预测': pred,
        '策略': strategy,
        '理由': reason,
        'confidence': conf,
        'diff': diff,
        'min_option': min_option,
        'max_option': max_option,
        'is_trap': is_trap
    }

# ==================== 主程序 ====================
def main():
    print("=" * 80)
    print("3.16 V7/V8 + 排除法策略预测分析")
    print("=" * 80)
    
    # 获取所有源数据文件
    files = sorted(DATA_DIR.glob("*_源数据.md"))
    print(f"\n找到 {len(files)} 场比赛")
    
    results = []
    
    for f in files:
        info = parse_source_file(f)
        
        if not info.get('initial_odds'):
            print(f"跳过 {info.get('match_id')}: 无赔率数据")
            continue
        
        # 计算V7/V8指标
        v7v8 = calculate_v7_v8(info)
        
        # 应用预测策略
        pred_result = apply_prediction_strategy(v7v8, info)
        
        results.append({
            '编号': info.get('match_id', ''),
            '对阵': f"{info.get('home_team', '')} vs {info.get('away_team', '')}",
            '主胜概率': f"{v7v8['prob_home']:.1f}%" if v7v8 else '-',
            '平局概率': f"{v7v8['prob_draw']:.1f}%" if v7v8 else '-',
            '客胜概率': f"{v7v8['prob_away']:.1f}%" if v7v8 else '-',
            '置信度': f"{v7v8['confidence']:.1f}%" if v7v8 else '-',
            '胜率差': f"{v7v8['diff']:+.1f}%" if v7v8 else '-',
            '8变化': f"[{v7v8['home_8']},{v7v8['draw_8']},{v7v8['away_8']}]" if v7v8 else '-',
            '策略': pred_result['策略'],
            '预测': pred_result['预测'],
            '理由': pred_result['理由'],
            '排除项': pred_result.get('min_option', '-'),
        })
    
    # ==================== 输出结果 ====================
    print("\n" + "=" * 120)
    print(f"{'编号':<8} {'对阵':<22} {'置信度':<8} {'胜率差':<10} {'8变化':<15} {'策略':<15} {'预测':<8} {'理由':<25}")
    print("-" * 120)
    
    for r in results:
        print(f"{r['编号']:<8} {r['对阵'][:20]:<20} {r['置信度']:<8} {r['胜率差']:<10} {r['8变化']:<15} {r['策略']:<15} {r['预测']:<8} {r['理由']:<25}")
    
    # ==================== 高置信度推荐 ====================
    print("\n" + "=" * 80)
    print("【高置信度推荐 (>=55%)】")
    print("=" * 80)
    
    high_conf = [r for r in results if float(r['置信度'].replace('%','')) >= 55]
    for r in high_conf:
        print(f"  {r['编号']} {r['对阵']:<20} 置信度{r['置信度']} 胜率差{r['胜率差']} → 预测:{r['预测']}")
    
    # ==================== 策略统计 ====================
    print("\n" + "=" * 80)
    print("【策略分布】")
    print("=" * 80)
    
    strategy_count = {}
    for r in results:
        s = r['策略']
        strategy_count[s] = strategy_count.get(s, 0) + 1
    
    for s, c in sorted(strategy_count.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}场")
    
    # ==================== 保存Excel ====================
    wb = Workbook()
    ws = wb.active
    ws.title = "3.16_V7V8预测"
    
    # 表头
    headers = ['编号', '对阵', '主胜概率', '平局概率', '客胜概率', '置信度', '胜率差', '8变化', '策略', '预测', '理由', '排除项']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = PatternFill('solid', start_color='366092')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # 数据
    for row_idx, r in enumerate(results, 2):
        ws.cell(row_idx, 1, r['编号'])
        ws.cell(row_idx, 2, r['对阵'])
        ws.cell(row_idx, 3, r['主胜概率'])
        ws.cell(row_idx, 4, r['平局概率'])
        ws.cell(row_idx, 5, r['客胜概率'])
        ws.cell(row_idx, 6, r['置信度'])
        ws.cell(row_idx, 7, r['胜率差'])
        ws.cell(row_idx, 8, r['8变化'])
        ws.cell(row_idx, 9, r['策略'])
        ws.cell(row_idx, 10, r['预测'])
        ws.cell(row_idx, 11, r['理由'])
        ws.cell(row_idx, 12, r['排除项'])
        
        # 高置信度标记绿色
        conf_val = float(r['置信度'].replace('%', ''))
        if conf_val >= 55:
            for col in range(1, 13):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='C6EFCE')
        elif conf_val <= 45:
            for col in range(1, 13):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='FFEB9C')
    
    # 列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 10
    
    wb.save(OUTPUT_FILE)
    print(f"\n已保存: {OUTPUT_FILE}")

if __name__ == '__main__':
    main()
