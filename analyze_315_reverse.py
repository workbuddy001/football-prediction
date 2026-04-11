import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# 读取预测数据
df = pd.read_excel('d:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx')

# 转换概率列
def to_float(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

df['主胜概率'] = df['主胜概率'].apply(to_float)
df['平局概率'] = df['平局概率'].apply(to_float)
df['客胜概率'] = df['客胜概率'].apply(to_float)

# 实际比赛结果
actual_results = {
    '周日001': ('日本女', 1), '周日002': ('福冈黄蜂', 1), '周日003': ('首尔FC', 0),
    '周日004': ('仁川联', 2), '周日005': ('墨胜利', 1), '周日006': ('乌德勒支', 0),
    '周日007': ('热那亚', 0), '周日008': ('汉诺威96', 2), '周日009': ('马洛卡', 1),
    '周日010': ('费耶诺德', 1), '周日011': ('布兰', 0), '周日012': ('水晶宫', 2),
    '周日013': ('诺丁汉', 2), '周日014': ('曼联', 1), '周日015': ('比萨', 1),
    '周日016': ('博洛尼亚', 0), '周日017': ('美因茨', 0), '周日018': ('巴萨', 1),
    '周日019': ('瓦勒伦加', 1), '周日020': ('里昂', 2), '周日021': ('利物浦', 2),
    '周日022': ('柏林联合', 0), '周日023': ('科莫', 1), '周日024': ('贝蒂斯', 2),
    '周日025': ('斯图加特', 1), '周日026': ('拉齐奥', 1), '周日027': ('皇家社会', 1),
    '周日028': ('波尔图', 1), '周日029': ('温哥华', 1),
}

# 获取预测选项
def get_pred_result(row):
    if '主胜' in str(row['首选']):
        return 0, '主胜'
    elif '客胜' in str(row['首选']):
        return 1, '客胜'
    return 2, '平局'

def get_max_prob(row):
    probs = [row['主胜概率'], row['平局概率'], row['客胜概率']]
    idx = probs.index(max(probs))
    return ['主胜', '平局', '客胜'][idx], max(probs)

# 分析每场比赛
results = []
for idx, row in df.iterrows():
    match_id = row['编号'].split('_')[0]
    pred_code, pred_str = get_pred_result(row)
    max_prob_str, max_prob = get_max_prob(row)
    panxing = row['盘型']
    
    actual = actual_results.get(match_id)
    if not actual:
        continue
    actual_code = actual[1]
    actual_str = ['主胜', '客胜', '平局'][actual_code]
    
    is_correct_orig = pred_code == actual_code
    confidence = max_prob
    
    # 新的智能推荐逻辑
    recommend = ""
    reason = ""
    is_correct_new = False
    
    if panxing == '诱盘' and confidence > 50:
        # 诱盘 + 高置信度 → 反诱盘(选概率最低)
        probs = {'主胜': row['主胜概率'], '平局': row['平局概率'], '客胜': row['客胜概率']}
        min_option = min(probs, key=probs.get)
        recommend = f"反诱盘:{min_option}"
        is_correct_new = (min_option == actual_str)
        reason = "诱盘高置信，反向选择"
    elif panxing == '实盘' and confidence <= 50:
        # 实盘 + 低置信度 → 反向推荐(选概率次低或第二高的)
        probs = [('主胜', row['主胜概率']), ('平局', row['平局概率']), ('客胜', row['客胜概率'])]
        probs_sorted = sorted(probs, key=lambda x: x[1], reverse=True)
        # 原预测是最高的，选第二高的
        second_option = probs_sorted[1][0]
        recommend = f"反向:{second_option}"
        is_correct_new = (second_option == actual_str)
        reason = "实盘低置信，反向选择"
    elif confidence > 55 and panxing == '实盘':
        # 高置信度 + 实盘 → 正向推荐
        recommend = f"正向:{max_prob_str}"
        is_correct_new = is_correct_orig
        reason = "高置信实盘，正向"
    elif confidence <= 40:
        recommend = "观望"
        is_correct_new = False
        reason = "置信度太低"
    else:
        recommend = f"正向:{max_prob_str}"
        is_correct_new = is_correct_orig
        reason = "正常分析"
    
    results.append({
        '编号': match_id,
        '盘型': panxing,
        '置信度': f"{max_prob:.1f}%",
        '原预测': pred_str,
        '实际': actual_str,
        '原对错': '对' if is_correct_orig else '错',
        '新推荐': recommend,
        '新对错': '对' if is_correct_new else '错',
        '原因': reason
    })

result_df = pd.DataFrame(results)

print("=" * 90)
print("反向策略验证")
print("=" * 90)

# 总体对比
orig_correct = (result_df['原对错'] == '对').sum()
new_correct = (result_df['新对错'] == '对').sum()
print(f"\n原预测准确率: {orig_correct}/29 = {orig_correct/29*100:.1f}%")
print(f"新策略准确率: {new_correct}/29 = {new_correct/29*100:.1f}%")

# 重点：实盘低置信的反向效果
print("\n" + "=" * 90)
print("【关键发现】实盘 + 低置信度 → 反向推荐效果")
print("=" * 90)

low_conf_real = result_df[(result_df['盘型'] == '实盘') & (result_df['置信度'].apply(lambda x: float(x.replace('%','')) <= 50))]
print(f"\n实盘+低置信度比赛共{len(low_conf_real)}场:")
print(low_conf_real[['编号', '置信度', '原预测', '实际', '原对错', '新推荐', '新对错']].to_string(index=False))

orig_correct_low = (low_conf_real['原对错'] == '对').sum()
new_correct_low = (low_conf_real['新对错'] == '对').sum()
print(f"\n原预测正确: {orig_correct_low}/{len(low_conf_real)} = {orig_correct_low/len(low_conf_real)*100:.1f}%")
print(f"反向推荐正确: {new_correct_low}/{len(low_conf_real)} = {new_correct_low/len(low_conf_real)*100:.1f}%")

print("\n" + "=" * 90)
print("完整对比表")
print("=" * 90)
print(result_df[['编号', '盘型', '置信度', '原预测', '实际', '原对错', '新推荐', '新对错']].to_string(index=False))

# 保存Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "反向策略验证"

headers = ['编号', '盘型', '置信度', '原预测', '实际', '原对错', '新推荐', '新对错', '原因']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = PatternFill('solid', start_color='366092')
    cell.font = Font(bold=True, color='FFFFFF')

for row_idx, r in enumerate(results, 2):
    for col_idx, key in enumerate(['编号', '盘型', '置信度', '原预测', '实际', '原对错', '新推荐', '新对错', '原因'], 1):
        ws.cell(row_idx, col_idx, r[key])
    
    # 颜色标记
    new_cell = ws.cell(row_idx, 7)
    if r['新对错'] == '对':
        new_cell.fill = PatternFill('solid', start_color='00FF00')
    else:
        new_cell.fill = PatternFill('solid', start_color='FF6666')

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 20

wb.save('3.15_反向策略验证.xlsx')
print("\n已保存: 3.15_反向策略验证.xlsx")
