# V5算法修正版 - 平衡3.14和3.15
# 问题：V5对3.14提升到40.6%但对3.15变成0%
# 修正：保留V4核心逻辑，只增加平局条件判断

import os
import re
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def analyze_v5b(info):
    """V5修正版算法"""
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    macao_tip = info.get('macao_tip', '')
    
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "盘型": "未知", "详情": []}
    
    initial_home = [x[0] for x in initial]
    initial_draw = [x[1] for x in initial]
    initial_away = [x[2] for x in initial]
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_initial_home = sum(initial_home) / len(initial_home)
    avg_initial_draw = sum(initial_draw) / len(initial_draw)
    avg_initial_away = sum(initial_away) / len(initial_away)
    
    avg_realtime_home = sum(realtime_home) / len(realtime_home)
    avg_realtime_draw = sum(realtime_draw) / len(realtime_draw)
    avg_realtime_away = sum(realtime_away) / len(realtime_away)
    
    prob_home = 1 / avg_realtime_home / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_draw = 1 / avg_realtime_draw / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_away = 1 / avg_realtime_away / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    
    home_pct_change = (avg_realtime_home - avg_initial_home) / avg_initial_home * 100
    draw_pct_change = (avg_realtime_draw - avg_initial_draw) / avg_initial_draw * 100
    away_pct_change = (avg_realtime_away - avg_initial_away) / avg_initial_away * 100
    
    home_up = sum(1 for i in range(len(initial)) if realtime[i][0] > initial[i][0])
    draw_up = sum(1 for i in range(len(initial)) if realtime[i][1] > initial[i][1])
    away_up = sum(1 for i in range(len(initial)) if realtime[i][2] > initial[i][2])
    
    home_up_pct = home_up / len(initial) * 100
    away_up_pct = away_up / len(initial) * 100
    draw_down = len(initial) - draw_up
    draw_down_pct = draw_down / len(initial) * 100
    
    macao_home = realtime[0][0] if realtime else avg_realtime_home
    macao_away = realtime[0][2] if realtime else avg_realtime_away
    
    details = []
    prediction = ""
    
    # === 保留V4核心规则 ===
    
    # 规则1: 强胆
    if avg_realtime_home < 1.5:
        prediction = f"{home}主胜"
        details.append("规则1: 强胆主胜")
    elif avg_realtime_away < 1.5:
        prediction = f"{away}客胜"
        details.append("规则1: 强胆客胜")
    
    # 规则2: 强队主场 + 近况好
    elif 1.5 < avg_realtime_home < 2.0 and home_w >= home_l:
        prediction = f"{home}主胜"
        details.append("规则2: 强队主场+近况好")
    
    # 规则3: 强队客场 + 近况好
    elif 1.5 < avg_realtime_away < 2.0 and away_w >= away_l:
        prediction = f"{away}客胜"
        details.append("规则3: 强队客场+近况好")
    
    # 规则4: 主胜大升
    elif home_pct_change > 15 and home_up_pct > 80:
        prediction = f"{away}客胜"
        details.append(f"规则4: 主胜升{home_pct_change:.1f}%")
    
    # 规则5: 客胜大升
    elif away_pct_change > 15 and away_up_pct > 70:
        prediction = f"{home}主胜"
        details.append(f"规则5: 客胜升{away_pct_change:.1f}%")
    
    # 规则6: 主队近况好
    elif home_w >= 3 and avg_realtime_home < 2.5:
        prediction = f"{home}主胜"
        details.append(f"规则6: 主队W{home_w}场")
    
    # 规则7: 客队近况好
    elif away_w >= 3 and avg_realtime_away < 2.5:
        prediction = f"{away}客胜"
        details.append(f"规则7: 客队W{away_w}场")
    
    # 规则8: 澳门客胜低
    elif macao_away < avg_realtime_away * 0.85:
        prediction = f"{away}客胜"
        details.append("规则8: 澳门客胜低")
    
    # 规则9: 澳门主胜低
    elif macao_home < avg_realtime_home * 0.90:
        prediction = f"{home}主胜"
        details.append("规则9: 澳门主胜低")
    
    # 规则10: 澳门推荐主队
    elif macao_tip and home in macao_tip:
        prediction = f"{home}主胜"
        details.append("规则10: 澳门推荐主队")
    
    # 规则11: 澳门推荐客队
    elif macao_tip and away in macao_tip:
        prediction = f"{away}客胜"
        details.append("规则11: 澳门推荐客队")
    
    # === V5新增：平局规则（只在非常明显时）===
    # 规则12: 平局趋势非常明显（降赔公司>65% + 降赔幅度>3% + 概率>32%）
    elif draw_down_pct > 65 and draw_pct_change < -3 and prob_draw > 0.32:
        prediction = "平局"
        details.append(f"规则12: 平局明显,降{abs(draw_pct_change):.1f}%")
    
    # 规则13: 主胜不稳 + 平局概率高
    elif home_pct_change > 8 and away_pct_change > 0 and prob_draw > 0.33:
        prediction = "平局"
        details.append("规则13: 主胜不稳+高平局概率")
    
    # 规则14: 主胜不稳
    elif home_pct_change > 5 and away_pct_change < 0:
        prediction = f"{away}客胜"
        details.append("规则14: 主胜不稳")
    
    # 规则15: 默认概率最高
    else:
        if prob_home > prob_away and prob_home > prob_draw:
            prediction = f"{home}主胜"
            details.append("规则15: 主胜概率最高")
        elif prob_away > prob_home and prob_away > prob_draw:
            prediction = f"{away}客胜"
            details.append("规则15: 客胜概率最高")
        else:
            prediction = "平局"
            details.append("规则15: 平局概率最高")
    
    # 盘型判断
    if home_up_pct > 80 and draw_down_pct > 50:
        panxing = "实盘"
    else:
        panxing = "诱盘"
    
    return {
        "预测": prediction,
        "盘型": panxing,
        "详情": details,
        "主胜概率": f"{prob_home*100:.1f}%",
        "平局概率": f"{prob_draw*100:.1f}%",
        "客胜概率": f"{prob_away*100:.1f}%",
        "主胜": f"{avg_realtime_home:.2f}",
        "平局": f"{avg_realtime_draw:.2f}",
        "客胜": f"{avg_realtime_away:.2f}"
    }

def parse_source_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    match_info = {}
    
    filename = Path(filepath).stem
    match = re.match(r'周六(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        match_info['match_id'] = f"周六{match.group(1)}"
        match_info['home_team'] = match.group(2)
        match_info['away_team'] = match.group(3)
    
    match = re.search(r'赛事.*?\|.*?(\S+)', content)
    if match:
        match_info['league'] = match.group(1)
    
    match = re.search(r'主队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['home_w'] = int(match.group(1))
        match_info['home_l'] = int(match.group(3))
    
    match = re.search(r'客队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['away_w'] = int(match.group(1))
        match_info['away_l'] = int(match.group(3))
    
    match = re.search(r'澳门推荐.*?\|.*?(\S+)', content)
    if match:
        match_info['macao_tip'] = match.group(1)
    
    initial_odds = []
    realtime_odds = []
    
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            initial_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            realtime_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match_info['initial_odds'] = initial_odds
    match_info['realtime_odds'] = realtime_odds
    
    return match_info

# 验证3.14
folder = "分析模板/3.14"
files = sorted(Path(folder).glob("*.md"))
results = []

for f in files:
    info = parse_source_file(f)
    if not info.get('home_team') or not info.get('initial_odds'):
        continue
    result = analyze_v5b(info)
    results.append({
        '编号': info.get('match_id'),
        '对阵': f"{info.get('home_team')} vs {info.get('away_team')}",
        '预测': result['预测'],
        '详情': ', '.join(result.get('详情', []))
    })

actual_14 = {
    "周六001": "平局", "周六002": "客胜", "周六003": "平局", "周六004": "客胜",
    "周六005": "主胜", "周六006": "主胜", "周六007": "平局", "周六008": "主胜",
    "周六009": "主胜", "周六010": "客胜", "周六011": "主胜", "周六012": "平局",
    "周六013": "平局", "周六014": "主胜", "周六015": "主胜", "周六016": "平局",
    "周六017": "平局", "周六018": "主胜", "周六019": "主胜", "周六020": "主胜",
    "周六021": "主胜", "周六022": "主胜", "周六023": "客胜", "周六024": "平局",
    "周六025": "主胜", "周六026": "客胜", "周六027": "客胜", "周六028": "客胜",
    "周六029": "平局", "周六030": "主胜", "周六031": "客胜", "周六032": "平局",
}

def check_match(pred, actual):
    pred = str(pred).lower()
    if "主胜" in pred:
        return "主胜" == actual
    elif "客胜" in pred:
        return "客胜" == actual
    elif "平局" in pred:
        return "平局" == actual
    return False

correct14 = 0
for r in results:
    if check_match(r['预测'], actual_14.get(r['编号'], "")):
        correct14 += 1

print(f"V5修正版 - 3.14准确率: {correct14/32*100:.1f}% ({correct14}/32)")

# 验证3.15
wb15 = openpyxl.load_workbook('3.15_比赛预测汇总.xlsx')
ws15 = wb15.active
pred15_count = 0
for row in range(2, 31):
    id_ = ws15.cell(row, 1).value
    pred = ws15.cell(row, 4).value
    actual = {
        "周日001": "主胜", "周日003": "客胜", "周日004": "平局", "周日006": "客胜",
        "周日007": "客胜", "周日008": "平局", "周日009": "主胜", "周日010": "主胜",
        "周日011": "主胜", "周日012": "平局", "周日013": "平局", "周日014": "主胜",
        "周日015": "主胜", "周日016": "客胜", "周日017": "客胜", "周日018": "主胜",
        "周日019": "主胜", "周日020": "平局", "周日021": "平局", "周日022": "客胜",
        "周日023": "主胜", "周日024": "平局", "周日025": "主胜", "周日026": "主胜",
        "周日027": "主胜", "周日028": "客胜", "周日029": "主胜",
    }.get(id_, "")
    if check_match(pred, actual):
        pred15_count += 1

print(f"V5修正版 - 3.15准确率: {pred15_count/29*100:.1f}% ({pred15_count}/29)")

# 生成Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.14_V5b预测"

headers = ['编号', '对阵', '预测', '详情']
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")

for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['预测'])
    ws.cell(row, 4, r['详情'])

wb.save("3.14_V5b预测.xlsx")
print("\n已保存: 3.14_V5b预测.xlsx")
