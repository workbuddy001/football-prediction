import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 初盘赔率数据
initial_odds = [
    (2.88, 2.80, 2.35), (3.00, 2.75, 2.50), (3.13, 2.91, 2.23), (3.10, 2.87, 2.50),
    (3.20, 3.00, 2.35), (3.15, 3.05, 2.45), (3.15, 3.05, 2.35), (3.15, 3.25, 2.41),
    (3.40, 3.10, 2.38), (3.10, 2.90, 2.45), (3.10, 2.87, 2.55), (3.02, 2.95, 2.57),
    (3.20, 3.00, 2.35), (3.10, 2.90, 2.50), (2.99, 2.92, 2.35), (3.15, 2.90, 2.38),
    (3.10, 2.95, 2.50), (3.15, 3.05, 2.32), (3.15, 3.05, 2.32), (3.15, 3.25, 2.41),
    (3.15, 2.80, 2.18), (3.15, 3.05, 2.32), (3.25, 2.98, 2.44), (2.80, 3.05, 2.60),
    (3.30, 3.09, 2.58), (3.15, 3.05, 2.35), (3.00, 3.10, 2.35), (3.10, 2.80, 2.30),
    (3.15, 3.05, 2.35), (3.15, 2.98, 2.52), (2.88, 3.00, 2.50), (3.10, 3.00, 2.30),
    (3.37, 3.00, 2.50), (3.20, 3.05, 2.41), (3.25, 3.15, 2.40), (3.15, 3.10, 2.45),
    (3.20, 2.90, 2.30), (3.18, 3.05, 2.52), (3.35, 3.10, 2.33), (3.15, 3.00, 2.40),
    (3.00, 2.88, 2.50), (2.85, 2.70, 2.40), (3.20, 3.10, 2.35), (2.80, 2.80, 2.70),
    (3.15, 3.00, 2.38), (3.20, 3.20, 2.30), (3.30, 3.04, 2.42), (3.20, 3.00, 2.35),
    (3.26, 2.98, 2.35), (3.15, 2.95, 2.45), (3.10, 2.70, 2.50)
]

realtime_odds = [
    (3.30, 2.69, 2.20), (3.30, 2.80, 2.30), (3.13, 2.91, 2.23), (3.30, 2.87, 2.40),
    (3.30, 3.00, 2.37), (3.35, 2.95, 2.45), (3.40, 2.90, 2.45), (3.60, 2.95, 2.38),
    (3.50, 3.00, 2.40), (3.50, 2.80, 2.38), (3.30, 2.87, 2.40), (3.56, 3.00, 2.41),
    (3.45, 2.95, 2.38), (3.40, 2.87, 2.45), (3.36, 2.90, 2.44), (3.35, 2.90, 2.40),
    (3.30, 2.87, 2.40), (3.20, 2.94, 2.51), (3.35, 2.90, 2.47), (3.50, 2.95, 2.40),
    (3.20, 2.77, 2.18), (3.35, 2.90, 2.47), (3.60, 3.05, 2.48), (3.40, 2.95, 2.40),
    (3.38, 3.08, 2.44), (3.40, 2.95, 2.40), (3.05, 3.10, 2.35), (3.40, 2.90, 2.40),
    (3.35, 2.95, 2.40), (3.25, 3.05, 2.42), (3.40, 2.90, 2.38), (3.30, 2.88, 2.30),
    (3.47, 2.95, 2.48), (3.40, 2.95, 2.46), (3.40, 2.95, 2.40), (3.15, 3.05, 2.45),
    (3.40, 2.90, 2.40), (3.44, 2.99, 2.44), (3.30, 2.90, 2.48), (3.30, 2.90, 2.40),
    (3.30, 2.80, 2.38), (3.35, 2.85, 2.35), (3.40, 2.90, 2.40), (3.40, 2.90, 2.30),
    (3.30, 2.95, 2.43), (3.50, 2.95, 2.30), (3.42, 3.01, 2.48), (3.45, 3.00, 2.46),
    (3.46, 2.91, 2.36), (3.45, 2.95, 2.35), (3.30, 2.80, 2.30)
]

companies = [
    '竞官', '威尔', '*门', '立*', 't3*5', 'Irwtn', 'SI', '*冠', '易*博', '伟*',
    'Bn', 'Pile平*', '10t', 'C*l', '利*', 'Ut', 'Stgt', 'IBCt', 'Mon88', '金*博',
    '香会', '12t', '必*', '18t', '1xt', 'Irtps', 'Et', 't-at-He', 'tfg', 'tsn S*tsbk',
    'tway', 'BeSts', 'Clt', 'DaFt', 'Font', 'Gnia S*t', 'Hska Lja', 'Mhon', 'Miant', 'Ny',
    'Py Pr', 'P*U', 'S*l', 'Sy*t', 'Styt', 'Tico', 'TS*t', 'Tot', 'Wt', '1t', '88St'
]

# 计算变化百分比
home_pct_change = []
draw_pct_change = []
away_pct_change = []

for i in range(len(initial_odds)):
    home_pct = (realtime_odds[i][0] - initial_odds[i][0]) / initial_odds[i][0] * 100
    draw_pct = (realtime_odds[i][1] - initial_odds[i][1]) / initial_odds[i][1] * 100
    away_pct = (realtime_odds[i][2] - initial_odds[i][2]) / initial_odds[i][2] * 100
    home_pct_change.append(home_pct)
    draw_pct_change.append(draw_pct)
    away_pct_change.append(away_pct)

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "水位变化百分比"

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
title_font = Font(bold=True, size=14)
center = Alignment(horizontal='center', vertical='center')

ws['A1'] = '维罗纳 vs 热那亚 - 水位变动幅度百分比分析'
ws['A1'].font = title_font
ws.merge_cells('A1:J1')
ws['A1'].alignment = center

# 统计汇总
ws['A3'] = '一、水位变动百分比统计'
ws['A3'].font = Font(bold=True, size=12)

stats_headers = ['指标', '主胜变化%', '平局变化%', '客胜变化%']
for c, h in enumerate(stats_headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

stats_data = [
    ['平均值', f'{np.mean(home_pct_change):.2f}%', f'{np.mean(draw_pct_change):.2f}%', f'{np.mean(away_pct_change):.2f}%'],
    ['中位数', f'{np.median(home_pct_change):.2f}%', f'{np.median(draw_pct_change):.2f}%', f'{np.median(away_pct_change):.2f}%'],
    ['最大升幅', f'{max(home_pct_change):.2f}%', f'{max(draw_pct_change):.2f}%', f'{max(away_pct_change):.2f}%'],
    ['最大降幅', f'{min(home_pct_change):.2f}%', f'{min(draw_pct_change):.2f}%', f'{min(away_pct_change):.2f}%'],
    ['标准差', f'{np.std(home_pct_change):.2f}%', f'{np.std(draw_pct_change):.2f}%', f'{np.std(away_pct_change):.2f}%'],
]

for r, row_data in enumerate(stats_data, 5):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

# 变化方向统计
ws['A11'] = '二、变化方向统计'
ws['A11'].font = Font(bold=True, size=12)

dir_headers = ['类型', '上升(家)', '下降(家)', '不变(家)', '上升占比', '下降占比']
for c, h in enumerate(dir_headers, 1):
    cell = ws.cell(row=12, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

up_home = sum(1 for x in home_pct_change if x > 0)
down_home = sum(1 for x in home_pct_change if x < 0)
same_home = sum(1 for x in home_pct_change if x == 0)

up_draw = sum(1 for x in draw_pct_change if x > 0)
down_draw = sum(1 for x in draw_pct_change if x < 0)
same_draw = sum(1 for x in draw_pct_change if x == 0)

up_away = sum(1 for x in away_pct_change if x > 0)
down_away = sum(1 for x in away_pct_change if x < 0)
same_away = sum(1 for x in away_pct_change if x == 0)

total = len(initial_odds)

dir_data = [
    ['主胜', up_home, down_home, same_home, f'{up_home/total*100:.1f}%', f'{down_home/total*100:.1f}%'],
    ['平局', up_draw, down_draw, same_draw, f'{up_draw/total*100:.1f}%', f'{down_draw/total*100:.1f}%'],
    ['客胜', up_away, down_away, same_away, f'{up_away/total*100:.1f}%', f'{down_away/total*100:.1f}%'],
]

for r, row_data in enumerate(dir_data, 13):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

# 澳门分析
ws['A18'] = '三、澳门赔率分析'
ws['A18'].font = Font(bold=True, size=12)

# 澳门是第3家(*门)
macao_idx = 2
macao_init = initial_odds[macao_idx]
macao_real = realtime_odds[macao_idx]

macao_home_pct = (macao_real[0] - macao_init[0]) / macao_init[0] * 100
macao_draw_pct = (macao_real[1] - macao_init[1]) / macao_init[1] * 100
macao_away_pct = (macao_real[2] - macao_init[2]) / macao_init[2] * 100

macao_headers = ['类型', '初盘', '即时', '变化', '变化%']
for c, h in enumerate(macao_headers, 1):
    cell = ws.cell(row=19, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

macao_data = [
    ['主胜', macao_init[0], macao_real[0], macao_real[0] - macao_init[0], f'{macao_home_pct:.2f}%'],
    ['平局', macao_init[1], macao_real[1], macao_real[1] - macao_init[1], f'{macao_draw_pct:.2f}%'],
    ['客胜', macao_init[2], macao_real[2], macao_real[2] - macao_init[2], f'{macao_away_pct:.2f}%'],
]

for r, row_data in enumerate(macao_data, 20):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val).alignment = center

# 详细数据
ws['A25'] = '四、详细水位变化数据'
ws['A25'].font = Font(bold=True, size=12)

detail_headers = ['公司', '初盘主胜', '即时主胜', '主胜变化%', '初盘平局', '即时平局', '平局变化%', '初盘客胜', '即时客胜', '客胜变化%']
for c, h in enumerate(detail_headers, 1):
    cell = ws.cell(row=26, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

for i in range(len(initial_odds)):
    row = 27 + i
    data = [
        companies[i],
        initial_odds[i][0], realtime_odds[i][0], f'{home_pct_change[i]:.2f}%',
        initial_odds[i][1], realtime_odds[i][1], f'{draw_pct_change[i]:.2f}%',
        initial_odds[i][2], realtime_odds[i][2], f'{away_pct_change[i]:.2f}%'
    ]
    for c, val in enumerate(data, 1):
        ws.cell(row=row, column=c, value=val).alignment = center

# 列宽
ws.column_dimensions['A'].width = 12
for col in 'BCDEFGHIJ':
    ws.column_dimensions[col].width = 14

wb.save('维罗纳vs热那亚_水位变化分析.xlsx')
print("Excel已生成")

# 打印关键数据供报告使用
print("\n=== 关键统计数据 ===")
print(f"主胜平均涨幅: {np.mean(home_pct_change):.2f}%")
print(f"平局平均变化: {np.mean(draw_pct_change):.2f}%")
print(f"客胜平均变化: {np.mean(away_pct_change):.2f}%")
print(f"主胜上升占比: {up_home/total*100:.1f}%")
print(f"平局下降占比: {down_draw/total*100:.1f}%")
print(f"客胜下降占比: {down_away/total*100:.1f}%")
print(f"澳门主胜变化: {macao_home_pct:.2f}%")
print(f"澳门平局变化: {macao_draw_pct:.2f}%")
print(f"澳门客胜变化: {macao_away_pct:.2f}%")
