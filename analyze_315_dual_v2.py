import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

# 读取预测数据
df = pd.read_excel('d:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx')

def to_float(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

df['主胜概率'] = df['主胜概率'].apply(to_float)
df['平局概率'] = df['平局概率'].apply(to_float)
df['客胜概率'] = df['客胜概率'].apply(to_float)

actual_results = {
    '周日001': ('日本女', 1), '周日002': ('福冈黄蜂', 1), '周日003': ('首尔FC', 0),
    '周日004': ('仁川联', 2), '周日005': ('墨胜利', 1), '周日006': ('乌德勒支', 0),
    '周日007': ('热那亚', 0), '周日008': ('汉诺威96', 2), '周日009': ('马洛卡', 1),
    '周日010': ('费耶诺德', 1), '周日011': ('布兰', 0), '周日012': ('水晶宫', 2),
    '周日013': ('诺丁汉', 2), '周日014': ('曼联', 1), '周日015': ('比萨', 1),
    '周日016': ('博洛尼亚', 0), '周日017': ('美因茨', 0), '周日018': ('巴萨', 1),
    '周日019': ('瓦勒伦加', 1), '周日020': ('里昂', 2), '周日021': ('利物浦', 2),
    '周日022': ('柏林联合', 0), '周日023': ('科莫', 1), '周日024': ('贝蒂斯', 2),
    '周日025': ('斯图加特', 1), '周日026': ('拉齐奥', 1), '周日027': ('皇家社会', 1),
    '周日028': ('波尔图', 1), '周日029': ('温哥华', 1),
}

results = []
for idx, row in df.iterrows():
    match_id = row['编号'].split('_')[0]
    
    # 从"首选"列直接获取原预测
    first_choice = str(row['首选']).strip()
    if '主胜' in first_choice:
        pred_code, pred_str = 0, '主胜'
    elif '客胜' in first_choice:
        pred_code, pred_str = 1, '客胜'
    else:
        pred_code, pred_str = 2, '平局'
    
    # 获取最高概率
    probs = {'主胜': row['主胜概率'], '平局': row['平局概率'], '客胜': row['客胜概率']}
    max_prob_str = max(probs, key=probs.get)
    max_prob = probs[max_prob_str]
    
    panxing = row['盘型']
    
    actual = actual_results.get(match_id)
    if not actual:
        continue
    actual_code = actual[1]
    actual_str = ['主胜', '客胜', '平局'][actual_code]
    
    is_correct_orig = pred_code == actual_code
    confidence = max_prob
    
    # 双选策略
    dual_pick = ""
    dual_hit = False
    reason = ""
    
    if panxing == '诱盘' and confidence > 50:
        # 诱盘高置信 → 反诱盘(单选概率最低)
        min_option = min(probs, key=probs.get)
        dual_pick = min_option
        dual_hit = (min_option == actual_str)
        reason = "诱盘高置信→反诱盘"
    elif panxing == '实盘' and confidence <= 50:
        # 实盘低置信 → 双选(原预测 + 第二高概率)
        prob_list = [('主胜', row['主胜概率']), ('平局', row['平局概率']), ('客胜', row['客胜概率'])]
        prob_sorted = sorted(prob_list, key=lambda x: x[1], reverse=True)
        second_option = prob_sorted[1][0]
        
        dual_pick = f"{pred_str}或{second_option}"
        dual_hit = (pred_str == actual_str) or (second_option == actual_str)
        reason = f"实盘低置信→双选"
    elif confidence > 55 and panxing == '实盘':
        dual_pick = max_prob_str
        dual_hit = is_correct_orig
        reason = "高置信实盘→正向"
    elif confidence <= 40:
        dual_pick = "观望"
        dual_hit = False
        reason = "置信度太低"
    else:
        dual_pick = max_prob_str
        dual_hit = is_correct_orig
        reason = "正常分析"
    
    results.append({
        '编号': match_id,
        '盘型': panxing,
        '置信度': f"{max_prob:.1f}%",
        '原预测': pred_str,
        '实际': actual_str,
        '原对错': 'O' if is_correct_orig else 'X',
        '双选': dual_pick,
        '双选中': 'O' if dual_hit else 'X',
        '原因': reason
    })

result_df = pd.DataFrame(results)

print("=" * 95)
print("双选策略效果验证 (修正版)")
print("=" * 95)

orig_correct = (result_df['原对错'] == 'O').sum()
dual_correct = (result_df['双选中'] == 'O').sum()

print(f"\n原单选准确率: {orig_correct}/29 = {orig_correct/29*100:.1f}%")
print(f"双选命中准确率: {dual_correct}/29 = {dual_correct/29*100:.1f}%")

# 重点：实盘低置信的双选效果
print("\n" + "=" * 95)
print("【核心】实盘+低置信度 → 双选效果")
print("=" * 95)

low_conf_real = result_df[(result_df['盘型'] == '实盘') & (result_df['置信度'].apply(lambda x: float(x.replace('%','')) <= 50))]
print(f"\n实盘+低置信度 共{len(low_conf_real)}场:")
print(low_conf_real[['编号', '置信度', '原预测', '实际', '原对错', '双选', '双选中']].to_string(index=False))

orig_c = (low_conf_real['原对错'] == 'O').sum()
dual_c = (low_conf_real['双选中'] == 'O').sum()
print(f"\n原单选正确: {orig_c}/{len(low_conf_real)} = {orig_c/len(low_conf_real)*100:.1f}%")
print(f"双选命中: {dual_c}/{len(low_conf_real)} = {dual_c/len(low_conf_real)*100:.1f}%")

# 统计汇总
print("\n" + "=" * 95)
print("策略汇总")
print("=" * 95)
summary = result_df.groupby('原因').agg({
    '原对错': lambda x: (x == 'O').sum(),
    '双选中': lambda x: (x == 'O').sum(),
    '编号': 'count'
}).reset_index()
summary.columns = ['策略', '原正确', '双选正确', '总数']
print(summary.to_string(index=False))

# 保存Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "双选策略"

headers = ['编号', '盘型', '置信度', '原预测', '实际', '原对错', '双选', '双选中', '原因']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = PatternFill('solid', start_color='366092')
    cell.font = Font(bold=True, color='FFFFFF')

for row_idx, r in enumerate(results, 2):
    for col_idx, key in enumerate(['编号', '盘型', '置信度', '原预测', '实际', '原对错', '双选', '双选中', '原因'], 1):
        ws.cell(row_idx, col_idx, r[key])
    
    if r['双选中'] == 'O':
        ws.cell(row_idx, 7).fill = PatternFill('solid', start_color='00FF00')
    else:
        ws.cell(row_idx, 7).fill = PatternFill('solid', start_color='FF6666')

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 16
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 25

wb.save('3.15_双选策略.xlsx')
print("\n已保存: 3.15_双选策略.xlsx")
