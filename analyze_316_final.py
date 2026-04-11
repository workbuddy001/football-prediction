# -*- coding: utf-8 -*-
"""
3.16比赛预测分析脚本 - 优化版
基于V7/V8分析 + 实盘排除法策略

核心策略:
1. 实盘 + 低置信度 → 排除法(选最低)
2. 实盘 → 正向
3. 非实盘 → 其他策略
"""

import re
import ast
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== 配置 ====================
DATA_DIR = Path("d:/work/workbuddy/足球预测/分析模板/3.16")
OUTPUT_FILE = "3.16_V7V8预测分析_优化版.xlsx"

# ==================== 解析源数据 ====================
def parse_source_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    info = {}
    filename = filepath.stem
    
    match = re.match(r'(周一|周二|周三|周四|周五|周六)(\d+)_([^vs]+)vs(.+?)_源数据', filename)
    if match:
        info['date'] = match.group(1)
        info['num'] = match.group(2)
        info['home_team'] = match.group(3).strip()
        info['away_team'] = match.group(4).strip()
        info['match_id'] = f"{match.group(1)}{match.group(2)}"
    
    match = re.search(r'澳门推荐.*?([^\n]+)', content)
    if match:
        info['macao_tip'] = match.group(1).strip()
    
    # 初盘赔率
    initial_odds = []
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            odds_str = '[' + match.group(1) + ']'
            odds_str = re.sub(r'#.*', '', odds_str)
            initial_odds = ast.literal_eval(odds_str)
        except: pass
    
    # 即时赔率
    realtime_odds = []
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            odds_str = '[' + match.group(1) + ']'
            odds_str = re.sub(r'#.*', '', odds_str)
            realtime_odds = ast.literal_eval(odds_str)
        except: pass
    
    info['initial_odds'] = initial_odds
    info['realtime_odds'] = realtime_odds
    
    return info

def calculate_v7_v8(info):
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    
    if not initial or not realtime:
        return None
    
    init_home = sum(x[0] for x in initial) / len(initial)
    init_draw = sum(x[1] for x in initial) / len(initial)
    init_away = sum(x[2] for x in initial) / len(initial)
    
    real_home = sum(x[0] for x in realtime) / len(realtime)
    real_draw = sum(x[1] for x in realtime) / len(realtime)
    real_away = sum(x[2] for x in realtime) / len(realtime)
    
    # 即时概率
    real_prob_home = 1/real_home / (1/real_home + 1/real_draw + 1/real_away)
    real_prob_draw = 1/real_draw / (1/real_home + 1/real_draw + 1/real_away)
    real_prob_away = 1/real_away / (1/real_home + 1/real_draw + 1/real_away)
    
    confidence = max(real_prob_home, real_prob_draw, real_prob_away) * 100
    diff = (real_prob_home - real_prob_away) * 100
    
    # 8变化
    home_8 = sum(1 for i in range(len(initial)) if realtime[i][0] < initial[i][0])
    draw_8 = sum(1 for i in range(len(initial)) if realtime[i][1] < initial[i][1])
    away_8 = sum(1 for i in range(len(initial)) if realtime[i][2] < initial[i][2])
    
    return {
        'confidence': confidence,
        'diff': diff,
        'home_8': home_8,
        'draw_8': draw_8,
        'away_8': away_8,
        'prob_home': real_prob_home * 100,
        'prob_draw': real_prob_draw * 100,
        'prob_away': real_prob_away * 100,
    }

def is_real_market(conf, diff, home_8, draw_8, away_8):
    """
    判断是否为实盘
    核心条件：
    1. 置信度和胜率差成比例
    2. 8变化中庸（每个方向|8变化| <= 2）
    """
    abs_diff = abs(diff)
    
    # 8变化中庸判断：每个方向|8| <= 2
    is_moderate_8 = abs(home_8) <= 2 and abs(draw_8) <= 2 and abs(away_8) <= 2
    
    if not is_moderate_8:
        return False, f"非实盘(8:[{home_8},{draw_8},{away_8}]非中庸)"
    
    # 实盘1: 55-65% + 10-20%
    if 55 <= conf < 65 and 10 <= abs_diff <= 20:
        return True, "实盘1(55-65%+10-20%)"
    
    # 实盘2: 65-75% + 30-40%
    if 65 <= conf < 75 and 30 <= abs_diff <= 40:
        return True, "实盘2(65-75%+30-40%)"
    
    # 实盘3: 75%+ + 40%+
    if conf >= 75 and abs_diff >= 40:
        return True, "实盘3(75%++40%+)"
    
    # 比例实盘: 胜率差/置信度 ≈ 0.4-0.9（作为补充）
    if conf >= 50 and abs_diff >= 10:
        ratio = abs_diff / conf
        if 0.4 <= ratio <= 0.9:
            return True, f"比例实盘(conf{conf:.0f}%*ratio{ratio:.1f}={abs_diff:.0f}%)"
    
    return False, "非实盘"

def apply_prediction_strategy(v7v8, info):
    if not v7v8:
        return {'预测': '数据不足', '策略': '-', '理由': '数据不足', '盘型': '-'}
    
    conf = v7v8['confidence']
    diff = v7v8['diff']
    home_8 = v7v8['home_8']
    draw_8 = v7v8['draw_8']
    away_8 = v7v8['away_8']
    abs_diff = abs(diff)
    
    probs = {
        '主胜': v7v8['prob_home'],
        '平局': v7v8['prob_draw'],
        '客胜': v7v8['prob_away']
    }
    
    max_option = max(probs, key=probs.get)
    min_option = min(probs, key=probs.get)
    
    is_real, market_type = is_real_market(conf, diff, home_8, draw_8, away_8)
    
    # ==================== 核心策略 ====================
    
    # 策略1: 实盘 + 低置信度(<50%) → 排除法(选最低)
    if is_real and conf < 50:
        pred = min_option
        reason = f"实盘+低置信({conf:.0f}%)→排除法"
        strategy = "排除法"
    
    # 策略2: 实盘 → 正向
    elif is_real:
        if diff > 0:
            pred = '主胜'
        else:
            pred = '客胜'
        reason = f"实盘({market_type})→正向"
        strategy = "实盘正向"
    
    # 策略3: 高置信度(>=70%) → 正向
    elif conf >= 70:
        pred = max_option
        reason = f"高置信({conf:.0f}%)→正向"
        strategy = "高置信正向"
    
    # 策略4: 低置信度(<45%) → 排除法
    elif conf < 45:
        pred = min_option
        reason = f"低置信({conf:.0f}%)→排除法"
        strategy = "排除法"
    
    # 策略5: 胜率差明显(>=25%)
    elif abs_diff >= 25:
        if diff > 0:
            pred = '主胜'
        else:
            pred = '客胜'
        reason = f"胜率差偏向({diff:+.0f}%)"
        strategy = "胜率差偏向"
    
    # 策略6: 8变化偏向
    else:
        if home_8 - away_8 >= 2:
            pred = '主胜'
            reason = f"主胜8优势({home_8}vs{away_8})"
            strategy = "8优势"
        elif away_8 - home_8 >= 2:
            pred = '客胜'
            reason = f"客胜8优势({away_8}vs{home_8})"
            strategy = "8优势"
        else:
            pred = max_option
            reason = f"默认({conf:.0f}%)"
            strategy = "默认"
    
    return {
        '预测': pred, '策略': strategy, '理由': reason, '盘型': market_type,
        'confidence': conf, 'diff': diff, 'min_option': min_option, 'max_option': max_option,
        'is_real': is_real
    }

# ==================== 主程序 ====================
def main():
    print("=" * 80)
    print("3.16 V7/V8 + 实盘排除法策略预测分析(优化版)")
    print("=" * 80)
    
    files = sorted(DATA_DIR.glob("*_源数据.md"))
    print(f"\n找到 {len(files)} 场比赛")
    
    results = []
    
    for f in files:
        info = parse_source_file(f)
        
        if not info.get('initial_odds'):
            print(f"跳过 {info.get('match_id')}: 无赔率数据")
            continue
        
        v7v8 = calculate_v7_v8(info)
        pred_result = apply_prediction_strategy(v7v8, info)
        
        results.append({
            '编号': info.get('match_id', ''),
            '对阵': f"{info.get('home_team', '')} vs {info.get('away_team', '')}",
            '主胜概率': f"{v7v8['prob_home']:.1f}%" if v7v8 else '-',
            '平局概率': f"{v7v8['prob_draw']:.1f}%" if v7v8 else '-',
            '客胜概率': f"{v7v8['prob_away']:.1f}%" if v7v8 else '-',
            '置信度': f"{v7v8['confidence']:.1f}%" if v7v8 else '-',
            '胜率差': f"{v7v8['diff']:+.1f}%" if v7v8 else '-',
            '8变化': f"[{v7v8['home_8']},{v7v8['draw_8']},{v7v8['away_8']}]" if v7v8 else '-',
            '盘型': pred_result['盘型'],
            '策略': pred_result['策略'],
            '排除项': pred_result['min_option'],
            '预测': pred_result['预测'],
            '理由': pred_result['理由'],
            '澳门推荐': info.get('macao_tip', ''),
        })
    
    # 打印结果
    print("\n" + "=" * 130)
    print(f"{'编号':<8} {'对阵':<24} {'置信度':<8} {'胜率差':<8} {'盘型':<22} {'策略':<12} {'排除项':<6} {'预测':<6} {'实际':<6} {'结果'}")
    print("=" * 130)
    
    actual_results = {
        '周一001': '客胜', '周一002': '客胜', '周一003': '客胜', '周一004': '平局',
        '周一005': '客胜', '周一006': '平局',
        '周二001': '客胜', '周二002': '客胜', '周二003': '平局', '周二004': '主胜',
        '周二005': '主胜', '周二006': '主胜', '周二007': '客胜', '周二008': '客胜',
    }
    
    for r in results:
        match_id = r['编号']
        actual = actual_results.get(match_id, '')
        is_correct = 'O' if r['预测'] == actual else ('X' if actual else '-')
        
        print(f"{r['编号']:<8} {r['对阵']:<24} {r['置信度']:<8} {r['胜率差']:<8} {r['盘型']:<22} {r['策略']:<12} {r['排除项']:<6} {r['预测']:<6} {actual:<6} {is_correct}")
    
    # 统计
    print("\n" + "=" * 80)
    print("统计结果")
    print("=" * 80)
    
    strategy_stats = {}
    for r in results:
        match_id = r['编号']
        actual = actual_results.get(match_id, '')
        if not actual:
            continue
        
        strategy = r['策略']
        if strategy not in strategy_stats:
            strategy_stats[strategy] = {'total': 0, 'correct': 0}
        
        strategy_stats[strategy]['total'] += 1
        if r['预测'] == actual:
            strategy_stats[strategy]['correct'] += 1
    
    print("\n按策略统计:")
    for strategy, stats in sorted(strategy_stats.items(), key=lambda x: -x[1]['correct']/max(x[1]['total'],1)):
        rate = stats['correct']/max(stats['total'],1)*100
        print(f"  {strategy}: {stats['correct']}/{stats['total']} = {rate:.1f}%")
    
    total = sum(s['total'] for s in strategy_stats.values())
    correct = sum(s['correct'] for s in strategy_stats.values())
    print(f"\n总体准确率: {correct}/{total} = {correct/max(total,1)*100:.1f}%")
    
    # 保存Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "预测结果"
    
    headers = ['编号', '对阵', '主胜概率', '平局概率', '客胜概率', '置信度', '胜率差', '8变化', '盘型', '策略', '排除项', '预测', '实际', '结果', '理由', '澳门推荐']
    ws.append(headers)
    
    for r in results:
        match_id = r['编号']
        actual = actual_results.get(match_id, '')
        is_correct = 'O' if r['预测'] == actual else ('X' if actual else '-')
        
        ws.append([
            r['编号'], r['对阵'], r['主胜概率'], r['平局概率'], r['客胜概率'],
            r['置信度'], r['胜率差'], r['8变化'], r['盘型'], r['策略'],
            r['排除项'], r['预测'], actual, is_correct, r['理由'], r['澳门推荐']
        ])
        
        row_idx = ws.max_row
        if is_correct == 'O':
            for col in range(1, len(headers)+1):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='C6EFCE')
        elif is_correct == 'X':
            for col in range(1, len(headers)+1):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='FFC7CE')
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 22)
    
    wb.save(OUTPUT_FILE)
    print(f"\n结果已保存到: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
