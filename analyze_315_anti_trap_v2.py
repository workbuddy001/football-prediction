import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 读取预测数据
df = pd.read_excel('d:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx')

# 转换概率列
def to_float(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

df['主胜概率'] = df['主胜概率'].apply(to_float)
df['平局概率'] = df['平局概率'].apply(to_float)
df['客胜概率'] = df['客胜概率'].apply(to_float)

# 实际比赛结果
actual_results = {
    '周日001': ('日本女', 1), '周日002': ('福冈黄蜂', 1), '周日003': ('首尔FC', 0),
    '周日004': ('仁川联', 2), '周日005': ('墨胜利', 1), '周日006': ('乌德勒支', 0),
    '周日007': ('热那亚', 0), '周日008': ('汉诺威96', 2), '周日009': ('马洛卡', 1),
    '周日010': ('费耶诺德', 1), '周日011': ('布兰', 0), '周日012': ('水晶宫', 2),
    '周日013': ('诺丁汉', 2), '周日014': ('曼联', 1), '周日015': ('比萨', 1),
    '周日016': ('博洛尼亚', 0), '周日017': ('美因茨', 0), '周日018': ('巴萨', 1),
    '周日019': ('瓦勒伦加', 1), '周日020': ('里昂', 2), '周日021': ('利物浦', 2),
    '周日022': ('柏林联合', 0), '周日023': ('科莫', 1), '周日024': ('贝蒂斯', 2),
    '周日025': ('斯图加特', 1), '周日026': ('拉齐奥', 1), '周日027': ('皇家社会', 1),
    '周日028': ('波尔图', 1), '周日029': ('温哥华', 1),
}

# 获取预测选项
def get_pred_result(row):
    if '主胜' in str(row['首选']):
        return 0, '主胜'
    elif '客胜' in str(row['首选']):
        return 1, '客胜'
    return 2, '平局'

def get_max_prob(row):
    probs = [row['主胜概率'], row['平局概率'], row['客胜概率']]
    idx = probs.index(max(probs))
    return ['主胜', '平局', '客胜'][idx], max(probs)

# 分析每场比赛
results = []
for idx, row in df.iterrows():
    match_id = row['编号'].split('_')[0]
    pred_code, pred_str = get_pred_result(row)
    max_prob_str, max_prob = get_max_prob(row)
    panxing = row['盘型']  # 诱盘/实盘
    
    # 获取实际结果
    actual = actual_results.get(match_id)
    if not actual:
        continue
    actual_code = actual[1]
    actual_str = ['主胜', '客胜', '平局'][actual_code]
    
    # 判断原预测是否正确
    is_correct_orig = pred_code == actual_code
    
    # 计算置信度
    confidence = max_prob
    
    # 反诱盘分析
    anti_trap = ""
    trap_reason = ""
    is_correct_anti = False
    
    if panxing == '诱盘' and confidence > 50:
        # 诱盘 + 高置信度 = 需要反诱盘
        # 推荐选择概率最低的选项
        probs = {'主胜': row['主胜概率'], '平局': row['平局概率'], '客胜': row['客胜概率']}
        min_option = min(probs, key=probs.get)
        anti_trap = min_option
        trap_reason = f"诱盘陷阱!原预测{max_prob_str}({max_prob:.1f}%)可能是庄家诱饵"
        is_correct_anti = (min_option == actual_str)
    elif panxing == '实盘' and confidence <= 50:
        # 实盘 + 低置信度 = 谨慎对待
        anti_trap = "谨慎"
        trap_reason = "实盘但低置信度，建议观望"
        is_correct_anti = False
    elif confidence > 55 and panxing == '实盘':
        # 高置信度 + 实盘 = 可靠推荐
        anti_trap = "推荐"
        trap_reason = f"高置信度({max_prob:.1f}%)实盘，可信度高"
        is_correct_anti = is_correct_orig
    elif confidence <= 40:
        anti_trap = "观望"
        trap_reason = "置信度太低，不建议投注"
        is_correct_anti = False
    else:
        anti_trap = "正常"
        trap_reason = "按正常分析执行"
        is_correct_anti = is_correct_orig
    
    results.append({
        '编号': match_id,
        '对阵': row['对阵'].replace(' | ', ' vs ').replace(' |', ''),
        '盘型': panxing,
        '原预测': pred_str,
        '置信度': f"{max_prob:.1f}%",
        '实际': actual_str,
        '原对错': '对' if is_correct_orig else '错',
        '反诱盘推荐': anti_trap,
        '反诱盘对错': '对' if is_correct_anti else '错',
        '原因': trap_reason
    })

result_df = pd.DataFrame(results)

# 统计
print("=" * 90)
print("3.15 反诱盘分析结果")
print("=" * 90)

# 原预测准确率
orig_correct = (result_df['原对错'] == '对').sum()
print(f"\n【原预测准确率】{orig_correct}/29 = {orig_correct/29*100:.1f}%")

# 按反诱盘推荐类型统计
print("\n【反诱盘推荐效果分析】")
recommend_analysis = []
for rec_type in result_df['反诱盘推荐'].unique():
    subset = result_df[result_df['反诱盘推荐'] == rec_type]
    anti_correct = (subset['反诱盘对错'] == '对').sum()
    orig_correct = (subset['原对错'] == '对').sum()
    recommend_analysis.append({
        '推荐类型': rec_type,
        '场次': len(subset),
        '原预测正确': orig_correct,
        '反诱盘正确': anti_correct,
        '原准确率': f"{orig_correct/len(subset)*100:.1f}%" if len(subset) > 0 else "0%",
        '反诱盘准确率': f"{anti_correct/len(subset)*100:.1f}%" if len(subset) > 0 else "0%"
    })

analysis_df = pd.DataFrame(recommend_analysis)
print(analysis_df.to_string(index=False))

# 重点：高置信度诱盘的反诱盘效果
print("\n【重点：高置信度诱盘的反诱盘效果】")
high_conf_trap = result_df[(result_df['盘型'] == '诱盘') & (result_df['反诱盘推荐'].isin(['主胜', '客胜', '平局']))]
if len(high_conf_trap) > 0:
    anti_correct = (high_conf_trap['反诱盘对错'] == '对').sum()
    orig_correct = (high_conf_trap['原对错'] == '对').sum()
    print(f"共{len(high_conf_trap)}场:")
    print(f"  原预测准确: {orig_correct}场 ({orig_correct/len(high_conf_trap)*100:.1f}%)")
    print(f"  反诱盘准确: {anti_correct}场 ({anti_correct/len(high_conf_trap)*100:.1f}%)")

# 详细表格
print("\n" + "=" * 90)
print("详细分析")
print("=" * 90)
print(result_df[['编号', '盘型', '原预测', '置信度', '实际', '原对错', '反诱盘推荐', '反诱盘对错', '原因']].to_string(index=False))

# 保存Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "反诱盘分析"

# 标题行
headers = ['编号', '盘型', '原预测', '置信度', '实际', '原对错', '反诱盘推荐', '反诱盘对错', '原因']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', start_color='366092')
    cell.font = Font(bold=True, color='FFFFFF')

# 数据行
for row_idx, r in enumerate(results, 2):
    ws.cell(row_idx, 1, r['编号'])
    ws.cell(row_idx, 2, r['盘型'])
    ws.cell(row_idx, 3, r['原预测'])
    ws.cell(row_idx, 4, r['置信度'])
    ws.cell(row_idx, 5, r['实际'])
    ws.cell(row_idx, 6, r['原对错'])
    
    # 反诱盘推荐列根据类型着色
    rec_cell = ws.cell(row_idx, 7, r['反诱盘推荐'])
    if r['反诱盘推荐'] == '推荐':
        rec_cell.fill = PatternFill('solid', start_color='00FF00')
    elif r['反诱盘推荐'] in ['主胜', '客胜', '平局']:
        rec_cell.fill = PatternFill('solid', start_color='FF6600')
    elif r['反诱盘推荐'] == '观望':
        rec_cell.fill = PatternFill('solid', start_color='FFFF00')
    
    ws.cell(row_idx, 8, r['反诱盘对错'])
    ws.cell(row_idx, 9, r['原因'])
    
    # 原对错着色
    orig_cell = ws.cell(row_idx, 6)
    if r['原对错'] == '对':
        orig_cell.font = Font(color='00AA00')
    else:
        orig_cell.font = Font(color='FF0000')
    
    # 反诱盘对错着色
    anti_cell = ws.cell(row_idx, 8)
    if r['反诱盘对错'] == '对':
        anti_cell.font = Font(color='00AA00')
    else:
        anti_cell.font = Font(color='FF0000')

# 列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 35

wb.save('3.15_反诱盘分析.xlsx')
print("\n已保存: 3.15_反诱盘分析.xlsx")
