# -*- coding: utf-8 -*-
"""
排除法策略分析脚本 v1.0
基于3.15复盘发现的必杀技:
- 诱盘 + 高置信度 → 反诱盘(选最低) → 53.8%
- 实盘 + 低置信度 → 排除法(选最低) → 100%
- 高置信度 + 实盘 → 正向 → 100%
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json
from pathlib import Path

# ==================== 配置 ====================
DATA_FILE = 'd:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx'
JSON_FILE = 'd:/work/workbuddy/足球预测/分析模板/matches_full_2026-03-15.json'
OUTPUT_FILE = '3.16_排除法预测.xlsx'

# 置信度阈值
HIGH_CONF_THRESHOLD = 50  # 高置信度阈值
LOW_CONF_THRESHOLD = 40   # 低置信度阈值

# ==================== 数据处理 ====================
def to_float(x):
    """转换百分比字符串为浮点数"""
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

def load_match_data():
    """加载比赛数据"""
    df = pd.read_excel(DATA_FILE)
    
    # 转换概率列
    df['主胜概率'] = df['主胜概率'].apply(to_float)
    df['平局概率'] = df['平局概率'].apply(to_float)
    df['客胜概率'] = df['客胜概率'].apply(to_float)
    
    return df

def load_json_data():
    """加载JSON原始数据"""
    try:
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            matches = json.load(f)
        return matches
    except:
        return []

# ==================== 排除法策略 ====================
def apply_exclusion_strategy(row, probs):
    """
    应用排除法策略
    
    策略逻辑:
    1. 诱盘 + 高置信度 → 反诱盘(选最低)
    2. 实盘 + 低置信度 → 排除法(选最低)
    3. 高置信度 + 实盘 → 正向
    4. 其他情况 → 观望
    """
    panxing = str(row.get('盘型', '')).strip()
    confidence = max(probs.values())  # 最高概率
    
    # 获取三个选项的概率
    home_prob = probs['主胜']
    draw_prob = probs['平局']
    away_prob = probs['客胜']
    
    # 找到概率最低的选项(排除法核心!)
    min_option = min(probs, key=probs.get)
    min_prob = probs[min_option]
    
    # 找到概率最高的选项
    max_option = max(probs, key=probs.get)
    
    # 获取原预测(从"首选"列)
    first_choice = str(row.get('首选', '')).strip()
    if '主胜' in first_choice:
        orig_pred = '主胜'
    elif '客胜' in first_choice:
        orig_pred = '客胜'
    elif '平局' in first_choice:
        orig_pred = '平局'
    else:
        orig_pred = max_option
    
    strategy = ""
    final_pick = ""
    reason = ""
    
    # 策略1: 诱盘 + 高置信度 → 反诱盘
    if panxing == '诱盘' and confidence > HIGH_CONF_THRESHOLD:
        strategy = "诱盘高置信→反诱盘"
        final_pick = min_option  # 选择最低概率
        reason = f"诱盘时庄家故意制造混乱，最低概率反而是答案"
    
    # 策略2: 实盘 + 低置信度 → 排除法
    elif panxing == '实盘' and confidence <= LOW_CONF_THRESHOLD:
        strategy = "实盘低置信→排除法"
        final_pick = min_option  # 选择最低概率
        reason = f"实盘低置信时庄家无明确倾向，排除法100%命中"
    
    # 策略3: 高置信度 + 实盘 → 正向
    elif panxing == '实盘' and confidence > HIGH_CONF_THRESHOLD:
        strategy = "高置信实盘→正向"
        final_pick = max_option  # 选择最高概率
        reason = f"实盘高置信时正向选择"
    
    # 策略4: 其他情况
    else:
        if confidence > 55:
            strategy = "高置信→正向"
            final_pick = max_option
            reason = f"高置信度正向"
        elif confidence <= 35:
            strategy = "极低置信→排除"
            final_pick = min_option
            reason = f"极低置信度使用排除法"
        else:
            strategy = "观望"
            final_pick = "观望"
            reason = f"不符合明确策略条件"
    
    return {
        'strategy': strategy,
        'final_pick': final_pick,
        'reason': reason,
        'confidence': confidence,
        'min_option': min_option,
        'max_option': max_option,
        'orig_pred': orig_pred
    }

# ==================== 主程序 ====================
def main():
    print("=" * 80)
    print("排除法策略分析 v1.0")
    print("=" * 80)
    
    # 加载数据
    df = load_match_data()
    print(f"\n加载了 {len(df)} 场比赛")
    
    # 实际结果(3.15) - 从复盘分析文档获取
    actual_results = {
        '周日001': '主胜', '周日002': '主胜', '周日003': '客胜', '周日004': '平局',
        '周日005': '主胜', '周日006': '客胜', '周日007': '客胜', '周日008': '平局',
        '周日009': '主胜', '周日010': '主胜', '周日011': '主胜', '周日012': '平局',
        '周日013': '平局', '周日014': '主胜', '周日015': '主胜', '周日016': '客胜',
        '周日017': '客胜', '周日018': '主胜', '周日019': '主胜', '周日020': '平局',
        '周日021': '平局', '周日022': '客胜', '周日023': '主胜', '周日024': '平局',
        '周日025': '主胜', '周日026': '主胜', '周日027': '主胜', '周日028': '主胜',
        '周日029': '主胜',
    }
    
    results = []
    
    for idx, row in df.iterrows():
        match_id = str(row.get('编号', '')).split('_')[0]
        
        # 获取概率
        probs = {
            '主胜': row['主胜概率'],
            '平局': row['平局概率'],
            '客胜': row['客胜概率']
        }
        
        # 应用排除法策略
        strategy_result = apply_exclusion_strategy(row, probs)
        
        # 获取实际结果
        actual = actual_results.get(match_id, '')
        
        # 检查是否命中
        is_correct = strategy_result['final_pick'] == actual if actual else None
        
        results.append({
            '编号': match_id,
            '对阵': str(row.get('对阵', '')).split('|')[-1].strip() if '|' in str(row.get('对阵', '')) else str(row.get('对阵', '')),
            '盘型': row.get('盘型', ''),
            '主胜概率': f"{probs['主胜']:.1f}%",
            '平局概率': f"{probs['平局']:.1f}%",
            '客胜概率': f"{probs['客胜']:.1f}%",
            '置信度': f"{strategy_result['confidence']:.1f}%",
            '原预测': strategy_result['orig_pred'],
            '排除项(最低)': strategy_result['min_option'],
            '策略': strategy_result['strategy'],
            '最终选择': strategy_result['final_pick'],
            '实际': actual,
            '结果': 'O' if is_correct else ('X' if is_correct is False else '-'),
            '理由': strategy_result['reason']
        })
    
    # 创建结果DataFrame
    result_df = pd.DataFrame(results)
    
    # ==================== 统计输出 ====================
    print("\n" + "=" * 100)
    print("完整预测结果")
    print("=" * 100)
    
    # 打印表格
    print(f"\n{'编号':<8} {'对阵':<25} {'盘型':<6} {'置信度':<8} {'排除项':<8} {'策略':<20} {'最终选择':<8} {'实际':<6} {'结果':<4}")
    print("-" * 100)
    
    for r in results:
        print(f"{r['编号']:<8} {r['对阵'][:22]:<22} {r['盘型']:<6} {r['置信度']:<8} {r['排除项(最低)']:<8} {r['策略'][:18]:<18} {r['最终选择']:<8} {r['实际']:<6} {r['结果']}")
    
    # ==================== 策略统计 ====================
    print("\n" + "=" * 80)
    print("策略统计")
    print("=" * 80)
    
    # 按策略分组统计
    strategy_stats = {}
    for r in results:
        strategy = r['策略']
        if strategy not in strategy_stats:
            strategy_stats[strategy] = {'total': 0, 'correct': 0}
        strategy_stats[strategy]['total'] += 1
        if r['结果'] == 'O':
            strategy_stats[strategy]['correct'] += 1
    
    print(f"\n{'策略':<25} {'场次':<8} {'正确':<8} {'准确率':<10}")
    print("-" * 60)
    
    for strategy, stats in sorted(strategy_stats.items(), key=lambda x: -x[1]['correct']):
        acc = stats['correct'] / stats['total'] * 100 if stats['total'] > 0 else 0
        print(f"{strategy:<25} {stats['total']:<8} {stats['correct']:<8} {acc:.1f}%")
    
    # ==================== 核心策略验证 ====================
    print("\n" + "=" * 80)
    print("【核心策略验证】排除法效果")
    print("=" * 80)
    
    # 筛选使用排除法的比赛
    exclusion_matches = [r for r in results if '排除' in r['策略']]
    if exclusion_matches:
        excl_correct = sum(1 for r in exclusion_matches if r['结果'] == 'O')
        print(f"\n排除法共 {len(exclusion_matches)} 场: 正确 {excl_correct} 场, 准确率 {excl_correct/len(exclusion_matches)*100:.1f}%")
    
    # 筛选反诱盘
    anti_trap = [r for r in results if '反诱盘' in r['策略']]
    if anti_trap:
        anti_correct = sum(1 for r in anti_trap if r['结果'] == 'O')
        print(f"反诱盘共 {len(anti_trap)} 场: 正确 {anti_correct} 场, 准确率 {anti_correct/len(anti_trap)*100:.1f}%")
    
    # 总体准确率
    valid_results = [r for r in results if r['结果'] != '-']
    total_correct = sum(1 for r in valid_results if r['结果'] == 'O')
    print(f"\n总体准确率: {total_correct}/{len(valid_results)} = {total_correct/len(valid_results)*100:.1f}%")
    
    # ==================== 保存Excel ====================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排除法预测"
    
    # 写入表头
    headers = ['编号', '对阵', '盘型', '主胜概率', '平局概率', '客胜概率', 
              '置信度', '原预测', '排除项(最低)', '策略', '最终选择', '实际', '结果', '理由']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = PatternFill('solid', start_color='366092')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # 写入数据
    for row_idx, r in enumerate(results, 2):
        ws.cell(row_idx, 1, r['编号'])
        ws.cell(row_idx, 2, r['对阵'])
        ws.cell(row_idx, 3, r['盘型'])
        ws.cell(row_idx, 4, r['主胜概率'])
        ws.cell(row_idx, 5, r['平局概率'])
        ws.cell(row_idx, 6, r['客胜概率'])
        ws.cell(row_idx, 7, r['置信度'])
        ws.cell(row_idx, 8, r['原预测'])
        ws.cell(row_idx, 9, r['排除项(最低)'])
        ws.cell(row_idx, 10, r['策略'])
        ws.cell(row_idx, 11, r['最终选择'])
        ws.cell(row_idx, 12, r['实际'])
        ws.cell(row_idx, 13, r['结果'])
        ws.cell(row_idx, 14, r['理由'])
        
        # 颜色标记
        if r['结果'] == 'O':
            for col in range(1, 15):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='C6EFCE')
        elif r['结果'] == 'X':
            for col in range(1, 15):
                ws.cell(row_idx, col).fill = PatternFill('solid', start_color='FFC7CE')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 8
    ws.column_dimensions['M'].width = 6
    ws.column_dimensions['N'].width = 35
    
    wb.save(OUTPUT_FILE)
    print(f"\n已保存: {OUTPUT_FILE}")
    
    return result_df

if __name__ == '__main__':
    main()
