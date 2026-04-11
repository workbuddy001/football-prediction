# 3.15比赛把握度分析
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import ast

def parse_315_source(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    info = {}
    filename = filepath.stem
    match = re.match(r'周日(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        info['match_id'] = f"周日{match.group(1)}"
        info['home_team'] = match.group(2)
        info['away_team'] = match.group(3)
    
    # 近况
    match = re.search(r'主队.*?近况走势.*?([WL]+)', content)
    if match:
        form = match.group(1).upper()
        info['home_w'] = form.count('W')
        info['home_l'] = form.count('L')
    
    match = re.search(r'客队.*?近况走势.*?([WL]+)', content)
    if match:
        form = match.group(1).upper()
        info['away_w'] = form.count('W')
        info['away_l'] = form.count('L')
    
    # 赔率
    initial_odds = []
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            initial_odds = ast.literal_eval(re.sub(r'#.*', '', tuple_str))
        except:
            pass
    
    realtime_odds = []
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            realtime_odds = ast.literal_eval(re.sub(r'#.*', '', tuple_str))
        except:
            pass
    
    info['initial_odds'] = initial_odds
    info['realtime_odds'] = realtime_odds
    
    return info

def analyze_with_confidence(info):
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "把握度": "低", "详情": []}
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_home = sum(realtime_home) / len(realtime_home)
    avg_draw = sum(realtime_draw) / len(realtime_draw)
    avg_away = sum(realtime_away) / len(realtime_away)
    
    prob_home = 1/avg_home / (1/avg_home + 1/avg_draw + 1/avg_away)
    prob_draw = 1/avg_draw / (1/avg_home + 1/avg_draw + 1/avg_away)
    prob_away = 1/avg_away / (1/avg_home + 1/avg_draw + 1/avg_away)
    
    initial_home = [x[0] for x in initial]
    initial_away = [x[2] for x in initial]
    home_change = (avg_home - sum(initial_home)/len(initial_home)) / (sum(initial_home)/len(initial_home)) * 100
    
    details = []
    prediction = ""
    confidence = "D"
    
    if avg_away < 1.5:
        prediction = f"{away}客胜"
        details.append("强胆客")
        confidence = "A"
    elif avg_home < 1.5:
        prediction = f"{home}主胜"
        details.append("强胆主")
        confidence = "A"
    elif 1.5 < avg_home < 2.0 and home_w >= home_l and prob_home > 0.45:
        prediction = f"{home}主胜"
        details.append(f"强主场")
        confidence = "B"
    elif 1.5 < avg_away < 2.0 and away_w >= away_l and prob_away > 0.45:
        prediction = f"{away}客胜"
        details.append(f"强客场")
        confidence = "B"
    elif prob_home > 0.55:
        prediction = f"{home}主胜"
        details.append(f"高概率")
        confidence = "B"
    elif prob_away > 0.55:
        prediction = f"{away}客胜"
        details.append(f"高概率")
        confidence = "B"
    elif prob_home > prob_away and prob_home > 0.35:
        prediction = f"{home}主胜"
        details.append(f"概率偏向")
        confidence = "C"
    elif prob_away > prob_home and prob_away > 0.35:
        prediction = f"{away}客胜"
        details.append(f"概率偏向")
        confidence = "C"
    else:
        if prob_home > prob_away:
            prediction = f"{home}主胜"
        elif prob_away > prob_home:
            prediction = f"{away}客胜"
        else:
            prediction = "平局"
        details.append("默认")
    
    return {
        "预测": prediction,
        "把握度": confidence,
        "详情": details
    }

# 分析3.15
from pathlib import Path

folder = Path("分析模板/3.15")
files = sorted(folder.glob("*.md"))

results = []
for f in files:
    info = parse_315_source(f)
    if not info.get('home_team') or not info.get('initial_odds'):
        continue
    result = analyze_with_confidence(info)
    results.append({
        '编号': info.get('match_id'),
        '对阵': f"{info.get('home_team')} vs {info.get('away_team')}",
        '预测': result['预测'],
        '把握度': result['把握度'],
        '详情': ', '.join(result['详情'])
    })

# 3.15实际结果
actual_15 = {
    "周日001": "主胜", "周日003": "客胜", "周日004": "平局", "周日006": "客胜",
    "周日007": "客胜", "周日008": "平局", "周日009": "主胜", "周日010": "主胜",
    "周日011": "主胜", "周日012": "平局", "周日013": "平局", "周日014": "主胜",
    "周日015": "主胜", "周日016": "客胜", "周日017": "客胜", "周日018": "主胜",
    "周日019": "主胜", "周日020": "平局", "周日021": "平局", "周日022": "客胜",
    "周日023": "主胜", "周日024": "平局", "周日025": "主胜", "周日026": "主胜",
    "周日027": "主胜", "周日028": "客胜", "周日029": "主胜",
}

def check_match(pred, actual):
    pred = str(pred).lower()
    if "主胜" in pred:
        return "主胜" == actual
    elif "客胜" in pred:
        return "客胜" == actual
    elif "平局" in pred:
        return "平局" == actual
    return False

conf_a = [r for r in results if r['把握度'] == 'A']
conf_b = [r for r in results if r['把握度'] == 'B']
conf_c = [r for r in results if r['把握度'] == 'C']
conf_d = [r for r in results if r['把握度'] == 'D']

print("=" * 60)
print("3.15 比赛把握度分析 (对比)")
print("=" * 60)

print("\n【把握度A - 高把握】")
if conf_a:
    c = sum(1 for r in conf_a if check_match(r['预测'], actual_15.get(r['编号'], "")))
    print(f"  共{len(conf_a)}场, 准确{c}场, 准确率{c/len(conf_a)*100:.0f}%")

print("\n【把握度B - 中高把握】")
if conf_b:
    c = sum(1 for r in conf_b if check_match(r['预测'], actual_15.get(r['编号'], "")))
    print(f"  共{len(conf_b)}场, 准确{c}场, 准确率{c/len(conf_b)*100:.0f}%")

print("\n【把握度C - 中把握】")
if conf_c:
    c = sum(1 for r in conf_c if check_match(r['预测'], actual_15.get(r['编号'], "")))
    print(f"  共{len(conf_c)}场, 准确{c}场, 准确率{c/len(conf_c)*100:.0f}%")

# 重点推荐
high_conf = conf_a + conf_b
high_correct = sum(1 for r in high_conf if check_match(r['预测'], actual_15.get(r['编号'], "")))
print(f"\n【重点推荐 A+B】共{len(high_conf)}场, 准确{high_correct}场, 准确率{high_correct/len(high_conf)*100:.0f}%")

# 总体
all_c = sum(1 for r in results if check_match(r['预测'], actual_15.get(r['编号'], "")))
print(f"总体准确率: {all_c}/29 = {all_c/29*100:.1f}%")

# 保存
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.15_把握度"

for col, h in enumerate(['编号', '对阵', '预测', '把握度', '详情'], 1):
    ws.cell(1, col, h)

for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['预测'])
    ws.cell(row, 4, r['把握度'])
    ws.cell(row, 5, r['详情'])

wb.save("3.15_把握度分析.xlsx")
print("\n已保存: 3.15_把握度分析.xlsx")
