# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

data = [
    {'编号': '周五001', '对阵': '布里斯班 vs 西悉尼', '主队状态': '10%', '客队状态': '30%', '状态差': '-20%', '主胜': 2.63, '平局': 3.43, '客胜': 2.54, '预测': '平局', '把握度': 'D', '理由': '澳门推荐平局'},
    {'编号': '周五002', '对阵': 'Austral女 vs 朝鲜女', '主队状态': '60%', '客队状态': '70%', '状态差': '-10%', '主胜': 2.11, '平局': 3.11, '客胜': 3.17, '预测': '主胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五003', '对阵': '马格德堡 vs 达姆施塔', '主队状态': '40%', '客队状态': '50%', '状态差': '-10%', '主胜': 2.34, '平局': 3.72, '客胜': 2.66, '预测': '主胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五004', '对阵': '胡巴尔 vs 吉达国民', '主队状态': '70%', '客队状态': '80%', '状态差': '-10%', '主胜': 2.33, '平局': 3.37, '客胜': 2.79, '预测': '主胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五005', '对阵': '克莱蒙 vs 波城FC', '主队状态': '30%', '客队状态': '20%', '状态差': '+10%', '主胜': 2.38, '平局': 3.29, '客胜': 2.75, '预测': '主胜', '把握度': 'C', '理由': '诱盘防冷'},
    {'编号': '周五006', '对阵': '兹沃勒 vs 格罗宁根', '主队状态': '20%', '客队状态': '20%', '状态差': '0%', '主胜': 2.67, '平局': 3.53, '客胜': 2.47, '预测': '客胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五007', '对阵': '坎布尔 vs 罗达JC', '主队状态': '80%', '客队状态': '30%', '状态差': '+50%', '主胜': 1.75, '平局': 3.85, '客胜': 3.89, '预测': '主胜', '把握度': 'B', '理由': '状态差距大+低赔=实盘'},
    {'编号': '周五008', '对阵': '门兴 vs 圣保利', '主队状态': '20%', '客队状态': '30%', '状态差': '-10%', '主胜': 1.94, '平局': 3.33, '客胜': 4.13, '预测': '主胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五009', '对阵': '都灵 vs 帕尔马', '主队状态': '30%', '客队状态': '40%', '状态差': '-10%', '主胜': 2.24, '平局': 2.97, '客胜': 3.62, '预测': '主胜', '把握度': 'C', '理由': '诱盘防冷'},
    {'编号': '周五010', '对阵': '马赛 vs 欧塞尔', '主队状态': '40%', '客队状态': '10%', '状态差': '+30%', '主胜': 1.43, '平局': 4.62, '客胜': 7.05, '预测': '主胜', '把握度': 'D', '理由': '概率最高'},
    {'编号': '周五011', '对阵': '雷克斯 vs 斯旺西', '主队状态': '60%', '客队状态': '60%', '状态差': '0%', '主胜': 2.26, '平局': 3.32, '客胜': 3.06, '预测': '主胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
    {'编号': '周五012', '对阵': '阿拉维斯 vs 比利亚雷', '主队状态': '30%', '客队状态': '40%', '状态差': '-10%', '主胜': 3.38, '平局': 3.36, '客胜': 2.15, '预测': '客胜', '把握度': 'C', '理由': '状态相近+平赔稳定=实盘'},
]

wb = Workbook()
ws = wb.active
ws.title = "V3预测3.13"

headers = ['编号', '对阵', '主队状态', '客队状态', '状态差', '主胜', '平局', '客胜', '预测', '把握度', '理由']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

for row_idx, d in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=d['编号'])
    ws.cell(row=row_idx, column=2, value=d['对阵'])
    ws.cell(row=row_idx, column=3, value=d['主队状态'])
    ws.cell(row=row_idx, column=4, value=d['客队状态'])
    ws.cell(row=row_idx, column=5, value=d['状态差'])
    ws.cell(row=row_idx, column=6, value=d['主胜'])
    ws.cell(row=row_idx, column=7, value=d['平局'])
    ws.cell(row=row_idx, column=8, value=d['客胜'])
    ws.cell(row=row_idx, column=9, value=d['预测'])
    ws.cell(row=row_idx, column=10, value=d['把握度'])
    ws.cell(row=row_idx, column=11, value=d['理由'])

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 8
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 35

for row_idx in range(2, len(data) + 2):
    conf = ws.cell(row=row_idx, column=10).value
    if conf == 'B':
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

wb.save('3.13_V3预测.xlsx')
print("已保存到 3.13_V3预测.xlsx")
