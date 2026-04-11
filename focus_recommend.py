import openpyxl

# 读取3.14分析结果
wb = openpyxl.load_workbook('3.14_把握度分析.xlsx')
ws = wb.active

print('=' * 70)
print('3.14 重点推荐比赛 (把握度A+B)')
print('=' * 70)

results_14 = []
for row in range(2, 34):
    id_ = ws.cell(row, 1).value
    match = ws.cell(row, 2).value
    pred = ws.cell(row, 3).value
    conf = ws.cell(row, 4).value
    results_14.append({'id': id_, 'match': match, 'pred': pred, 'conf': conf})

actual_14 = {
    '周六001': '平局', '周六002': '客胜', '周六003': '平局', '周六004': '客胜',
    '周六005': '主胜', '周六006': '主胜', '周六007': '平局', '周六008': '主胜',
    '周六009': '主胜', '周六010': '客胜', '周六011': '主胜', '周六012': '平局',
    '周六013': '平局', '周六014': '主胜', '周六015': '主胜', '周六016': '平局',
    '周六017': '平局', '周六018': '主胜', '周六019': '主胜', '周六020': '主胜',
    '周六021': '主胜', '周六022': '主胜', '周六023': '客胜', '周六024': '平局',
    '周六025': '主胜', '周六026': '客胜', '周六027': '客胜', '周六028': '客胜',
    '周六029': '平局', '周六030': '主胜', '周六031': '客胜', '周六032': '平局',
}

high_conf = [r for r in results_14 if r['conf'] in ['A', 'B']]
print(f'共{len(high_conf)}场重点推荐\n')

ok_count = 0
for r in high_conf:
    actual = actual_14.get(r['id'], '')
    is_ok = (('主胜' in r['pred'] and actual == '主胜') or
             ('客胜' in r['pred'] and actual == '客胜') or
             ('平局' in r['pred'] and actual == '平局'))
    status = 'OK' if is_ok else 'X'
    if is_ok: ok_count += 1
    star = 'A' if r['conf'] == 'A' else 'B'
    print(f'[{star}] {status} {r["id"]}: {r["pred"]:10s} vs {actual:4s}')

print(f'\n复盘结果: {ok_count}/{len(high_conf)} = {ok_count/len(high_conf)*100:.0f}%')
