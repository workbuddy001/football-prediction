# 3.14比赛批量分析脚本 - 使用V4算法
import os
import re
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# 读取源数据文件
def parse_source_file(filepath):
    """解析源数据文件"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    match_info = {}
    
    # 比赛对阵 - 查找文件名
    filename = Path(filepath).stem
    # 周六002_名古屋鲸vs神户胜利_源数据
    match = re.match(r'周六(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        match_info['match_id'] = f"周六{match.group(1)}"
        match_info['home_team'] = match.group(2)
        match_info['away_team'] = match.group(3)
    
    # 比赛时间
    match = re.search(r'比赛时间.*?\|.*?(\d{4}-\d{2}-\d{2}\s*\d{1,2}:\d{2})', content)
    if match:
        match_info['match_time'] = match.group(1)
    
    # 赛事
    match = re.search(r'赛事.*?\|.*?(\S+)', content)
    if match:
        match_info['league'] = match.group(1)
    
    # 主队近况
    match = re.search(r'主队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['home_w'] = int(match.group(1))
        match_info['home_d'] = int(match.group(2))
        match_info['home_l'] = int(match.group(3))
    
    # 客队近况
    match = re.search(r'客队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['away_w'] = int(match.group(1))
        match_info['away_d'] = int(match.group(2))
        match_info['away_l'] = int(match.group(3))
    
    # 澳门推荐
    match = re.search(r'澳门推荐.*?\|.*?(\S+)', content)
    if match:
        macao_tip = match.group(1)
        match_info['macao_tip'] = macao_tip
    
    # 提取Python格式的赔率数据
    initial_odds = []
    realtime_odds = []
    
    # 提取初盘赔率
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            # 提取元组列表
            tuple_str = '[' + match.group(1) + ']'
            # 清理注释
            tuple_str = re.sub(r'#.*', '', tuple_str)
            initial_odds = ast.literal_eval(tuple_str)
        except Exception as e:
            print(f"解析初盘失败: {e}")
    
    # 提取即时赔率
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            realtime_odds = ast.literal_eval(tuple_str)
        except Exception as e:
            print(f"解析即时失败: {e}")
    
    match_info['initial_odds'] = initial_odds
    match_info['realtime_odds'] = realtime_odds
    
    return match_info

# V4分析算法
def analyze_v4(info):
    """V4算法分析"""
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    macao_tip = info.get('macao_tip', '')
    
    # 近况
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "盘型": "未知", "详情": []}
    
    # 计算统计
    initial_home = [x[0] for x in initial]
    initial_draw = [x[1] for x in initial]
    initial_away = [x[2] for x in initial]
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_initial_home = sum(initial_home) / len(initial_home)
    avg_initial_draw = sum(initial_draw) / len(initial_draw)
    avg_initial_away = sum(initial_away) / len(initial_away)
    
    avg_realtime_home = sum(realtime_home) / len(realtime_home)
    avg_realtime_draw = sum(realtime_draw) / len(realtime_draw)
    avg_realtime_away = sum(realtime_away) / len(realtime_away)
    
    # 即时概率
    prob_home = 1 / avg_realtime_home / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_draw = 1 / avg_realtime_draw / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_away = 1 / avg_realtime_away / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    
    # 变化百分比
    home_pct_change = (avg_realtime_home - avg_initial_home) / avg_initial_home * 100
    draw_pct_change = (avg_realtime_draw - avg_initial_draw) / avg_initial_draw * 100
    away_pct_change = (avg_realtime_away - avg_initial_away) / avg_initial_away * 100
    
    # 升赔公司数
    home_up = sum(1 for i in range(len(initial)) if realtime[i][0] > initial[i][0])
    draw_up = sum(1 for i in range(len(initial)) if realtime[i][1] > initial[i][1])
    away_up = sum(1 for i in range(len(initial)) if realtime[i][2] > initial[i][2])
    
    home_up_pct = home_up / len(initial) * 100
    away_up_pct = away_up / len(initial) * 100
    draw_down = len(initial) - draw_up
    
    # 澳门赔率 (取第一个公司的赔率作为澳门参考)
    macao_home = realtime[0][0] if realtime else avg_realtime_home
    macao_away = realtime[0][2] if realtime else avg_realtime_away
    
    details = []
    prediction = ""
    
    # 规则1: 主胜<1.5
    if avg_realtime_home < 1.5:
        prediction = f"{home}主胜"
        details.append("规则1: 主胜<1.5")
    
    # 规则2: 客胜<1.5
    elif avg_realtime_away < 1.5:
        prediction = f"{away}客胜"
        details.append("规则2: 客胜<1.5")
    
    # 规则3: 强队主场
    elif 1.5 < avg_realtime_home < 2.0 and home_w >= home_l:
        prediction = f"{home}主胜"
        details.append("规则3: 强队主场+近况好")
    
    # 规则4: 强队客场
    elif 1.5 < avg_realtime_away < 2.0 and away_w >= away_l:
        prediction = f"{away}客胜"
        details.append("规则4: 强队客场+近况好")
    
    # 规则5: 主胜升>15%
    elif home_pct_change > 15 and home_up_pct > 80:
        prediction = f"{away}客胜"
        details.append(f"规则5: 主胜升{home_pct_change:.1f}%")
    
    # 规则6: 客胜升>15%
    elif away_pct_change > 15 and away_up_pct > 70:
        prediction = f"{home}主胜"
        details.append(f"规则6: 客胜升{away_pct_change:.1f}%")
    
    # 规则7: 主队近况好
    elif home_w >= 3 and avg_realtime_home < 2.5:
        prediction = f"{home}主胜"
        details.append(f"规则7: 主队W{home_w}场")
    
    # 规则8: 客队近况好
    elif away_w >= 3 and avg_realtime_away < 2.5:
        prediction = f"{away}客胜"
        details.append(f"规则8: 客队W{away_w}场")
    
    # 规则9: 澳门客胜低
    elif macao_away < avg_realtime_away * 0.85:
        prediction = f"{away}客胜"
        details.append("规则9: 澳门客胜低")
    
    # 规则10: 澳门主胜低
    elif macao_home < avg_realtime_home * 0.90:
        prediction = f"{home}主胜"
        details.append("规则10: 澳门主胜低")
    
    # 规则11: 澳门推荐主队
    elif macao_tip and home in macao_tip:
        prediction = f"{home}主胜"
        details.append("规则11: 澳门推荐主队")
    
    # 规则12: 澳门推荐客队
    elif macao_tip and away in macao_tip:
        prediction = f"{away}客胜"
        details.append("规则12: 澳门推荐客队")
    
    # 规则13: 平局条件
    elif draw_pct_change < -2 and draw_up > len(initial) * 0.6 and prob_draw > 0.35:
        prediction = "平局"
        details.append(f"规则13: 平局降{abs(draw_pct_change):.1f}%")
    
    # 规则14: 主胜不稳
    elif home_pct_change > 5 and away_pct_change < 0:
        prediction = f"{away}客胜"
        details.append("规则14: 主胜不稳")
    
    # 规则15: 默认概率最高
    else:
        if prob_home > prob_away and prob_home > prob_draw:
            prediction = f"{home}主胜"
            details.append("规则15: 主胜概率最高")
        elif prob_away > prob_home and prob_away > prob_draw:
            prediction = f"{away}客胜"
            details.append("规则15: 客胜概率最高")
        else:
            prediction = "平局"
            details.append("规则15: 平局概率最高")
    
    # 盘型判断
    if home_up_pct > 80 and draw_down > len(initial) * 0.5:
        panxing = "实盘"
    else:
        panxing = "诱盘"
    
    return {
        "预测": prediction,
        "盘型": panxing,
        "详情": details,
        "主胜概率": f"{prob_home*100:.1f}%",
        "平局概率": f"{prob_draw*100:.1f}%",
        "客胜概率": f"{prob_away*100:.1f}%",
        "主胜": f"{avg_realtime_home:.2f}",
        "平局": f"{avg_realtime_draw:.2f}",
        "客胜": f"{avg_realtime_away:.2f}"
    }

# 主程序
folder = "分析模板/3.14"
files = sorted(Path(folder).glob("*.md"))

results = []

for f in files:
    print(f"处理: {f.name}")
    info = parse_source_file(f)
    
    if not info.get('home_team'):
        print(f"  - 跳过: {f.name}")
        continue
    
    if not info.get('initial_odds'):
        print(f"  - 无赔率数据: {f.name}")
        continue
    
    result = analyze_v4(info)
    
    results.append({
        '编号': info.get('match_id', f.stem.split('_')[0]),
        '对阵': f"{info.get('home_team', '')} vs {info.get('away_team', '')}",
        '赛事': info.get('league', ''),
        '预测': result['预测'],
        '盘型': result['盘型'],
        '主胜': result.get('主胜', ''),
        '平局': result.get('平局', ''),
        '客胜': result.get('客胜', ''),
        '主胜概率': result.get('主胜概率', ''),
        '平局概率': result.get('平局概率', ''),
        '客胜概率': result.get('客胜概率', ''),
        '详情': ', '.join(result.get('详情', []))
    })

print(f"\n共分析 {len(results)} 场比赛")

# 生成Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.14预测汇总"

# 标题行
headers = ['编号', '对阵', '赛事', '预测', '盘型', '主胜', '平局', '客胜', '主胜概率', '平局概率', '客胜概率', '详情']
for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal='center')

# 数据行
for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['赛事'])
    ws.cell(row, 4, r['预测'])
    ws.cell(row, 5, r['盘型'])
    ws.cell(row, 6, r['主胜'])
    ws.cell(row, 7, r['平局'])
    ws.cell(row, 8, r['客胜'])
    ws.cell(row, 9, r['主胜概率'])
    ws.cell(row, 10, r['平局概率'])
    ws.cell(row, 11, r['客胜概率'])
    ws.cell(row, 12, r['详情'])

# 调整列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
ws.column_dimensions['K'].width = 10
ws.column_dimensions['L'].width = 40

wb.save("3.14_比赛预测汇总.xlsx")
print("已保存: 3.14_比赛预测汇总.xlsx")
