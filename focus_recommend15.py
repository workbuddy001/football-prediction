import openpyxl

# 3.15重点推荐
wb15 = openpyxl.load_workbook('3.15_把握度分析.xlsx')
ws15 = wb15.active

print('=' * 70)
print('3.15 重点推荐比赛 (把握度A+B)')
print('=' * 70)

results_15 = []
for row in range(2, 31):
    id_ = ws15.cell(row, 1).value
    match = ws15.cell(row, 2).value
    pred = ws15.cell(row, 3).value
    conf = ws15.cell(row, 4).value
    results_15.append({'id': id_, 'match': match, 'pred': pred, 'conf': conf})

actual_15 = {
    '周日001': '主胜', '周日003': '客胜', '周日004': '平局', '周日006': '客胜',
    '周日007': '客胜', '周日008': '平局', '周日009': '主胜', '周日010': '主胜',
    '周日011': '主胜', '周日012': '平局', '周日013': '平局', '周日014': '主胜',
    '周日015': '主胜', '周日016': '客胜', '周日017': '客胜', '周日018': '主胜',
    '周日019': '主胜', '周日020': '平局', '周日021': '平局', '周日022': '客胜',
    '周日023': '主胜', '周日024': '平局', '周日025': '主胜', '周日026': '主胜',
    '周日027': '主胜', '周日028': '客胜', '周日029': '主胜',
}

high_conf15 = [r for r in results_15 if r['conf'] in ['A', 'B']]
print(f'共{len(high_conf15)}场重点推荐\n')

ok_count15 = 0
for r in high_conf15:
    actual = actual_15.get(r['id'], '')
    is_ok = (('主胜' in r['pred'] and actual == '主胜') or
             ('客胜' in r['pred'] and actual == '客胜') or
             ('平局' in r['pred'] and actual == '平局'))
    status = 'OK' if is_ok else 'X'
    if is_ok: ok_count15 += 1
    star = 'A' if r['conf'] == 'A' else 'B'
    print(f'[{star}] {status} {r["id"]}: {r["pred"]:10s} vs {actual:4s}')

print(f'\n复盘结果: {ok_count15}/{len(high_conf15)} = {ok_count15/len(high_conf15)*100:.0f}%')
