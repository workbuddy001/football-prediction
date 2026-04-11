import openpyxl
wb = openpyxl.load_workbook('3.14_比赛预测汇总.xlsx')
ws = wb.active
print('|编号|对阵|赛事|预测|盘型|详情|')
print('|---|---|---|---|---|---|')
for row in range(2, 34):
    id_ = ws.cell(row, 1).value
    match = ws.cell(row, 2).value
    league = ws.cell(row, 3).value
    pred = ws.cell(row, 4).value
    pan = ws.cell(row, 5).value
    detail = ws.cell(row, 12).value[:35] if ws.cell(row, 12).value else ''
    print(f'|{id_}|{match}|{league}|{pred}|{pan}|{detail}|')
