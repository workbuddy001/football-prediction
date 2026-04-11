# 直接分析错误案例
import openpyxl

# 读取把握度分析结果
wb14 = openpyxl.load_workbook('3.14_把握度分析.xlsx')
ws14 = wb14.active

actual_14 = {
    '周六001': '平局', '周六002': '客胜', '周六003': '平局', '周六004': '客胜',
    '周六005': '主胜', '周六006': '主胜', '周六007': '平局', '周六008': '主胜',
    '周六009': '主胜', '周六010': '客胜', '周六011': '主胜', '周六012': '平局',
    '周六013': '平局', '周六014': '主胜', '周六015': '主胜', '周六016': '平局',
    '周六017': '平局', '周六018': '主胜', '周六019': '主胜', '周六020': '主胜',
    '周六021': '主胜', '周六022': '主胜', '周六023': '客胜', '周六024': '平局',
    '周六025': '主胜', '周六026': '客胜', '周六027': '客胜', '周六028': '客胜',
    '周六029': '平局', '周六030': '主胜', '周六031': '客胜', '周六032': '平局',
}

actual_15 = {
    '周日001': '主胜', '周日003': '客胜', '周日004': '平局', '周日006': '客胜',
    '周日007': '客胜', '周日008': '平局', '周日009': '主胜', '周日010': '主胜',
    '周日011': '主胜', '周日012': '平局', '周日013': '平局', '周日014': '主胜',
    '周日015': '主胜', '周六016': '客胜', '周六017': '客胜', '周日018': '主胜',
    '周日019': '主胜', '周日020': '平局', '周日021': '平局', '周日022': '客胜',
    '周日023': '主胜', '周日024': '平局', '周日025': '主胜', '周日026': '主胜',
    '周日027': '主胜', '周日028': '客胜', '周日029': '主胜',
}

print("=" * 70)
print("错误案例详细分析")
print("=" * 70)

# 分析3.14错误
print("\n【3.14错误案例分析】")
errors_14 = []
for row in range(2, 34):
    id_ = ws14.cell(row, 1).value
    pred = ws14.cell(row, 3).value
    conf = ws14.cell(row, 4).value
    home_odds = ws14.cell(row, 6).value
    draw_odds = ws14.cell(row, 7).value
    away_odds = ws14.cell(row, 8).value
    
    actual = actual_14.get(id_, '')
    
    # 判断是否错误
    is_correct = False
    if pred and actual:
        if ('主胜' in pred and actual == '主胜') or \
           ('客胜' in pred and actual == '客胜') or \
           ('平局' in pred and actual == '平局'):
            is_correct = True
    
    if not is_correct and actual:
        try:
            errors_14.append({
                'id': id_,
                'pred': pred,
                'conf': conf,
                'actual': actual,
                'home_odds': float(home_odds.replace('%','')) if home_odds else 0,
                'draw_odds': float(draw_odds.replace('%','')) if draw_odds else 0,
                'away_odds': float(away_odds.replace('%','')) if away_odds else 0,
            })
        except:
            pass

print(f"\n共{len(errors_14)}个错误:")
print(f"{'编号':<8} {'预测':<12} {'把握度':<4} {'实际':<6} {'主胜%':<8} {'平局%':<8} {'客胜%':<8}")
print("-" * 60)
for e in errors_14:
    print(f"{e['id']:<8} {e['pred']:<12} {e['conf']:<4} {e['actual']:<6} {e['home_odds']:<8.0f} {e['draw_odds']:<8.0f} {e['away_odds']:<8.0f}")

# 规律总结
print("\n【错误规律总结】")
print("-" * 50)

# 1. 强队客场
strong_away_errors = [e for e in errors_14 if '客胜' in e['pred'] and e['away_odds'] < 50]
print(f"1. 强队客场(客胜概率<50%)错误: {len(strong_away_errors)}场")

# 2. 高概率主胜
high_home_errors = [e for e in errors_14 if '主胜' in e['pred'] and e['home_odds'] > 40]
print(f"2. 高概率主胜(>40%)错误: {len(high_home_errors)}场")

# 3. 平局问题
draw_actual = [e for e in errors_14 if e['actual'] == '平局']
print(f"3. 实际平局被错判: {len(draw_actual)}场")

# 4. 把握度分析
conf_a_errors = [e for e in errors_14 if e['conf'] == 'A']
conf_b_errors = [e for e in errors_14 if e['conf'] == 'B']
print(f"4. A类把握度错误: {len(conf_a_errors)}场")
print(f"5. B类把握度错误: {len(conf_b_errors)}场")

# 优化建议
print("\n【优化建议】")
print("-" * 50)
print("1. 强队客场(客胜<2.0) - 改为防平/防主胜")
print("2. 高概率主胜(>45%)但信心不足时 - 适当防平")
print("3. 平局特征: 降赔公司>50% + 概率>28%")
print("4. A类错误多为: 强队低赔但不胜")
