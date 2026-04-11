# V7算法 - 基于规则分析优化
# 结论:
# - 强胆客: 100% (保留)
# - 强胆主: 57% (保留)
# - 强主场: 57% (保留)
# - 强客场: 0% (删除/反转)
# - 主升15: 0% (删除)
# - 主W3: 44% (保留但降低权重)
# - 客W3: 36% (保留但降低权重)
# - 主概率: 50% (保留)
# - 客概率: 36% (保留但降低权重)
# - 主不稳: 15% (删除)

import os
import re
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

def analyze_v7(info):
    """V7算法 - 基于规则分析优化"""
    home = info.get('home_team', '主队')
    away = info.get('away_team', '客队')
    initial = info.get('initial_odds', [])
    realtime = info.get('realtime_odds', [])
    
    home_w = info.get('home_w', 0)
    home_l = info.get('home_l', 0)
    away_w = info.get('away_w', 0)
    away_l = info.get('away_l', 0)
    
    if not initial or not realtime:
        return {"预测": "数据不足", "盘型": "未知", "详情": []}
    
    initial_home = [x[0] for x in initial]
    initial_draw = [x[1] for x in initial]
    initial_away = [x[2] for x in initial]
    
    realtime_home = [x[0] for x in realtime]
    realtime_draw = [x[1] for x in realtime]
    realtime_away = [x[2] for x in realtime]
    
    avg_realtime_home = sum(realtime_home) / len(realtime_home)
    avg_realtime_draw = sum(realtime_draw) / len(realtime_draw)
    avg_realtime_away = sum(realtime_away) / len(realtime_away)
    
    prob_home = 1 / avg_realtime_home / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_draw = 1 / avg_realtime_draw / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    prob_away = 1 / avg_realtime_away / (1/avg_realtime_home + 1/avg_realtime_draw + 1/avg_realtime_away)
    
    home_pct_change = (avg_realtime_home - sum(initial_home)/len(initial_home)) / (sum(initial_home)/len(initial_home)) * 100
    away_pct_change = (avg_realtime_away - sum(initial_away)/len(initial_away)) / (sum(initial_away)/len(initial_away)) * 100
    draw_pct_change = (avg_realtime_draw - sum(initial_draw)/len(initial_draw)) / (sum(initial_draw)/len(initial_draw)) * 100
    
    draw_up = sum(1 for i in range(len(initial)) if realtime[i][1] > initial[i][1])
    draw_down_pct = (len(initial) - draw_up) / len(initial) * 100
    
    details = []
    prediction = ""
    
    # ===== V7优化规则 =====
    
    # 规则1: 强胆客 (100%准确率!)
    if avg_realtime_away < 1.5:
        prediction = f"{away}客胜"
        details.append("强胆客")
    
    # 规则2: 强胆主 (57%准确率)
    elif avg_realtime_home < 1.5:
        prediction = f"{home}主胜"
        details.append("强胆主")
    
    # 规则3: 强主场 (57%准确率)
    elif 1.5 < avg_realtime_home < 2.0 and home_w >= home_l:
        prediction = f"{home}主胜"
        details.append("强主场")
    
    # 规则4: 强客场 - 改为预测主胜或平局 (原规则0%准确率)
    elif 1.5 < avg_realtime_away < 2.0:
        # 强客场不一定是客胜，可能出冷门
        if prob_draw > 0.30:
            prediction = "平局"
            details.append("强客场防冷")
        else:
            prediction = f"{home}主胜"
            details.append("强客场防冷")
    
    # 规则5: 客胜大升>15% - 删除或反转
    # 原规则20%准确率，改为预测主胜
    
    # 规则6: 概率优先 - 使用概率最高的选项
    elif prob_home > prob_away and prob_home > prob_draw:
        prediction = f"{home}主胜"
        details.append(f"主概率{prob_home*100:.0f}%")
    elif prob_away > prob_home and prob_away > prob_draw:
        prediction = f"{away}客胜"
        details.append(f"客概率{prob_away*100:.0f}%")
    else:
        prediction = "平局"
        details.append(f"平概率{prob_draw*100:.0f}%")
    
    panxing = "实盘" if draw_down_pct > 50 else "诱盘"
    
    return {
        "预测": prediction,
        "盘型": panxing,
        "详情": details,
    }

def parse_source_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    match_info = {}
    
    filename = Path(filepath).stem
    match = re.match(r'周六(\d+)_(.+?)vs(.+?)_源数据', filename)
    if match:
        match_info['match_id'] = f"周六{match.group(1)}"
        match_info['home_team'] = match.group(2)
        match_info['away_team'] = match.group(3)
    
    match = re.search(r'主队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['home_w'] = int(match.group(1))
        match_info['home_l'] = int(match.group(3))
    
    match = re.search(r'客队近况.*?(\d+)胜(\d+)平(\d+)负', content)
    if match:
        match_info['away_w'] = int(match.group(1))
        match_info['away_l'] = int(match.group(3))
    
    initial_odds = []
    realtime_odds = []
    
    match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            initial_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if match:
        try:
            tuple_str = '[' + match.group(1) + ']'
            tuple_str = re.sub(r'#.*', '', tuple_str)
            realtime_odds = ast.literal_eval(tuple_str)
        except:
            pass
    
    match_info['initial_odds'] = initial_odds
    match_info['realtime_odds'] = realtime_odds
    
    return match_info

# 测试V7
folder = "分析模板/3.14"
files = sorted(Path(folder).glob("*.md"))
results = []

for f in files:
    info = parse_source_file(f)
    if not info.get('home_team') or not info.get('initial_odds'):
        continue
    result = analyze_v7(info)
    results.append({
        '编号': info.get('match_id'),
        '对阵': f"{info.get('home_team')} vs {info.get('away_team')}",
        '预测': result['预测'],
        '详情': ', '.join(result['详情'])
    })

actual_14 = {
    "周六001": "平局", "周六002": "客胜", "周六003": "平局", "周六004": "客胜",
    "周六005": "主胜", "周六006": "主胜", "周六007": "平局", "周六008": "主胜",
    "周六009": "主胜", "周六010": "客胜", "周六011": "主胜", "周六012": "平局",
    "周六013": "平局", "周六014": "主胜", "周六015": "主胜", "周六016": "平局",
    "周六017": "平局", "周六018": "主胜", "周六019": "主胜", "周六020": "主胜",
    "周六021": "主胜", "周六022": "主胜", "周六023": "客胜", "周六024": "平局",
    "周六025": "主胜", "周六026": "客胜", "周六027": "客胜", "周六028": "客胜",
    "周六029": "平局", "周六030": "主胜", "周六031": "客胜", "周六032": "平局",
}

def check_match(pred, actual):
    pred = str(pred).lower()
    if "主胜" in pred:
        return "主胜" == actual
    elif "客胜" in pred:
        return "客胜" == actual
    elif "平局" in pred:
        return "平局" == actual
    return False

correct14 = 0
for r in results:
    actual = actual_14.get(r['编号'], "")
    if check_match(r['预测'], actual):
        correct14 += 1

print(f"V7算法 - 3.14准确率: {correct14/32*100:.1f}% ({correct14}/32)")

# 保存结果
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3.14_V7预测"

for col, header in enumerate(['编号', '对阵', '预测', '详情'], 1):
    cell = ws.cell(1, col, header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")

for row, r in enumerate(results, 2):
    ws.cell(row, 1, r['编号'])
    ws.cell(row, 2, r['对阵'])
    ws.cell(row, 3, r['预测'])
    ws.cell(row, 4, r['详情'])

wb.save("3.14_V7预测.xlsx")
print("已保存: 3.14_V7预测.xlsx")
