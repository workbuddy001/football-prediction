import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 读取预测数据
df = pd.read_excel('d:/work/workbuddy/足球预测/3.15_比赛预测汇总.xlsx')

# 转换概率列
def to_float(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    return float(x)

df['主胜概率'] = df['主胜概率'].apply(to_float)
df['平局概率'] = df['平局概率'].apply(to_float)
df['客胜概率'] = df['客胜概率'].apply(to_float)

# 实际比赛结果
actual_results = {
    '周日001': ('日本女', 1), '周日002': ('福冈黄蜂', 1), '周日003': ('首尔FC', 0),
    '周日004': ('仁川联', 2), '周日005': ('墨胜利', 1), '周日006': ('乌德勒支', 0),
    '周日007': ('热那亚', 0), '周日008': ('汉诺威96', 2), '周日009': ('马洛卡', 1),
    '周日010': ('费耶诺德', 1), '周日011': ('布兰', 0), '周日012': ('水晶宫', 2),
    '周日013': ('诺丁汉', 2), '周日014': ('曼联', 1), '周日015': ('比萨', 1),
    '周日016': ('博洛尼亚', 0), '周日017': ('美因茨', 0), '周日018': ('巴萨', 1),
    '周日019': ('瓦勒伦加', 1), '周日020': ('里昂', 2), '周日021': ('利物浦', 2),
    '周日022': ('柏林联合', 0), '周日023': ('科莫', 1), '周日024': ('贝蒂斯', 2),
    '周日025': ('斯图加特', 1), '周日026': ('拉齐奥', 1), '周日027': ('皇家社会', 1),
    '周日028': ('波尔图', 1), '周日029': ('温哥华', 1),
}

# 获取预测选项
def get_pred_result(row):
    if '主胜' in str(row['首选']):
        return 0, '主胜'
    elif '客胜' in str(row['首选']):
        return 1, '客胜'
    return 2, '平局'

def get_max_prob(row):
    probs = [row['主胜概率'], row['平局概率'], row['客胜概率']]
    idx = probs.index(max(probs))
    return ['主胜', '平局', '客胜'][idx], max(probs)

# 分析每场比赛
results = []
for idx, row in df.iterrows():
    match_id = row['编号'].split('_')[0]
    pred_code, pred_str = get_pred_result(row)
    max_prob_str, max_prob = get_max_prob(row)
    panxing = row['盘型']  # 诱盘/实盘
    
    # 获取实际结果
    actual = actual_results.get(match_id)
    if not actual:
        continue
    actual_code = actual[1]
    actual_str = ['主胜', '客胜', '平局'][actual_code]
    
    # 判断是否正确
    is_correct = 1 if pred_code == actual_code else 0
    correct_str = "对" if is_correct else "错"
    
    # 计算置信度
    confidence = max_prob
    
    # 反诱盘分析
    anti_trap = ""
    trap_reason = ""
    
    if panxing == '诱盘' and confidence > 50:
        # 诱盘 + 高置信度 = 需要反诱盘
        # 推荐选择概率最低的选项
        probs = {'主胜': row['主胜概率'], '平局': row['平局概率'], '客胜': row['客胜概率']}
        min_option = min(probs, key=probs.get)
        anti_trap = min_option
        trap_reason = f"诱盘陷阱!原预测{max_prob_str}({max_prob:.1f}%)可能是庄家诱饵"
    elif panxing == '实盘' and confidence <= 50:
        # 实盘 + 低置信度 = 谨慎对待
        anti_trap = "谨慎"
        trap_reason = "实盘但低置信度，建议观望"
    elif confidence > 55 and panxing == '实盘':
        # 高置信度 + 实盘 = 可靠推荐
        anti_trap = "推荐"
        trap_reason = f"高置信度({max_prob:.1f}%)实盘，可信度高"
    elif confidence <= 40:
        anti_trap = "观望"
        trap_reason = "置信度太低，不建议投注"
    else:
        anti_trap = "正常"
        trap_reason = "按正常分析执行"
    
    results.append({
        '编号': match_id,
        '对阵': row['对阵'].replace(' | ', ' vs ').replace(' |', ''),
        '盘型': panxing,
        '原预测': pred_str,
        '置信度': f"{max_prob:.1f}%",
        '实际': actual_str,
        '对错': correct_str,
        '反诱盘推荐': anti_trap,
        '原因': trap_reason
    })

result_df = pd.DataFrame(results)

# 统计
print("=" * 80)
print("3.15 反诱盘分析结果")
print("=" * 80)

# 按推荐分类
print("\n【反诱盘推荐汇总】")
recommend_groups = result_df.groupby('反诱盘推荐').agg({
    '编号': 'count',
    '对错': lambda x: (x == '✓').sum()
}).reset_index()
recommend_groups.columns = ['推荐类型', '场次', '正确']
recommend_groups['准确率'] = (recommend_groups['正确'] / recommend_groups['场次'] * 100).round(1)
print(recommend_groups.to_string(index=False))

# 重点：反诱盘推荐的准确率
anti_trap = result_df[result_df['反诱盘推荐'] != '正常']
if len(anti_trap) > 0:
    anti_correct = (anti_trap['对错'] == '✓').sum()
    print(f"\n【反诱盘建议】共{len(anti_trap)}场, 正确{anti_correct}场, 准确率{anti_correct/len(anti_trap)*100:.1f}%")

# 详细表格
print("\n" + "=" * 80)
print("详细分析")
print("=" * 80)
print(result_df[['编号', '对阵', '盘型', '原预测', '置信度', '实际', '对错', '反诱盘推荐', '原因']].to_string(index=False))

# 保存Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "反诱盘分析"

# 标题行
headers = ['编号', '对阵', '盘型', '原预测', '置信度', '实际', '对错', '反诱盘推荐', '原因']
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', start_color='366092')
    cell.font = Font(bold=True, color='FFFFFF')

# 数据行
for row_idx, r in enumerate(results, 2):
    ws.cell(row_idx, 1, r['编号'])
    ws.cell(row_idx, 2, r['对阵'])
    ws.cell(row_idx, 3, r['盘型'])
    ws.cell(row_idx, 4, r['原预测'])
    ws.cell(row_idx, 5, r['置信度'])
    ws.cell(row_idx, 6, r['实际'])
    ws.cell(row_idx, 7, r['对错'])
    
    # 反诱盘推荐列根据类型着色
    rec_cell = ws.cell(row_idx, 8, r['反诱盘推荐'])
    if r['反诱盘推荐'] == '推荐':
        rec_cell.fill = PatternFill('solid', start_color='00FF00')
    elif r['反诱盘推荐'] in ['主胜', '客胜', '平局']:
        rec_cell.fill = PatternFill('solid', start_color='FF6600')
    elif r['反诱盘推荐'] == '观望':
        rec_cell.fill = PatternFill('solid', start_color='FFFF00')
    
    ws.cell(row_idx, 9, r['原因'])
    
    # 对错着色
    result_cell = ws.cell(row_idx, 7)
    if r['对错'] == '✓':
        result_cell.font = Font(color='00AA00')
    else:
        result_cell.font = Font(color='FF0000')

# 列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 6
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 35

wb.save('3.15_反诱盘分析.xlsx')
print("\n已保存: 3.15_反诱盘分析.xlsx")
