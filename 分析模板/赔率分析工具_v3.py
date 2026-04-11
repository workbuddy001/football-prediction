import os
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extract_odds_from_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    home_team = re.search(r'主队\s*\|\s*(.+)', content)
    away_team = re.search(r'客队\s*\|\s*(.+)', content)
    match_time = re.search(r'比赛时间\s*\|\s*(.+)', content)
    league = re.search(r'赛事\s*\|\s*(.+)', content)
    home_form = re.search(r'主队近况走势\s*\|\s*(.+)', content)
    away_form = re.search(r'客队近况走势\s*\|\s*(.+)', content)
    home_handicap = re.search(r'主队盘路走势\s*\|\s*(.+)', content)
    away_handicap = re.search(r'客队盘路走势\s*\|\s*(.+)', content)
    history = re.search(r'历史交锋\s*\|\s*(.+)', content)
    macao_tip = re.search(r'澳门推荐\s*\|\s*(.+)', content)
    
    init_match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if init_match:
        odds_str = init_match.group(1)
        initial_odds = eval('[' + odds_str + ']')
    else:
        initial_odds = []
    
    real_match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if real_match:
        odds_str = real_match.group(1)
        realtime_odds = eval('[' + odds_str + ']')
    else:
        realtime_odds = []
    
    return {
        'home_team': home_team.group(1).strip() if home_team else '',
        'away_team': away_team.group(1).strip() if away_team else '',
        'match_time': match_time.group(1).strip() if match_time else '',
        'league': league.group(1).strip() if league else '',
        'home_form': home_form.group(1).strip() if home_form else '',
        'away_form': away_form.group(1).strip() if away_form else '',
        'home_handicap': home_handicap.group(1).strip() if home_handicap else '',
        'away_handicap': away_handicap.group(1).strip() if away_handicap else '',
        'history': history.group(1).strip() if history else '',
        'macao_tip': macao_tip.group(1).strip() if macao_tip else '',
        'initial_odds': initial_odds,
        'realtime_odds': realtime_odds,
    }

def find_macao_odds(odds_list):
    for i, odds in enumerate(odds_list):
        home, draw, away = odds
        if 1.0 <= home <= 1.05 and 8 <= draw <= 20 and 15 <= away <= 30:
            return odds, i
    return odds_list[0] if odds_list else (0, 0, 0), 0

def analyze_match_v3(data):
    if not data['initial_odds'] or not data['realtime_odds']:
        return None
    
    init_home = [o[0] for o in data['initial_odds']]
    init_draw = [o[1] for o in data['initial_odds']]
    init_away = [o[2] for o in data['initial_odds']]
    
    real_home = [o[0] for o in data['realtime_odds']]
    real_draw = [o[1] for o in data['realtime_odds']]
    real_away = [o[2] for o in data['realtime_odds']]
    
    # 变化百分比
    home_pct = [(data['realtime_odds'][i][0] - data['initial_odds'][i][0]) / data['initial_odds'][i][0] * 100 
                for i in range(len(data['initial_odds']))]
    draw_pct = [(data['realtime_odds'][i][1] - data['initial_odds'][i][1]) / data['initial_odds'][i][1] * 100 
                for i in range(len(data['initial_odds']))]
    away_pct = [(data['realtime_odds'][i][2] - data['initial_odds'][i][2]) / data['initial_odds'][i][2] * 100 
                for i in range(len(data['initial_odds']))]
    
    # 概率
    real_home_prob = [1/x*100 for x in real_home]
    real_draw_prob = [1/x*100 for x in real_draw]
    real_away_prob = [1/x*100 for x in real_away]
    
    # 统计
    home_up_pct = sum(1 for x in home_pct if x > 0) / len(home_pct) * 100
    home_down_pct = sum(1 for x in home_pct if x < 0) / len(home_pct) * 100
    draw_up_pct = sum(1 for x in draw_pct if x > 0) / len(draw_pct) * 100
    draw_down_pct = sum(1 for x in draw_pct if x < 0) / len(draw_pct) * 100
    away_up_pct = sum(1 for x in away_pct if x > 0) / len(away_pct) * 100
    away_down_pct = sum(1 for x in away_pct if x < 0) / len(away_pct) * 100
    
    # 平均值
    avg_home = np.mean(real_home)
    avg_draw = np.mean(real_draw)
    avg_away = np.mean(real_away)
    
    avg_home_prob = np.mean(real_home_prob)
    avg_draw_prob = np.mean(real_draw_prob)
    avg_away_prob = np.mean(real_away_prob)
    
    # 澳门
    macao, macao_idx = find_macao_odds(data['realtime_odds'])
    macao_home, macao_draw, macao_away = macao
    
    # 近况
    def count_wins(form):
        if not form:
            return 0
        return sum(1 for c in form.upper() if c == 'W')
    
    home_wins = count_wins(data['home_form'])
    away_wins = count_wins(data['away_form'])
    
    # ========== V3算法：更激进的判断 ==========
    
    # 规则1: 强队主场 (主胜<1.5) → 主胜
    if avg_home < 1.5:
        first_choice = f"{data['home_team']}主胜"
        first_prob = f"{avg_home_prob:.0f}%"
        reason = "强队主场"
    # 规则2: 强队客场 (客胜<1.5) → 客胜
    elif avg_away < 1.5:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{avg_away_prob:.0f}%"
        reason = "强队客场"
    # 规则3: 主胜被明显看衰 (升>15%) 且 上升公司>80% → 客胜
    elif np.mean(home_pct) > 15 and home_up_pct > 80:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{avg_away_prob:.0f}%"
        reason = "主胜被大幅看衰"
    # 规则4: 客胜被明显看衰 (升>15%) 且 上升公司>70% → 主胜
    elif np.mean(away_pct) > 15 and away_up_pct > 70:
        first_choice = f"{data['home_team']}主胜"
        first_prob = f"{avg_home_prob:.0f}%"
        reason = "客胜被大幅看衰"
    # 规则5: 主队主场近况好 (W>=3) 且 主胜赔率<2.5 → 主胜
    elif home_wins >= 3 and avg_home < 2.5:
        first_choice = f"{data['home_team']}主胜"
        first_prob = f"{avg_home_prob:.0f}%"
        reason = "主队近况好+主场"
    # 规则6: 客队近况好 (W>=3) 且 客胜赔率<2.5 → 客胜
    elif away_wins >= 3 and avg_away < 2.5:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{avg_away_prob:.0f}%"
        reason = "客队近况好"
    # 规则7: 澳门客胜明显低于市场 → 客胜
    elif macao_away < avg_away * 0.85:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{avg_away_prob:.0f}%"
        reason = "澳门客胜偏低"
    # 规则8: 澳门主胜明显低于市场 → 主胜
    elif macao_home < avg_home * 0.9:
        first_choice = f"{data['home_team']}主胜"
        first_prob = f"{avg_home_prob:.0f}%"
        reason = "澳门主胜偏低"
    # 规则9: 平局降赔>60% 且 平局概率>32% → 平局 (严格条件)
    elif draw_down_pct > 60 and avg_draw_prob > 32:
        first_choice = "平局"
        first_prob = f"{avg_draw_prob:.0f}%"
        reason = "平局防范很强"
    # 规则10: 主胜变化大但客胜稳定 → 客胜
    elif home_up_pct > 60 and away_down_pct > 40:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{avg_away_prob:.0f}%"
        reason = "主胜不稳+客胜受保护"
    # 默认: 选概率最高，不轻易选平局
    else:
        if avg_home_prob >= avg_away_prob + 3:  # 主胜概率明显高
            first_choice = f"{data['home_team']}主胜"
            first_prob = f"{avg_home_prob:.0f}%"
        elif avg_away_prob >= avg_home_prob + 3:  # 客胜概率明显高
            first_choice = f"{data['away_team']}客胜"
            first_prob = f"{avg_away_prob:.0f}%"
        elif avg_draw_prob > 35:  # 平局概率特别高
            first_choice = "平局"
            first_prob = f"{avg_draw_prob:.0f}%"
        elif avg_home_prob >= avg_away_prob:
            first_choice = f"{data['home_team']}主胜"
            first_prob = f"{avg_home_prob:.0f}%"
        else:
            first_choice = f"{data['away_team']}客胜"
            first_prob = f"{avg_away_prob:.0f}%"
        reason = "概率最高"
    
    # 实盘判断
    is_real = home_up_pct > 60 and draw_down_pct > 40
    panxing = "实盘" if is_real else "诱盘"
    
    draw_possible = "高" if draw_down_pct > 55 and avg_draw_prob > 30 else \
                   "中" if draw_down_pct > 40 or avg_draw_prob > 25 else "低"
    
    return {
        'home_team': data['home_team'],
        'away_team': data['away_team'],
        'league': data['league'],
        'macao_tip': data['macao_tip'],
        'home_form': data['home_form'],
        'away_form': data['away_form'],
        'panxing': panxing,
        'draw_possible': draw_possible,
        'first_choice': first_choice,
        'first_prob': first_prob,
        'reason': reason,
        'real_home_prob': f"{avg_home_prob:.1f}%",
        'real_draw_prob': f"{avg_draw_prob:.1f}%",
        'real_away_prob': f"{avg_away_prob:.1f}%",
        'home_pct_avg': f"{np.mean(home_pct):.1f}%",
        'away_pct_avg': f"{np.mean(away_pct):.1f}%",
    }

# 处理文件
folder = "分析模板/3.15"
files = [f for f in os.listdir(folder) if f.endswith('_源数据.md')]

results = []
for f in sorted(files):
    filepath = os.path.join(folder, f)
    try:
        data = extract_odds_from_file(filepath)
        result = analyze_match_v3(data)
        if result:
            result['filename'] = f.replace('_源数据.md', '')
            results.append(result)
    except Exception as e:
        print(f"Error: {f} - {e}")

# 生成Excel
wb = Workbook()
ws = wb.active
ws.title = "比赛预测汇总"

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
center = Alignment(horizontal='center', vertical='center')

headers = ['编号', '赛事', '对阵', '澳门推荐', '盘型', '平局可能', '首选', '概率', '依据']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

for r, data in enumerate(results, 2):
    ws.cell(row=r, column=1, value=data['filename'])
    ws.cell(row=r, column=2, value=data['league'])
    ws.cell(row=r, column=3, value=f"{data['home_team']} vs {data['away_team']}")
    ws.cell(row=r, column=4, value=data['macao_tip'])
    ws.cell(row=r, column=5, value=data['panxing'])
    ws.cell(row=r, column=6, value=data['draw_possible'])
    ws.cell(row=r, column=7, value=data['first_choice'])
    ws.cell(row=r, column=8, value=data['first_prob'])
    ws.cell(row=r, column=9, value=data['reason'])
    
    for c in range(1, 10):
        ws.cell(row=r, column=c).alignment = center

for col in range(1, 10):
    ws.column_dimensions[chr(64+col)].width = 14

wb.save('3.15_比赛预测汇总_v3.xlsx')
print(f"Done: 3.15_比赛预测汇总_v3.xlsx")
