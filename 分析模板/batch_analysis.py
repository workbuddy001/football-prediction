import os
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extract_odds_from_file(filepath):
    """从源数据文件提取赔率数据"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 提取比赛信息
    home_team = re.search(r'主队\s*\|\s*(.+)', content)
    away_team = re.search(r'客队\s*\|\s*(.+)', content)
    match_time = re.search(r'比赛时间\s*\|\s*(.+)', content)
    league = re.search(r'赛事\s*\|\s*(.+)', content)
    home_form = re.search(r'主队近况走势\s*\|\s*(.+)', content)
    away_form = re.search(r'客队近况走势\s*\|\s*(.+)', content)
    home_handicap = re.search(r'主队盘路走势\s*\|\s*(.+)', content)
    away_handicap = re.search(r'客队盘路走势\s*\|\s*(.+)', content)
    history = re.search(r'历史交锋\s*\|\s*(.+)', content)
    macao_tip = re.search(r'澳门推荐\s*\|\s*(.+)', content)
    
    # 提取初盘赔率
    init_match = re.search(r'initial_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if init_match:
        odds_str = init_match.group(1)
        initial_odds = eval('[' + odds_str + ']')
    else:
        initial_odds = []
    
    # 提取即时赔率
    real_match = re.search(r'realtime_odds\s*=\s*\[(.*?)\]', content, re.DOTALL)
    if real_match:
        odds_str = real_match.group(1)
        realtime_odds = eval('[' + odds_str + ']')
    else:
        realtime_odds = []
    
    return {
        'home_team': home_team.group(1).strip() if home_team else '',
        'away_team': away_team.group(1).strip() if away_team else '',
        'match_time': match_time.group(1).strip() if match_time else '',
        'league': league.group(1).strip() if league else '',
        'home_form': home_form.group(1).strip() if home_form else '',
        'away_form': away_form.group(1).strip() if away_form else '',
        'home_handicap': home_handicap.group(1).strip() if home_handicap else '',
        'away_handicap': away_handicap.group(1).strip() if away_handicap else '',
        'history': history.group(1).strip() if history else '',
        'macao_tip': macao_tip.group(1).strip() if macao_tip else '',
        'initial_odds': initial_odds,
        'realtime_odds': realtime_odds,
    }

def analyze_match(data):
    """分析单场比赛"""
    if not data['initial_odds'] or not data['realtime_odds']:
        return None
    
    init_home = [o[0] for o in data['initial_odds']]
    init_draw = [o[1] for o in data['initial_odds']]
    init_away = [o[2] for o in data['initial_odds']]
    
    real_home = [o[0] for o in data['realtime_odds']]
    real_draw = [o[1] for o in data['realtime_odds']]
    real_away = [o[2] for o in data['realtime_odds']]
    
    # 计算变化百分比
    home_pct = [(data['realtime_odds'][i][0] - data['initial_odds'][i][0]) / data['initial_odds'][i][0] * 100 
                for i in range(len(data['initial_odds']))]
    draw_pct = [(data['realtime_odds'][i][1] - data['initial_odds'][i][1]) / data['initial_odds'][i][1] * 100 
                for i in range(len(data['initial_odds']))]
    away_pct = [(data['realtime_odds'][i][2] - data['initial_odds'][i][2]) / data['initial_odds'][i][2] * 100 
                for i in range(len(data['initial_odds']))]
    
    # 概率
    real_home_prob = [1/x*100 for x in real_home]
    real_draw_prob = [1/x*100 for x in real_draw]
    real_away_prob = [1/x*100 for x in real_away]
    
    # 统计
    home_up_pct = sum(1 for x in home_pct if x > 0) / len(home_pct) * 100
    draw_down_pct = sum(1 for x in draw_pct if x < 0) / len(draw_pct) * 100
    away_down_pct = sum(1 for x in away_pct if x < 0) / len(away_pct) * 100
    
    # 实盘判断
    is_real = home_up_pct > 60 and draw_down_pct > 40
    panxing = "实盘" if is_real else "诱盘"
    
    # 平局可能
    draw_prob = np.mean(real_draw_prob)
    draw_possible = "中高" if draw_down_pct > 40 and draw_prob > 25 else "中低"
    
    # 首选判断
    real_home_avg = np.mean(real_home)
    real_draw_avg = np.mean(real_draw)
    real_away_avg = np.mean(real_away)
    
    # 澳门(假设第3家)
    macao_idx = 2
    if len(data['realtime_odds']) > macao_idx:
        macao = data['realtime_odds'][macao_idx]
        macao_home = macao[0]
        macao_draw = macao[1]
        macao_away = macao[2]
    else:
        macao_home = macao_draw = macao_away = 0
    
    # 判断逻辑
    if macao_draw < real_draw_avg and macao_draw < macao_away:
        first_choice = "平局"
        first_prob = f"{draw_prob:.0f}%"
    elif macao_away < real_away_avg:
        first_choice = f"{data['away_team']}客胜"
        first_prob = f"{np.mean(real_away_prob):.0f}%"
    else:
        if real_home_avg < 1.5:
            first_choice = f"{data['home_team']}主胜"
            first_prob = f"{np.mean(real_home_prob):.0f}%"
        else:
            first_choice = f"{data['away_team']}客胜"
            first_prob = f"{np.mean(real_away_prob):.0f}%"
    
    return {
        'home_team': data['home_team'],
        'away_team': data['away_team'],
        'league': data['league'],
        'macao_tip': data['macao_tip'],
        'home_form': data['home_form'],
        'away_form': data['away_form'],
        'panxing': panxing,
        'draw_possible': draw_possible,
        'draw_prob': f"{draw_prob:.1f}%",
        'draw_change': f"{np.mean(draw_pct):.1f}%",
        'draw_down_pct': draw_down_pct,
        'first_choice': first_choice,
        'first_prob': first_prob,
        'home_pct_avg': f"{np.mean(home_pct):.1f}%",
        'away_pct_avg': f"{np.mean(away_pct):.1f}%",
        'real_home_prob': f"{np.mean(real_home_prob):.1f}%",
        'real_draw_prob': f"{np.mean(real_draw_prob):.1f}%",
        'real_away_prob': f"{np.mean(real_away_prob):.1f}%",
        'companies': len(data['initial_odds']),
    }

# 处理所有文件
folder = "分析模板/3.15"
files = [f for f in os.listdir(folder) if f.endswith('_源数据.md')]

results = []
for f in sorted(files):
    print(f"正在分析: {f}")
    filepath = os.path.join(folder, f)
    try:
        data = extract_odds_from_file(filepath)
        result = analyze_match(data)
        if result:
            result['filename'] = f.replace('_源数据.md', '')
            results.append(result)
            print(f"  -> {result['first_choice']} ({result['first_prob']})")
    except Exception as e:
        print(f"  -> 错误: {e}")

# 生成汇总Excel
wb = Workbook()
ws = wb.active
ws.title = "比赛预测汇总"

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
center = Alignment(horizontal='center', vertical='center')

headers = ['编号', '赛事', '对阵', '澳门推荐', '盘型', '平局可能', '首选', '主胜概率', '平局概率', '客胜概率', '主胜变化', '平局变化', '客胜变化']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

for r, data in enumerate(results, 2):
    ws.cell(row=r, column=1, value=data['filename'])
    ws.cell(row=r, column=2, value=data['league'])
    ws.cell(row=r, column=3, value=f"{data['home_team']} vs {data['away_team']}")
    ws.cell(row=r, column=4, value=data['macao_tip'])
    ws.cell(row=r, column=5, value=data['panxing'])
    ws.cell(row=r, column=6, value=data['draw_possible'])
    ws.cell(row=r, column=7, value=data['first_choice'])
    ws.cell(row=r, column=8, value=data['real_home_prob'])
    ws.cell(row=r, column=9, value=data['real_draw_prob'])
    ws.cell(row=r, column=10, value=data['real_away_prob'])
    ws.cell(row=r, column=11, value=data['home_pct_avg'])
    ws.cell(row=r, column=12, value=data['draw_change'])
    ws.cell(row=r, column=13, value=data['away_pct_avg'])
    
    for c in range(1, 14):
        ws.cell(row=r, column=c).alignment = center

for col in range(1, 14):
    ws.column_dimensions[chr(64+col)].width = 14

wb.save('3.15_比赛预测汇总.xlsx')
print(f"\n已生成汇总文件: 3.15_比赛预测汇总.xlsx")
print(f"共分析 {len(results)} 场比赛")
