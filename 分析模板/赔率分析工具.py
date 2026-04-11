import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def analyze_match(home_team, away_team, match_time, league, home_form, away_form, 
                   home_handicap, away_handicap, history, macao_tip,
                   initial_odds, realtime_odds, companies):
    """
    赔率分析函数
    
    参数:
    - home_team: 主队名称
    - away_team: 客队名称
    - match_time: 比赛时间
    - league: 赛事
    - home_form: 主队近况(如WLLLDL)
    - away_form: 客队近况
    - home_handicap: 主队盘路
    - away_handicap: 客队盘路
    - history: 历史交锋(如"2胜4和4负")
    - macao_tip: 澳门推荐
    - initial_odds: 初盘赔率列表 [(主,平,负), ...]
    - realtime_odds: 即时赔率列表
    - companies: 公司名称列表
    """
    
    # 提取数据
    init_home = [o[0] for o in initial_odds]
    init_draw = [o[1] for o in initial_odds]
    init_away = [o[2] for o in initial_odds]
    
    real_home = [o[0] for o in realtime_odds]
    real_draw = [o[1] for o in realtime_odds]
    real_away = [o[2] for o in realtime_odds]
    
    # 计算变化
    home_change = [real_home[i] - init_home[i] for i in range(len(initial_odds))]
    draw_change = [real_draw[i] - init_draw[i] for i in range(len(initial_odds))]
    away_change = [real_away[i] - init_away[i] for i in range(len(initial_odds))]
    
    # 计算百分比变化
    home_pct_change = [(realtime_odds[i][0] - initial_odds[i][0]) / initial_odds[i][0] * 100 for i in range(len(initial_odds))]
    draw_pct_change = [(realtime_odds[i][1] - initial_odds[i][1]) / initial_odds[i][1] * 100 for i in range(len(initial_odds))]
    away_pct_change = [(realtime_odds[i][2] - initial_odds[i][2]) / initial_odds[i][2] * 100 for i in range(len(initial_odds))]
    
    # 概率计算
    init_home_prob = [1/x*100 for x in init_home]
    init_draw_prob = [1/x*100 for x in init_draw]
    init_away_prob = [1/x*100 for x in init_away]
    
    real_home_prob = [1/x*100 for x in real_home]
    real_draw_prob = [1/x*100 for x in real_draw]
    real_away_prob = [1/x*100 for x in real_away]
    
    # 统计函数
    def fmt(val): return f"{val:.2f}"
    def pct(val): return f"{val:.2f}%"
    
    results = {
        'home_team': home_team,
        'away_team': away_team,
        'match_time': match_time,
        'league': league,
        'home_form': home_form,
        'away_form': away_form,
        'home_handicap': home_handicap,
        'away_handicap': away_handicap,
        'history': history,
        'macao_tip': macao_tip,
        
        # 初盘统计
        'init_home_avg': fmt(np.mean(init_home)),
        'init_draw_avg': fmt(np.mean(init_draw)),
        'init_away_avg': fmt(np.mean(init_away)),
        'init_home_mid': fmt(np.median(init_home)),
        'init_draw_mid': fmt(np.median(init_draw)),
        'init_away_mid': fmt(np.median(init_away)),
        'init_home_min': fmt(min(init_home)),
        'init_draw_min': fmt(min(init_draw)),
        'init_away_min': fmt(min(init_away)),
        'init_home_max': fmt(max(init_home)),
        'init_draw_max': fmt(max(init_draw)),
        'init_away_max': fmt(max(init_away)),
        
        # 即时统计
        'real_home_avg': fmt(np.mean(real_home)),
        'real_draw_avg': fmt(np.mean(real_draw)),
        'real_away_avg': fmt(np.mean(real_away)),
        'real_home_mid': fmt(np.median(real_home)),
        'real_draw_mid': fmt(np.median(real_draw)),
        'real_away_mid': fmt(np.median(real_away)),
        'real_home_min': fmt(min(real_home)),
        'real_draw_min': fmt(min(real_draw)),
        'real_away_min': fmt(min(real_away)),
        'real_home_max': fmt(max(real_home)),
        'real_draw_max': fmt(max(real_draw)),
        'real_away_max': fmt(max(real_away)),
        
        # 变化统计
        'home_change_avg': fmt(np.mean(home_change)),
        'draw_change_avg': fmt(np.mean(draw_change)),
        'away_change_avg': fmt(np.mean(away_change)),
        'home_change_mid': fmt(np.median(home_change)),
        'draw_change_mid': fmt(np.median(draw_change)),
        'away_change_mid': fmt(np.median(away_change)),
        'home_up': sum(1 for x in home_change if x > 0),
        'home_down': sum(1 for x in home_change if x < 0),
        'home_same': sum(1 for x in home_change if x == 0),
        'draw_up': sum(1 for x in draw_change if x > 0),
        'draw_down': sum(1 for x in draw_change if x < 0),
        'draw_same': sum(1 for x in draw_change if x == 0),
        'away_up': sum(1 for x in away_change if x > 0),
        'away_down': sum(1 for x in away_change if x < 0),
        'away_same': sum(1 for x in away_change if x == 0),
        
        # 百分比变化
        'home_pct_avg': pct(np.mean(home_pct_change)),
        'draw_pct_avg': pct(np.mean(draw_pct_change)),
        'away_pct_avg': pct(np.mean(away_pct_change)),
        'home_pct_up': pct(sum(1 for x in home_pct_change if x > 0)/len(home_pct_change)*100),
        'home_pct_down': pct(sum(1 for x in home_pct_change if x < 0)/len(home_pct_change)*100),
        'draw_pct_up': pct(sum(1 for x in draw_pct_change if x > 0)/len(draw_pct_change)*100),
        'draw_pct_down': pct(sum(1 for x in draw_pct_change if x < 0)/len(draw_pct_change)*100),
        'away_pct_up': pct(sum(1 for x in away_pct_change if x > 0)/len(away_pct_change)*100),
        'away_pct_down': pct(sum(1 for x in away_pct_change if x < 0)/len(away_pct_change)*100),
        
        # 概率分析
        'init_home_prob': fmt(np.mean(init_home_prob)),
        'init_draw_prob': fmt(np.mean(init_draw_prob)),
        'init_away_prob': fmt(np.mean(init_away_prob)),
        'real_home_prob': fmt(np.mean(real_home_prob)),
        'real_draw_prob': fmt(np.mean(real_draw_prob)),
        'real_away_prob': fmt(np.mean(real_away_prob)),
        'home_prob_change': fmt(np.mean(real_home_prob) - np.mean(init_home_prob)),
        'draw_prob_change': fmt(np.mean(real_draw_prob) - np.mean(init_draw_prob)),
        'away_prob_change': fmt(np.mean(real_away_prob) - np.mean(init_away_prob)),
        
        # 返还率
        'init_return': pct(np.mean([(init_home_prob[i]+init_draw_prob[i]+init_away_prob[i])/100 for i in range(len(initial_odds))])*100),
        'real_return': pct(np.mean([(real_home_prob[i]+real_draw_prob[i]+real_away_prob[i])/100 for i in range(len(initial_odds))])*100),
        
        # 平局分析
        'real_draw_prob_val': fmt(np.mean(real_draw_prob)),
        'real_draw_min': fmt(min(real_draw)),
        'real_draw_max': fmt(max(real_draw)),
        'real_draw_mid_val': fmt(np.median(real_draw)),
        
        # 公司数量
        'total_companies': len(initial_odds),
    }
    
    # 澳门分析(假设澳门是第3家)
    macao_idx = 2
    macao_init = initial_odds[macao_idx]
    macao_real = realtime_odds[macao_idx]
    
    results['macao_home'] = macao_init[0]
    results['macao_draw'] = macao_init[1]
    results['macao_away'] = macao_init[2]
    results['macao_home_now'] = macao_real[0]
    results['macao_draw_now'] = macao_real[1]
    results['macao_away_now'] = macao_real[2]
    results['macao_home_chg'] = macao_real[0] - macao_init[0]
    results['macao_draw_chg'] = macao_real[1] - macao_init[1]
    results['macao_away_chg'] = macao_real[2] - macao_init[2]
    
    return results

def generate_excel(results, filename):
    """生成Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "赔率分析"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')
    
    # 标题
    ws['A1'] = f"{results['home_team']} vs {results['away_team']} 赔率分析"
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = center
    
    # 基本信息
    ws['A3'] = '基本信息'
    ws['A3'].font = Font(bold=True)
    info = [
        ['比赛时间', results['match_time']],
        ['赛事', results['league']],
        ['对阵', f"{results['home_team']}(主) vs {results['away_team']}"],
        ['主队近况', results['home_form']],
        ['客队近况', results['away_form']],
        ['澳门推荐', results['macao_tip']],
        ['历史交锋', results['history']],
    ]
    for r, (k, v) in enumerate(info, 4):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    
    # 赔率统计
    ws['A12'] = '一、赔率统计'
    ws['A12'].font = Font(bold=True, size=12)
    
    stats_headers = ['指标', '主胜', '平局', '客胜']
    for c, h in enumerate(stats_headers, 1):
        cell = ws.cell(row=13, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    stats_data = [
        ['初盘平均', results['init_home_avg'], results['init_draw_avg'], results['init_away_avg']],
        ['即时平均', results['real_home_avg'], results['real_draw_avg'], results['real_away_avg']],
        ['初盘中位数', results['init_home_mid'], results['init_draw_mid'], results['init_away_mid']],
        ['即时中位数', results['real_home_mid'], results['real_draw_mid'], results['real_away_mid']],
    ]
    for r, row_data in enumerate(stats_data, 14):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = center
    
    # 变化统计
    ws['A20'] = '二、变化统计'
    ws['A20'].font = Font(bold=True, size=12)
    
    chg_headers = ['指标', '主胜变化', '平局变化', '客胜变化']
    for c, h in enumerate(chg_headers, 1):
        cell = ws.cell(row=21, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    chg_data = [
        ['平均变化', results['home_change_avg'], results['draw_change_avg'], results['away_change_avg']],
        ['变化%', results['home_pct_avg'], results['draw_pct_avg'], results['away_pct_avg']],
        ['上升(家)', results['home_up'], results['draw_up'], results['away_up']],
        ['下降(家)', results['home_down'], results['draw_down'], results['away_down']],
    ]
    for r, row_data in enumerate(chg_data, 22):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = center
    
    # 概率分析
    ws['A28'] = '三、概率分析'
    ws['A28'].font = Font(bold=True, size=12)
    
    prob_headers = ['分析项', '主胜%', '平局%', '客胜%']
    for c, h in enumerate(prob_headers, 1):
        cell = ws.cell(row=29, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    prob_data = [
        ['初盘概率', results['init_home_prob'], results['init_draw_prob'], results['init_away_prob']],
        ['即时概率', results['real_home_prob'], results['real_draw_prob'], results['real_away_prob']],
        ['概率变化', results['home_prob_change'], results['draw_prob_change'], results['away_prob_change']],
    ]
    for r, row_data in enumerate(prob_data, 30):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = center
    
    # 澳门
    ws['A35'] = '四、澳门赔率'
    ws['A35'].font = Font(bold=True, size=12)
    
    macao_headers = ['类型', '初盘', '即时', '变化']
    for c, h in enumerate(macao_headers, 1):
        cell = ws.cell(row=36, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    
    macao_data = [
        ['主胜', results['macao_home'], results['macao_home_now'], results['macao_home_chg']],
        ['平局', results['macao_draw'], results['macao_draw_now'], results['macao_draw_chg']],
        ['客胜', results['macao_away'], results['macao_away_now'], results['macao_away_chg']],
    ]
    for r, row_data in enumerate(macao_data, 37):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val).alignment = center
    
    for col in 'ABCDEFGHI':
        ws.column_dimensions[col].width = 14
    
    wb.save(filename)
    print(f"Excel已生成: {filename}")

def generate_report(results, filename):
    """生成Markdown报告"""
    
    # 判断实盘/诱盘
    home_pct = float(results['home_pct_avg'].replace('%',''))
    draw_pct = float(results['draw_pct_avg'].replace('%',''))
    away_pct = float(results['away_pct_avg'].replace('%',''))
    home_up_pct = float(results['home_pct_up'].replace('%',''))
    draw_down_pct = float(results['draw_pct_down'].replace('%',''))
    away_down_pct = float(results['away_pct_down'].replace('%',''))
    
    # 判断逻辑
    is_real = home_up_pct > 80 and draw_down_pct > 50
    盘型 = "实盘" if is_real else "诱盘"
    
    # 判断平局可能
    draw_prob = float(results['real_draw_prob_val'])
    draw_possible = "中高" if draw_down_pct > 50 and draw_prob > 30 else "中低"
    
    # 首选判断
    if float(results['macao_draw_now']) < float(results['real_draw_avg']):
        first_choice = "平局"
        first_prob = f"{draw_prob:.0f}%"
    elif float(results['macao_away_now']) < float(results['real_away_avg']):
        first_choice = f"{results['away_team']}客胜"
        first_prob = f"{float(results['real_away_prob']):.0f}%"
    else:
        first_choice = results['away_team'] + "客胜"
        first_prob = f"{float(results['real_away_prob']):.0f}%"
    
    second_choice = "平局"
    second_prob = f"{draw_prob:.0f}%"
    
    report = f"""# {results['home_team']} vs {results['away_team']} 赔率分析报告

## 比赛基本信息
- **比赛时间**: {results['match_time']}
- **赛事**: {results['league']}
- **对阵**: {results['home_team']}(主) vs {results['away_team']}

---

## 一、赔率统计指标

### 1. 初盘赔率统计
| 指标 | 主胜 | 平局 | 客胜 |
|------|------|------|------|
| 平均值 | {results['init_home_avg']} | {results['init_draw_avg']} | {results['init_away_avg']} |
| 中位数 | {results['init_home_mid']} | {results['init_draw_mid']} | {results['init_away_mid']} |
| 最小值 | {results['init_home_min']} | {results['init_draw_min']} | {results['init_away_min']} |
| 最大值 | {results['init_home_max']} | {results['init_draw_max']} | {results['init_away_max']} |

### 2. 即时赔率统计
| 指标 | 主胜 | 平局 | 客胜 |
|------|------|------|------|
| 平均值 | {results['real_home_avg']} | {results['real_draw_avg']} | {results['real_away_avg']} |
| 中位数 | {results['real_home_mid']} | {results['real_draw_mid']} | {results['real_away_mid']} |
| 最小值 | {results['real_home_min']} | {results['real_draw_min']} | {results['real_away_min']} |
| 最大值 | {results['real_home_max']} | {results['real_draw_max']} | {results['real_away_max']} |

---

## 二、赔率变化分析

### 1. 变化统计
| 变化指标 | 平均变化 | 变化% | 上升(家) | 下降(家) | 不变(家) |
|----------|----------|-------|----------|----------|----------|
| 主胜变化 | {results['home_change_avg']} | {results['home_pct_avg']} | {results['home_up']} | {results['home_down']} | {results['home_same']} |
| 平局变化 | {results['draw_change_avg']} | {results['draw_pct_avg']} | {results['draw_up']} | {results['draw_down']} | {results['draw_same']} |
| 客胜变化 | {results['away_change_avg']} | {results['away_pct_avg']} | {results['away_up']} | {results['away_down']} | {results['away_same']} |

### 2. 关键发现
- 主胜变化: 平均{results['home_change_avg']}，{results['home_pct_up']}公司上升
- 平局变化: 平均{results['draw_change_avg']}，{results['draw_pct_down']}公司下降
- 客胜变化: 平均{results['away_change_avg']}，{results['away_pct_down']}公司下降

---

## 三、概率与返还率分析

### 1. 概率变化
| 分析项 | 主胜概率% | 平局概率% | 客胜概率% |
|--------|-----------|-----------|-----------|
| 初盘平均值 | {results['init_home_prob']} | {results['init_draw_prob']} | {results['init_away_prob']} |
| 即时平均值 | {results['real_home_prob']} | {results['real_draw_prob']} | {results['real_away_prob']} |
| 变化 | {results['home_prob_change']} | {results['draw_prob_change']} | {results['away_prob_change']} |

---

## 四、澳门心水分析

### 澳门推荐
- **推介**: {results['macao_tip']}
- **近况走势**:
  - {results['home_team']}: {results['home_form']}
  - {results['away_team']}: {results['away_form']}
- **盘路走势**:
  - {results['home_team']}: {results['home_handicap']}
  - {results['away_team']}: {results['away_handicap']}
- **对赛成绩**: {results['history']}

### 澳门赔率变化
| 赔率类型 | 主胜 | 平局 | 客胜 |
|----------|------|------|------|
| 初盘 | {results['macao_home']} | {results['macao_draw']} | {results['macao_away']} |
| 即时 | {results['macao_home_now']} | {results['macao_draw_now']} | {results['macao_away_now']} |
| 变化 | {results['macao_home_chg']} | {results['macao_draw_chg']} | {results['macao_away_chg']} |

---

## 五、水位变动百分比分析

### 1. 变化百分比统计
| 指标 | 主胜 | 平局 | 客胜 |
|------|------|------|------|
| 平均变化% | {results['home_pct_avg']} | {results['draw_pct_avg']} | {results['away_pct_avg']} |
| 上升占比 | {results['home_pct_up']} | {results['draw_pct_up']} | {results['away_pct_up']} |
| 下降占比 | {results['home_pct_down']} | {results['draw_pct_down']} | {results['away_pct_down']} |

---

## 六、平局定律分析

### 1. 平局赔率分析
- 即时平局赔率: {results['real_draw_avg']}(平均) / {results['real_draw_mid_val']}(中位数)
- 区间: {results['real_draw_min']} ~ {results['real_draw_max']}
- 即时平局概率: {results['real_draw_prob_val']}%

### 2. 平局判断
- 变化趋势: {'下降' if float(results['draw_change_avg']) < 0 else '上升'}
- 降赔公司占比: {results['draw_pct_down']}

---

## 七、实盘/诱盘判断

### 判断依据

| 特征 | 数据 | 判断 |
|------|------|------|
| 主胜升幅 | {results['home_pct_avg']}，{results['home_pct_up']}上升 | {'真实不看好主胜' if home_up_pct > 80 else '存在诱盘可能'} |
| 平局变化 | {results['draw_pct_avg']}，{results['draw_pct_down']}下降 | {'真实防范平局' if draw_down_pct > 50 else '防范有限'} |
| 客胜变化 | {results['away_pct_avg']}，{results['away_pct_down']}下降 | {'有保护' if away_down_pct > 30 else '保护有限'} |
| 澳门态度 | {'全场不动' if results['macao_home_chg'] == 0 else '有调整'} | {'保持中立' if results['macao_home_chg'] == 0 else '跟随市场'} |

## 🎯 判定：{盘型}

**数据依据**:
1. 主胜{results['home_pct_avg']}变化，{results['home_pct_up']}公司上升 → {'真实不看好' if home_up_pct > 80 else '可能诱盘'}
2. 平局{results['draw_pct_avg']}变化，{results['draw_pct_down']}公司下降 → {'真实防范' if draw_down_pct > 50 else '防范有限'}
3. 客胜{results['away_pct_avg']}变化，{results['away_pct_down']}公司下降
4. 澳门{'保持不动' if results['macao_home_chg'] == 0 else '有调整'}

---

## 八、综合预测结论

### 1. 概率排序
| 选项 | 即时概率 | 变化 |
|------|----------|------|
| {results['away_team']}(客胜) | {results['real_away_prob']}% | {results['away_prob_change']}% |
| 平局 | {results['real_draw_prob']}% | {results['draw_prob_change']}% |
| {results['home_team']}(主胜) | {results['real_home_prob']}% | {results['home_prob_change']}% |

### 2. 预测结论

**首选: {first_choice}** ({first_prob})

**次选: {second_choice}** ({second_prob})

**理由**:
- 澳门推荐: {results['macao_tip']}
- 平局概率: {draw_prob}%，{'较高' if draw_prob > 32 else '中等'}
- 盘型判断: {盘型}

### 3. 风险提示
- 足球比赛存在不确定性，本分析仅供参考
- 请根据自身判断做出投注决策

---

*分析公司数量: {results['total_companies']}家*
"""
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"报告已生成: {filename}")

# ========== 使用示例 ==========
if __name__ == "__main__":
    
    # 维罗纳 vs 热那亚 数据
    home_team = "维罗纳"
    away_team = "热那亚"
    match_time = "2026-03-15 19:30"
    league = "25/26意甲第29轮"
    home_form = "WLLLDL"
    away_form = "WLWDLL"
    home_handicap = "W1/LLLDL"
    away_handicap = "WLW1/LLL"
    history = "维罗纳 2胜4和4负"
    macao_tip = "和局"
    
    initial_odds = [
        (2.88, 2.80, 2.35), (3.00, 2.75, 2.50), (3.13, 2.91, 2.23), (3.10, 2.87, 2.50),
        (3.20, 3.00, 2.35), (3.15, 3.05, 2.45), (3.15, 3.05, 2.35), (3.15, 3.25, 2.41),
        (3.40, 3.10, 2.38), (3.10, 2.90, 2.45), (3.10, 2.87, 2.55), (3.02, 2.95, 2.57),
        (3.20, 3.00, 2.35), (3.10, 2.90, 2.50), (2.99, 2.92, 2.35), (3.15, 2.90, 2.38),
        (3.10, 2.95, 2.50), (3.15, 3.05, 2.32), (3.15, 3.05, 2.32), (3.15, 3.25, 2.41),
        (3.15, 2.80, 2.18), (3.15, 3.05, 2.32), (3.25, 2.98, 2.44), (2.80, 3.05, 2.60),
        (3.30, 3.09, 2.58), (3.15, 3.05, 2.35), (3.00, 3.10, 2.35), (3.10, 2.80, 2.30),
        (3.15, 3.05, 2.35), (3.15, 2.98, 2.52), (2.88, 3.00, 2.50), (3.10, 3.00, 2.30),
        (3.37, 3.00, 2.50), (3.20, 3.05, 2.41), (3.25, 3.15, 2.40), (3.15, 3.10, 2.45),
        (3.20, 2.90, 2.30), (3.18, 3.05, 2.52), (3.35, 3.10, 2.33), (3.15, 3.00, 2.40),
        (3.00, 2.88, 2.50), (2.85, 2.70, 2.40), (3.20, 3.10, 2.35), (2.80, 2.80, 2.70),
        (3.15, 3.00, 2.38), (3.20, 3.20, 2.30), (3.30, 3.04, 2.42), (3.20, 3.00, 2.35),
        (3.26, 2.98, 2.35), (3.15, 2.95, 2.45), (3.10, 2.70, 2.50)
    ]
    
    realtime_odds = [
        (3.30, 2.69, 2.20), (3.30, 2.80, 2.30), (3.13, 2.91, 2.23), (3.30, 2.87, 2.40),
        (3.30, 3.00, 2.37), (3.35, 2.95, 2.45), (3.40, 2.90, 2.45), (3.60, 2.95, 2.38),
        (3.50, 3.00, 2.40), (3.50, 2.80, 2.38), (3.30, 2.87, 2.40), (3.56, 3.00, 2.41),
        (3.45, 2.95, 2.38), (3.40, 2.87, 2.45), (3.36, 2.90, 2.44), (3.35, 2.90, 2.40),
        (3.30, 2.87, 2.40), (3.20, 2.94, 2.51), (3.35, 2.90, 2.47), (3.50, 2.95, 2.40),
        (3.20, 2.77, 2.18), (3.35, 2.90, 2.47), (3.60, 3.05, 2.48), (3.40, 2.95, 2.40),
        (3.38, 3.08, 2.44), (3.40, 2.95, 2.40), (3.05, 3.10, 2.35), (3.40, 2.90, 2.40),
        (3.35, 2.95, 2.40), (3.25, 3.05, 2.42), (3.40, 2.90, 2.38), (3.30, 2.88, 2.30),
        (3.47, 2.95, 2.48), (3.40, 2.95, 2.46), (3.40, 2.95, 2.40), (3.15, 3.05, 2.45),
        (3.40, 2.90, 2.40), (3.44, 2.99, 2.44), (3.30, 2.90, 2.48), (3.30, 2.90, 2.40),
        (3.30, 2.80, 2.38), (3.35, 2.85, 2.35), (3.40, 2.90, 2.40), (3.40, 2.90, 2.30),
        (3.30, 2.95, 2.43), (3.50, 2.95, 2.30), (3.42, 3.01, 2.48), (3.45, 3.00, 2.46),
        (3.46, 2.91, 2.36), (3.45, 2.95, 2.35), (3.30, 2.80, 2.30)
    ]
    
    companies = [f"公司{i}" for i in range(1, 52)]
    
    # 分析并生成报告
    results = analyze_match(home_team, away_team, match_time, league, 
                           home_form, away_form, home_handicap, away_handicap,
                           history, macao_tip, initial_odds, realtime_odds, companies)
    
    generate_excel(results, f"{home_team}vs{away_team}_分析.xlsx")
    generate_report(results, f"{home_team}vs{away_team}_分析报告.md")
